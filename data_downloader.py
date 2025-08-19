import os
import json
import zipfile
import io
import pandas as pd
from openpyxl import load_workbook
import streamlit as st
from datetime import datetime

class DataDownloader:
    """Utility class for downloading Excel files and document data"""
    
    def __init__(self):
        self.output_dir = "code/output"
        self.docs_dir = "docsss"  # Assuming this is where documents are stored
    
    def get_excel_files_info(self):
        """Get information about all available Excel files"""
        excel_files = []
        
        if os.path.exists(self.output_dir):
            for filename in os.listdir(self.output_dir):
                if filename.endswith('.xlsx'):
                    file_path = os.path.join(self.output_dir, filename)
                    file_size = os.path.getsize(file_path)
                    
                    # Get file info
                    file_info = {
                        'filename': filename,
                        'filepath': file_path,
                        'size_bytes': file_size,
                        'size_mb': round(file_size / (1024 * 1024), 2),
                        'last_modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S')
                    }
                    
                    # Try to get sheet names and basic info
                    try:
                        wb = load_workbook(file_path, read_only=True)
                        file_info['sheet_names'] = wb.sheetnames
                        file_info['sheet_count'] = len(wb.sheetnames)
                        wb.close()
                    except Exception as e:
                        file_info['sheet_names'] = []
                        file_info['sheet_count'] = 0
                        file_info['error'] = str(e)
                    
                    excel_files.append(file_info)
        
        return excel_files
    
    def get_document_files_info(self):
        """Get information about document files"""
        doc_files = []
        
        if os.path.exists(self.docs_dir):
            for filename in os.listdir(self.docs_dir):
                if filename.endswith(('.docx', '.pdf', '.txt')):
                    file_path = os.path.join(self.docs_dir, filename)
                    file_size = os.path.getsize(file_path)
                    
                    file_info = {
                        'filename': filename,
                        'filepath': file_path,
                        'size_bytes': file_size,
                        'size_mb': round(file_size / (1024 * 1024), 2),
                        'last_modified': datetime.fromtimestamp(os.path.getmtime(file_path)).strftime('%Y-%m-%d %H:%M:%S'),
                        'file_type': filename.split('.')[-1].upper()
                    }
                    
                    doc_files.append(file_info)
        
        return doc_files
    
    def create_excel_summary_report(self, excel_files):
        """Create a summary report of Excel files"""
        if not excel_files:
            return "No Excel files found."
        
        report = f"Excel Files Summary Report\n"
        report += f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        report += f"Total files: {len(excel_files)}\n\n"
        
        for i, file_info in enumerate(excel_files, 1):
            report += f"{i}. {file_info['filename']}\n"
            report += f"   Size: {file_info['size_mb']} MB\n"
            report += f"   Sheets: {file_info['sheet_count']}\n"
            report += f"   Modified: {file_info['last_modified']}\n"
            
            if file_info.get('sheet_names'):
                report += f"   Sheet names: {', '.join(file_info['sheet_names'])}\n"
            
            report += "\n"
        
        return report
    
    def create_document_summary_report(self, doc_files):
        """Create a summary report of document files"""
        if not doc_files:
            return "No document files found."
        
        report = f"Document Files Summary Report\n"
        report += f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        report += f"Total files: {len(doc_files)}\n\n"
        
        for i, file_info in enumerate(doc_files, 1):
            report += f"{i}. {file_info['filename']}\n"
            report += f"   Type: {file_info['file_type']}\n"
            report += f"   Size: {file_info['size_mb']} MB\n"
            report += f"   Modified: {file_info['last_modified']}\n\n"
        
        return report
    
    def download_single_excel_file(self, file_path, filename):
        """Download a single Excel file"""
        try:
            with open(file_path, 'rb') as f:
                file_data = f.read()
            
            return file_data
        except Exception as e:
            st.error(f"Error reading file {filename}: {str(e)}")
            return None
    
    def download_all_excel_files(self):
        """Download all Excel files as a ZIP"""
        try:
            excel_files = self.get_excel_files_info()
            
            if not excel_files:
                return None
            
            # Create ZIP buffer
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add each Excel file
                for file_info in excel_files:
                    zip_file.write(file_info['filepath'], file_info['filename'])
                
                # Add summary report
                summary_report = self.create_excel_summary_report(excel_files)
                zip_file.writestr("Excel_Files_Summary.txt", summary_report)
            
            zip_buffer.seek(0)
            return zip_buffer
            
        except Exception as e:
            st.error(f"Error creating Excel ZIP: {str(e)}")
            return None
    
    def download_all_document_files(self):
        """Download all document files as a ZIP"""
        try:
            doc_files = self.get_document_files_info()
            
            if not doc_files:
                return None
            
            # Create ZIP buffer
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add each document file
                for file_info in doc_files:
                    zip_file.write(file_info['filepath'], file_info['filename'])
                
                # Add summary report
                summary_report = self.create_document_summary_report(doc_files)
                zip_file.writestr("Document_Files_Summary.txt", summary_report)
            
            zip_buffer.seek(0)
            return zip_buffer
            
        except Exception as e:
            st.error(f"Error creating document ZIP: {str(e)}")
            return None
    
    def download_complete_data_package(self, session_state=None):
        """Download complete data package including Excel, documents, and session data"""
        try:
            # Create ZIP buffer
            zip_buffer = io.BytesIO()
            
            with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
                # Add Excel files
                excel_files = self.get_excel_files_info()
                if excel_files:
                    for file_info in excel_files:
                        zip_file.write(file_info['filepath'], f"Excel_Files/{file_info['filename']}")
                    
                    # Add Excel summary
                    excel_summary = self.create_excel_summary_report(excel_files)
                    zip_file.writestr("Excel_Files/Excel_Summary.txt", excel_summary)
                
                # Add document files
                doc_files = self.get_document_files_info()
                if doc_files:
                    for file_info in doc_files:
                        zip_file.write(file_info['filepath'], f"Document_Files/{file_info['filename']}")
                    
                    # Add document summary
                    doc_summary = self.create_document_summary_report(doc_files)
                    zip_file.writestr("Document_Files/Document_Summary.txt", doc_summary)
                
                # Add session state data if available
                if session_state:
                    # Add comprehensive data
                    if session_state.get('comprehensive_data'):
                        comp_data = session_state.comprehensive_data
                        zip_file.writestr("Session_Data/Project_Data.json", 
                                        json.dumps(comp_data, indent=2, default=str))
                    
                    # Add TRL results
                    if session_state.get('trl_result'):
                        trl_data = session_state.trl_result
                        zip_file.writestr("Session_Data/TRL_Analysis.json", 
                                        json.dumps(trl_data, indent=2, default=str))
                    
                    # Add Frascati results
                    if session_state.get('frascati_result'):
                        frascati_data = session_state.frascati_result
                        zip_file.writestr("Session_Data/Frascati_Analysis.json", 
                                        json.dumps(frascati_data, indent=2, default=str))
                    
                    # Add market results
                    if session_state.get('market_result'):
                        market_data = session_state.market_result
                        zip_file.writestr("Session_Data/Market_Analysis.json", 
                                        json.dumps(market_data, indent=2, default=str))
                    
                    # Add SME results
                    if session_state.get('sme_result'):
                        sme_data = session_state.sme_result
                        zip_file.writestr("Session_Data/SME_Analysis.json", 
                                        json.dumps(sme_data, indent=2, default=str))
                    
                    # Add document results
                    if session_state.get('document_results'):
                        doc_results = session_state.document_results
                        zip_file.writestr("Session_Data/Document_Results.json", 
                                        json.dumps(doc_results, indent=2, default=str))
                    
                    # Add Excel processing results
                    if session_state.get('excel_processing_result'):
                        excel_results = session_state.excel_processing_result
                        zip_file.writestr("Session_Data/Excel_Processing_Results.json", 
                                        json.dumps(excel_results, indent=2, default=str))
                
                # Add overall package summary
                package_summary = self.create_package_summary(excel_files, doc_files, session_state)
                zip_file.writestr("Package_Summary.txt", package_summary)
            
            zip_buffer.seek(0)
            return zip_buffer
            
        except Exception as e:
            st.error(f"Error creating complete package: {str(e)}")
            return None
    
    def create_package_summary(self, excel_files, doc_files, session_state):
        """Create a summary of the complete package"""
        summary = f"Complete Data Package Summary\n"
        summary += f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n"
        summary += f"=" * 50 + "\n\n"
        
        # Excel files summary
        summary += f"Excel Files: {len(excel_files)}\n"
        if excel_files:
            total_excel_size = sum(f['size_bytes'] for f in excel_files)
            summary += f"Total Excel size: {round(total_excel_size / (1024 * 1024), 2)} MB\n"
            for file_info in excel_files:
                summary += f"  - {file_info['filename']} ({file_info['size_mb']} MB)\n"
        summary += "\n"
        
        # Document files summary
        summary += f"Document Files: {len(doc_files)}\n"
        if doc_files:
            total_doc_size = sum(f['size_bytes'] for f in doc_files)
            summary += f"Total document size: {round(total_doc_size / (1024 * 1024), 2)} MB\n"
            for file_info in doc_files:
                summary += f"  - {file_info['filename']} ({file_info['size_mb']} MB)\n"
        summary += "\n"
        
        # Session state summary
        if session_state:
            summary += "Session Data Available:\n"
            if session_state.get('trl_result'):
                summary += "  - TRL Analysis Results\n"
            if session_state.get('frascati_result'):
                summary += "  - Frascati Analysis Results\n"
            if session_state.get('market_result'):
                summary += "  - Market Analysis Results\n"
            if session_state.get('sme_result'):
                summary += "  - SME Eligibility Results\n"
            if session_state.get('comprehensive_data'):
                summary += "  - Project Data (46 fields)\n"
            if session_state.get('document_results'):
                summary += "  - Document Processing Results\n"
            if session_state.get('excel_processing_result'):
                summary += "  - Excel Processing Results\n"
        
        return summary

def main():
    """Main function for testing the data downloader"""
    st.title("Data Download Utility")
    
    # Initialize downloader
    downloader = DataDownloader()
    
    # Get file information
    excel_files = downloader.get_excel_files_info()
    doc_files = downloader.get_document_files_info()
    
    # Display file information
    st.header("📊 Available Files")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("Excel Files")
        if excel_files:
            for file_info in excel_files:
                st.write(f"📈 **{file_info['filename']}**")
                st.write(f"   Size: {file_info['size_mb']} MB | Sheets: {file_info['sheet_count']}")
                st.write(f"   Modified: {file_info['last_modified']}")
                st.write("---")
        else:
            st.info("No Excel files found")
    
    with col2:
        st.subheader("Document Files")
        if doc_files:
            for file_info in doc_files:
                st.write(f"📄 **{file_info['filename']}**")
                st.write(f"   Type: {file_info['file_type']} | Size: {file_info['size_mb']} MB")
                st.write(f"   Modified: {file_info['last_modified']}")
                st.write("---")
        else:
            st.info("No document files found")
    
    # Download options
    st.header("💾 Download Options")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        if st.button("📊 Download Excel Files (ZIP)"):
            if excel_files:
                zip_buffer = downloader.download_all_excel_files()
                if zip_buffer:
                    st.download_button(
                        label="💾 Download Excel ZIP",
                        data=zip_buffer.getvalue(),
                        file_name=f"Excel_Files_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                        mime="application/zip"
                    )
                    st.success("✅ Excel files ZIP created!")
            else:
                st.warning("No Excel files available for download")
    
    with col2:
        if st.button("📄 Download Documents (ZIP)"):
            if doc_files:
                zip_buffer = downloader.download_all_document_files()
                if zip_buffer:
                    st.download_button(
                        label="💾 Download Documents ZIP",
                        data=zip_buffer.getvalue(),
                        file_name=f"Document_Files_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                        mime="application/zip"
                    )
                    st.success("✅ Documents ZIP created!")
            else:
                st.warning("No document files available for download")
    
    with col3:
        if st.button("📦 Download Complete Package"):
            # Create sample session state for testing
            sample_session = {
                'trl_result': {'trl_level': '4'},
                'comprehensive_data': {'extracted_count': 46}
            }
            
            zip_buffer = downloader.download_complete_data_package(sample_session)
            if zip_buffer:
                st.download_button(
                    label="💾 Download Complete Package",
                    data=zip_buffer.getvalue(),
                    file_name=f"Complete_Data_Package_{datetime.now().strftime('%Y%m%d_%H%M')}.zip",
                    mime="application/zip"
                )
                st.success("✅ Complete package created!")

if __name__ == "__main__":
    main()
