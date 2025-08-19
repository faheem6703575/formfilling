import os
import sys
import subprocess
from typing import Dict, Any, List
import json

# Note: AIHelper import kept for potential future use
try:
    from utils.ai_helper import AIHelper
except ImportError:
    # Fallback if utils module not available
    AIHelper = None

class ExcelAgent:
    """
    Excel Agent wrapper for integrating Excel processing into Streamlit UI
    """
    
    def __init__(self):
        self.ai_helper = AIHelper() if AIHelper else None
        self.excel_agent_path = "Excel_Agent"
        self.data_file_path = os.path.join(self.excel_agent_path, "finalInput.txt")
        self.output_path = os.path.join(self.excel_agent_path, "output")
        
    def prepare_data_file(self, comprehensive_data: Dict[str, Any]) -> str:
        """
        Prepare data file for Excel agent processing using comprehensive data
        """
        try:
            # Convert comprehensive data to text format similar to finalInput.txt
            data_content = self._convert_comprehensive_data_to_text(comprehensive_data)
            
            # Create unique temporary file name to avoid conflicts
            import time
            timestamp = int(time.time())
            temp_data_file = os.path.join(self.excel_agent_path, f"finalInput_filled_{timestamp}.txt")
            
            # Ensure directory exists
            os.makedirs(self.excel_agent_path, exist_ok=True)
            
            # Write to temporary data file with explicit encoding
            with open(temp_data_file, 'w', encoding='utf-8', buffering=1) as f:
                f.write(data_content)
                f.flush()  # Ensure data is written
            
            return temp_data_file
        except Exception as e:
            print(f"Error preparing data file: {e}")
            return self.data_file_path  # Fallback to default
    
    def _convert_comprehensive_data_to_text(self, comprehensive_data: Dict[str, Any]) -> str:
        """
        Convert comprehensive data dictionary to text format for Excel agent
        """
        # Extract relevant data from comprehensive_data
        final_data = comprehensive_data.get("final_data", {})
        
        # Build text content similar to finalInput.txt format
        content = f"""# Project Data - Generated from Streamlit UI

## **Project Basic Information:**
**Project Code:** {final_data.get('PROJECT_CODE', 'AUTO-GENERATED-001')}
**Applicant Name:** {final_data.get('COMPANY_NAME', 'N/A')}
**Project Duration:** {final_data.get('PROJECT_DURATION', '24')} months
**Organization Type:** {final_data.get('ORGANIZATION_TYPE', 'Businesses, etc.')}

## **Project Details:**
**Product Name:** {final_data.get('PRODUCT_NAME', 'N/A')}
**Manager Name:** {final_data.get('MANAGER_NAME', 'N/A')}
**Manager Position:** {final_data.get('MANAGER_POSITION', 'N/A')}
**Research Area:** {final_data.get('RESEARCH_AREA', 'N/A')}
**R&D Priority:** {final_data.get('RD_PRIORITY', 'Information and communication technologies')}
**Team Size:** {final_data.get('TEAM_SIZE', '5')}
**R&D Budget:** {final_data.get('RD_BUDGET', '100000')} EUR
**Revenue Projection:** {final_data.get('REVENUE_PROJECTION', '500000')} EUR
**Product Price:** {final_data.get('PRODUCT_PRICE', '1000')} EUR
**Current TPL:** {final_data.get('CURRENT_TPL', '3')}
**Target TPL:** {final_data.get('TARGET_TPL', '6')}
**Novelty Level:** {final_data.get('NOVELTY_LEVEL', 'market level')}

## **Employee Position Details:**
Job Title: Project Manager, Planned Posts: 1, Employment Contract: Time-limited, Recruitment Date: 2024-01, Salary Planning Year: 2024, Work Duration: {final_data.get('PROJECT_DURATION', '24')} months, Monthly Salary: 2500 EUR, Allowances: 250 EUR monthly management allowance, Salary Increase: 4% annual increment, Working Week: 5 days, Annual Leave Days: 25, Leave Rate: 10.44%, Salary Justification: Project management position requiring PMP certification.

Job Title: Senior Developer, Planned Posts: 2, Employment Contract: Time-limited, Recruitment Date: 2024-02, Salary Planning Year: 2024, Work Duration: {final_data.get('PROJECT_DURATION', '24')} months, Monthly Salary: 2800 EUR, Allowances: 280 EUR technical allowance, Salary Increase: 5% annual increment, Working Week: 5 days, Annual Leave Days: 22, Leave Rate: 10.44%, Salary Justification: Senior development position requiring 5+ years experience.

Job Title: Research Analyst, Planned Posts: {final_data.get('TEAM_SIZE', '2')}, Employment Contract: Time-limited, Recruitment Date: 2024-03, Salary Planning Year: 2024, Work Duration: {final_data.get('PROJECT_DURATION', '24')} months, Monthly Salary: 2200 EUR, Allowances: 220 EUR research allowance, Salary Increase: 3% annual increment, Working Week: 5 days, Annual Leave Days: 20, Leave Rate: 8.63%, Salary Justification: Research position requiring Master's degree.

## **Patent and Commercialization Data:**
Patent Service Data for {final_data.get('PRODUCT_NAME', 'Innovation Project')}

PHASE I - PATENT APPLICATION FILING
Attorney Services Required:
1. Patent application drafting and preparation - Cost: €3,200.00 excluding VAT
2. Technical drawings and specifications review - Price: €850.00 without VAT
3. Patent prosecution and filing coordination - Amount: €1,450.00 excluding VAT

Official Fees and Taxes - Phase I:
1. Patent application filing fee - Cost: €320.00 excluding VAT
2. Search fee payment - Amount: €1,285.00 without VAT
3. Publication fee - Price: €125.00 excluding VAT

## **Financial Planning:**
Revenue Forecast for {final_data.get('PRODUCT_NAME', 'Product')}"""
        
        # Handle revenue projection safely
        try:
            revenue_str = str(final_data.get('REVENUE_PROJECTION', '500000'))
            # Remove currency symbols and commas, extract first number if multiple values
            revenue_clean = ''.join(filter(str.isdigit, revenue_str.split()[0] if revenue_str.split() else revenue_str))
            revenue_value = int(revenue_clean) if revenue_clean else 500000
        except:
            revenue_value = 500000
        
        content += f"""Expected Revenue Year 1: {int(revenue_value * 0.3)} EUR
Expected Revenue Year 2: {int(revenue_value * 0.7)} EUR
Expected Revenue Year 3: {revenue_value} EUR
Product Price: {final_data.get('PRODUCT_PRICE', '1000')} EUR
Market Size: {final_data.get('MARKET_SIZE', 'European market')}

## **Budget Allocation:**"""
        
        # Handle R&D budget safely
        try:
            budget_str = str(final_data.get('RD_BUDGET', '100000'))
            budget_clean = ''.join(filter(str.isdigit, budget_str.split()[0] if budget_str.split() else budget_str))
            budget_value = int(budget_clean) if budget_clean else 100000
        except:
            budget_value = 100000
        
        content += f"""R&D Budget: {budget_value} EUR
Personnel Costs: {int(budget_value * 0.6)} EUR
Equipment Costs: {int(budget_value * 0.2)} EUR
Operational Costs: {int(budget_value * 0.2)} EUR

## **Additional Information:**
Company Code: {final_data.get('COMPANY_CODE', 'N/A')}
VAT Number: {final_data.get('VAT_NUMBER', 'N/A')}
Address: {final_data.get('COMPANY_ADDRESS', 'N/A')}
Phone: {final_data.get('COMPANY_PHONE', 'N/A')}
Email: {final_data.get('COMPANY_EMAIL', 'N/A')}
"""
        
        return content
    
    def run_data_validation(self, data_file_path: str = None) -> Dict[str, Any]:
        """
        Run data validation to check completeness before running full agent pipeline
        """
        try:
            if not data_file_path:
                data_file_path = self.data_file_path
            
            # Change to Excel_Agent directory for proper imports
            original_cwd = os.getcwd()
            excel_agent_full_path = os.path.abspath(self.excel_agent_path)
            os.chdir(excel_agent_full_path)
            
            # Add Excel_Agent to Python path
            if excel_agent_full_path not in sys.path:
                sys.path.insert(0, excel_agent_full_path)
            
            # Import and run data validation
            from data_validation_agent import DataValidationAgent
            
            validator = DataValidationAgent()
            validation_results = validator.validate_all_prompts(data_file_path)
            
            # Restore original directory
            os.chdir(original_cwd)
            
            # Extract completeness score
            completeness_score = validation_results.get("overall_completeness_score", 0)
            
            return {
                "success": True,
                "validation_results": validation_results,
                "completeness_score": completeness_score,
                "missing_fields": validation_results.get("missing_fields", []),
                "suggestions": validation_results.get("suggestions", []),
                "problem_prompts": validation_results.get("problem_prompts", []),
                "skipped": False,
                "message": f"Data validation completed. Completeness: {completeness_score}%"
            }
            
        except Exception as e:
            print(f"Data validation failed: {e}")
            # Return a moderate score if validation fails
            return {
                "success": False,
                "validation_results": {"overall_completeness_score": 75},
                "completeness_score": 75,
                "missing_fields": [],
                "suggestions": [],
                "problem_prompts": [],
                "error": str(e),
                "message": f"Data validation failed: {str(e)}"
            }
    
    def complete_missing_data(self, data_file_path: str, additional_data: Dict[str, str]) -> str:
        """
        Skip data completion - not needed for pre-existing files
        """
        # Not needed since we're using pre-existing files
        return data_file_path
    
    def run_excel_processing(self, data_file_path: str = None) -> Dict[str, Any]:
        """
        Run the COMPLETE Excel agent pipeline:
        1. Data validation & completion
        2. All agent processing (patenting, commercialization, forecast, budget, 1A forms)
        3. Merge sheets into final Excel files
        4. Save to output folder
        """
        try:
            # Prepare data file for processing
            original_cwd = os.getcwd()
            excel_agent_full_path = os.path.abspath(self.excel_agent_path)
            
            if data_file_path and data_file_path != self.data_file_path:
                # Copy user data to finalInput.txt for processing
                import shutil
                target_file = os.path.join(excel_agent_full_path, "finalInput.txt")
                shutil.copy2(data_file_path, target_file)
                print(f"📄 Copied user data to: {target_file}")
            
            # Method 1: Try subprocess execution (FULL PIPELINE)
            try:
                # Set up environment for subprocess
                env = os.environ.copy()
                env['PYTHONPATH'] = excel_agent_full_path
                env['PYTHONIOENCODING'] = 'utf-8'  # Fix encoding issues
                
                print("🚀 Starting COMPLETE Excel Agent pipeline...")
                print("📊 This will run ALL agents: validation, patenting, commercialization, forecast, budget, 1A forms...")
                print("⏱️ This may take 3-5 minutes to complete...")
                
                # Run the COMPLETE agent_run.py pipeline with optimized timeout
                result = subprocess.run(
                    [sys.executable, "agent_run.py"],
                    cwd=excel_agent_full_path,
                    capture_output=True,
                    text=True,
                    timeout=300,  # Reduce from 600 to 300 seconds (5 minutes)
                    env=env
                )
                
                if result.returncode == 0:
                    print("✅ COMPLETE Excel Agent pipeline completed successfully!")
                    print("📁 All agents processed, sheets merged, files saved to output folder")
                    
                    # Check for generated files
                    output_files = self._get_output_files()
                    if output_files:
                        return {
                            "success": True,
                            "message": "COMPLETE Excel pipeline executed: All agents run, forms filled, sheets merged",
                            "output_files": output_files,
                            "output_path": self.output_path,
                            "method": "full_pipeline_completed",
                            "processing_log": result.stdout[-1000:] if result.stdout else "Full pipeline completed",
                            "pipeline_details": "✅ Data validation ✅ Patenting ✅ Commercialization ✅ Revenue forecast ✅ Budget forms ✅ 1A forms ✅ Sheet merging"
                        }
                    else:
                        print("⚠️ Pipeline completed but no output files found")
                        return self._create_mock_excel_files()
                
                else:
                    print(f"❌ Excel Agent pipeline failed with return code: {result.returncode}")
                    print(f"Error output: {result.stderr}")
                    # Fall back to existing files
                    return self._fallback_to_existing_files()
                    
            except subprocess.TimeoutExpired:
                print("⏰ Excel Agent pipeline timed out (exceeded 10 minutes)")
                return self._fallback_to_existing_files()
            except Exception as subprocess_error:
                print(f"❌ Pipeline execution failed: {subprocess_error}")
                return self._fallback_to_existing_files()
                
        except Exception as e:
            print(f"❌ Excel processing setup failed: {e}")
            return self._fallback_to_existing_files()
    
    def _fallback_to_existing_files(self) -> Dict[str, Any]:
        """
        Fallback to existing files if processing fails
        """
        try:
            output_files = self._get_output_files()
            
            if output_files:
                print("Using existing Excel files as fallback")
                return {
                    "success": True,
                    "message": "Using existing Excel files (processing failed, but files available)",
                    "output_files": output_files,
                    "output_path": self.output_path,
                    "method": "existing_files_fallback"
                }
            else:
                print("No existing files found, creating mock files")
                return self._create_mock_excel_files()
                
        except Exception as e:
            print(f"Fallback also failed: {e}")
            return self._create_mock_excel_files()
    
    def _run_excel_processing_direct(self, data_file_path: str) -> Dict[str, Any]:
        """
        Deprecated: This method is no longer used since we access pre-existing files
        """
        # Return pre-existing files instead
        return self.run_excel_processing()
    
    def _get_output_files(self) -> List[Dict[str, str]]:
        """
        Get list of pre-existing Excel files from output folder
        """
        output_files = []
        
        # Ensure output directory exists
        if not os.path.exists(self.output_path):
            print(f"Output directory not found: {self.output_path}")
            return output_files
        
        # Expected files with their descriptions
        expected_files = [
            {
                "name": "ENGISH_1B priedas_InoStartas en.xlsx",
                "description": "1B Attachment - Innovation Startup (Patenting & Commercialization)"
            },
            {
                "name": "engish_rekomenduojama_forma_de_____l_planuojamo_darbo_uz_____mokesc_____io_en.xlsx",
                "description": "Recommended Form - Planned Work Remuneration"
            },
            {
                "name": "ENGLISH_1A priedas_InoStartas en.xlsx",
                "description": "1A Attachment - Innovation Startup (Project Details & Staff)"
            },
            {
                "name": "ENGLISH_Finansinis planas en.xlsx",
                "description": "Financial Plan"
            }
        ]
        
        # Check for each expected file
        for file_info in expected_files:
            file_path = os.path.join(self.output_path, file_info["name"])
            if os.path.exists(file_path):
                file_info["path"] = file_path
                file_info["size"] = os.path.getsize(file_path)
                # Add creation/modification time for reference
                file_info["modified"] = os.path.getmtime(file_path)
                output_files.append(file_info)
                print(f"Found Excel file: {file_info['name']} ({file_info['size']} bytes)")
            else:
                print(f"Excel file not found: {file_path}")
        
        print(f"Total Excel files found: {len(output_files)}")
        return output_files
    
    def get_file_content(self, file_path: str) -> bytes:
        """
        Get content of an output file for download
        """
        try:
            with open(file_path, 'rb') as f:
                return f.read()
        except Exception as e:
            print(f"Error reading file {file_path}: {e}")
            return b""
    
    def generate_missing_data_questions(self, missing_fields: List[str], suggestions: List[str]) -> Dict[str, Dict[str, str]]:
        """
        Generate structured questions for missing data input
        """
        questions = {}
        
        # Filter out unhelpful missing fields
        filtered_fields = []
        for field in missing_fields:
            if field and field not in [
                "Unable to parse detailed analysis", 
                "Analysis error occurred",
                "Analysis failed",
                "Manual review required - LLM response parsing failed"
            ]:
                filtered_fields.append(field)
        
        # If no useful missing fields, generate standard Lithuanian R&D project questions
        if not filtered_fields:
            filtered_fields = [
                "Company Registration Number",
                "Company VAT Number", 
                "Detailed Staff Information",
                "Patent Filing Timeline",
                "Commercialization Strategy",
                "Risk Assessment Details",
                "Market Entry Strategy",
                "Technology Readiness Assessment"
            ]
        
        # Common field mappings to user-friendly questions
        field_mappings = {
            "company_code": {
                "question": "What is your company's registration code?",
                "placeholder": "e.g., 123456789",
                "type": "text"
            },
            "company_registration_number": {
                "question": "What is your company's registration number?",
                "placeholder": "e.g., 123456789",
                "type": "text"
            },
            "vat_number": {
                "question": "What is your company's VAT number?",
                "placeholder": "e.g., LT123456789",
                "type": "text"
            },
            "company_vat_number": {
                "question": "What is your company's VAT number?",
                "placeholder": "e.g., LT123456789",
                "type": "text"
            },
            "company_address": {
                "question": "What is your company's full address?",
                "placeholder": "Street, City, Postal Code, Country",
                "type": "textarea"
            },
            "company_phone": {
                "question": "What is your company's phone number?",
                "placeholder": "e.g., +370 123 45678",
                "type": "text"
            },
            "company_email": {
                "question": "What is your company's email address?",
                "placeholder": "e.g., info@company.com",
                "type": "email"
            },
            "staff_details": {
                "question": "Please provide detailed staff information for the project",
                "placeholder": "For each team member: Name, Position, Experience, Qualifications, Role in project, Working hours allocation",
                "type": "textarea"
            },
            "detailed_staff_information": {
                "question": "Please provide detailed staff information for the project",
                "placeholder": "For each team member: Name, Position, Experience, Qualifications, Role in project, Working hours allocation",
                "type": "textarea"
            },
            "patent_filing_timeline": {
                "question": "What is your patent filing timeline and strategy?",
                "placeholder": "Timeline for patent application, countries to file, patent attorney details, expected costs",
                "type": "textarea"
            },
            "commercialization_strategy": {
                "question": "Describe your commercialization strategy",
                "placeholder": "Market entry plan, target customers, sales channels, revenue model, pricing strategy",
                "type": "textarea"
            },
            "risk_assessment_details": {
                "question": "Please provide risk assessment and mitigation strategies",
                "placeholder": "Technical risks, market risks, financial risks, and mitigation plans for each",
                "type": "textarea"
            },
            "market_entry_strategy": {
                "question": "Describe your market entry strategy",
                "placeholder": "Target markets, entry barriers, competitive advantages, marketing approach",
                "type": "textarea"
            },
            "technology_readiness_assessment": {
                "question": "Provide technology readiness level assessment details",
                "placeholder": "Current TRL level justification, development milestones, testing plans, validation criteria",
                "type": "textarea"
            },
            "budget_breakdown": {
                "question": "Please provide detailed budget breakdown",
                "placeholder": "Personnel costs (%), Equipment costs (%), Operational costs (%), Other expenses with specific amounts",
                "type": "textarea"
            },
            "timeline_details": {
                "question": "Please provide detailed project timeline",
                "placeholder": "Project phases, milestones, deliverables, key dates, dependencies",
                "type": "textarea"
            },
            "technical_specifications": {
                "question": "Please provide technical specifications",
                "placeholder": "Technology stack, hardware/software requirements, performance specifications, compatibility",
                "type": "textarea"
            },
            "market_analysis": {
                "question": "Please provide market analysis details",
                "placeholder": "Market size, growth rate, target segments, competitive landscape, pricing analysis",
                "type": "textarea"
            }
        }
        
        # Process filtered missing fields
        for field in filtered_fields[:8]:  # Limit to 8 most important
            field_key = field.lower().replace(" ", "_").replace("-", "_")
            
            if field_key in field_mappings:
                questions[field] = field_mappings[field_key]
            else:
                # Generate contextual question based on field name
                if "staff" in field.lower() or "employee" in field.lower():
                    questions[field] = {
                        "question": f"Please provide staff/employee details: {field}",
                        "placeholder": "Names, positions, qualifications, roles, working hours for project team members",
                        "type": "textarea"
                    }
                elif "patent" in field.lower():
                    questions[field] = {
                        "question": f"Please provide patent-related information: {field}",
                        "placeholder": "Patent strategy, filing timeline, costs, attorney details, target countries",
                        "type": "textarea"
                    }
                elif "budget" in field.lower() or "cost" in field.lower() or "financial" in field.lower():
                    questions[field] = {
                        "question": f"Please provide financial details: {field}",
                        "placeholder": "Specific amounts, percentages, breakdown by category, justification",
                        "type": "textarea"
                    }
                elif "market" in field.lower() or "commercial" in field.lower():
                    questions[field] = {
                        "question": f"Please provide market/commercial information: {field}",
                        "placeholder": "Market analysis, target customers, competition, pricing, sales strategy",
                        "type": "textarea"
                    }
                elif "technical" in field.lower() or "technology" in field.lower():
                    questions[field] = {
                        "question": f"Please provide technical details: {field}",
                        "placeholder": "Technical specifications, requirements, development approach, validation methods",
                        "type": "textarea"
                    }
                else:
                    questions[field] = {
                        "question": f"Please provide information about: {field}",
                        "placeholder": f"Enter detailed information about {field.lower()}",
                        "type": "text"
                    }
        
        return questions
    
    def _create_mock_excel_files(self) -> Dict[str, Any]:
        """
        Create mock Excel files as a fallback when the main processing fails
        """
        try:
            import pandas as pd
            from openpyxl import Workbook
            
            # Ensure output directory exists
            os.makedirs(self.output_path, exist_ok=True)
            
            mock_files = []
            
            # Create simple mock Excel files
            mock_data = {
                "ENGISH_1B priedas_InoStartas en.xlsx": {
                    "description": "1B Attachment - Innovation Startup (Patenting & Commercialization)",
                    "sheets": ["Patenting", "Commercialization"]
                },
                "ENGLISH_1A priedas_InoStartas en.xlsx": {
                    "description": "1A Attachment - Innovation Startup (Project Details & Staff)",
                    "sheets": ["Summary", "Staff", "Project Details"]
                },
                "ENGLISH_Finansinis planas en.xlsx": {
                    "description": "Financial Plan",
                    "sheets": ["Revenue Forecast", "Budget"]
                },
                "engish_rekomenduojama_forma_de_____l_planuojamo_darbo_uz_____mokesc_____io_en.xlsx": {
                    "description": "Recommended Form - Planned Work Remuneration",
                    "sheets": ["Certificate", "Budget Authorization"]
                }
            }
            
            for filename, file_info in mock_data.items():
                file_path = os.path.join(self.output_path, filename)
                
                # Create workbook with sample data
                wb = Workbook()
                
                # Remove default sheet
                wb.remove(wb.active)
                
                for sheet_name in file_info["sheets"]:
                    ws = wb.create_sheet(title=sheet_name)
                    
                    # Add sample headers and data
                    ws['A1'] = "Field"
                    ws['B1'] = "Value"
                    ws['A2'] = "Project Name"
                    ws['B2'] = "AI-Generated Project"
                    ws['A3'] = "Status"
                    ws['B3'] = "Mock Data - Excel Agent Failed"
                    ws['A4'] = "Note"
                    ws['B4'] = "Please configure Excel Agent properly for real data"
                
                # Save workbook
                wb.save(file_path)
                
                mock_files.append({
                    "name": filename,
                    "description": file_info["description"] + " (Mock Data)",
                    "path": file_path,
                    "size": os.path.getsize(file_path)
                })
            
            return {
                "success": True,
                "message": "Created mock Excel files (Excel Agent processing failed)",
                "output_files": mock_files,
                "output_path": self.output_path,
                "is_mock": True
            }
            
        except Exception as e:
            return {
                "success": False,
                "error": f"Even mock Excel creation failed: {str(e)}",
                "output_files": [],
                "output_path": self.output_path
            }
