from openpyxl import load_workbook, Workbook
from copy import copy

def merge_excels_to_subsheets_one(sheet_paths, output_path="combined.xlsx"):
    """
    Merges the active sheets from the given Excel files into one workbook.
    Each input file becomes a sheet in the output file, preserving structure.

    Parameters:
    - sheet_paths: List of 4 file paths to Excel files
    - output_path: Path to save the new workbook (default: combined.xlsx)
    """
    if len(sheet_paths) != 4:
        raise ValueError("Exactly 4 sheet paths must be provided.")

    # Create a new workbook
    new_wb = Workbook()
    # Remove the default sheet
    default_sheet = new_wb.active
    new_wb.remove(default_sheet)

    # Add sheets from the given files
    for i, path in enumerate(sheet_paths):
        print(f"Processing: {path}")
        src_wb = load_workbook(path)
        src_sheet = src_wb.active

        # Use the same sheet title if available, or make a custom name
        sheet_title = src_sheet.title.strip() or f"Sheet{i+1}"
        new_sheet = new_wb.create_sheet(title=sheet_title)

        # Copy all cells with their values and formatting
        for row in src_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                
                # Copy cell formatting
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # Copy column dimensions
        for col in src_sheet.column_dimensions:
            new_sheet.column_dimensions[col].width = src_sheet.column_dimensions[col].width
        
        # Copy row dimensions
        for row in src_sheet.row_dimensions:
            new_sheet.row_dimensions[row].height = src_sheet.row_dimensions[row].height
        
        # Copy merged cells
        for merged_range in src_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
            
        src_wb.close()

    # Save the final combined file
    new_wb.save(output_path)
    print(f"✅ Combined file saved to: {output_path}")

def merge_excels_to_subsheets_two(sheet_paths, output_path="combined.xlsx"):
    """
    Merges the active sheets from the given Excel files into one workbook.
    Each input file becomes a sheet in the output file, preserving structure.

    Parameters:
    - sheet_paths: List of 4 file paths to Excel files
    - output_path: Path to save the new workbook (default: combined.xlsx)
    """
    if len(sheet_paths) != 4:
        raise ValueError("Exactly 4 sheet paths must be provided.")

    # Create a new workbook
    new_wb = Workbook()
    # Remove the default sheet
    default_sheet = new_wb.active
    new_wb.remove(default_sheet)

    # Add sheets from the given files
    for i, path in enumerate(sheet_paths):
        print(f"Processing: {path}")
        src_wb = load_workbook(path)
        src_sheet = src_wb.active

        # Use the same sheet title if available, or make a custom name
        sheet_title = src_sheet.title.strip() or f"Sheet{i+1}"
        new_sheet = new_wb.create_sheet(title=sheet_title)

        # Copy all cells with their values and formatting
        for row in src_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                
                # Copy cell formatting
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # Copy column dimensions
        for col in src_sheet.column_dimensions:
            new_sheet.column_dimensions[col].width = src_sheet.column_dimensions[col].width
        
        # Copy row dimensions
        for row in src_sheet.row_dimensions:
            new_sheet.row_dimensions[row].height = src_sheet.row_dimensions[row].height
        
        # Copy merged cells
        for merged_range in src_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
            
        src_wb.close()

    # Save the final combined file
    new_wb.save(output_path)
    print(f"✅ Combined file saved to: {output_path}")




def merge_excels_to_subsheets_three(sheet_paths, output_path="combined.xlsx"):
    """
    Merges the active sheets from the given Excel files into one workbook.
    Each input file becomes a sheet in the output file, preserving structure.

    Parameters:
    - sheet_paths: List of 4 file paths to Excel files
    - output_path: Path to save the new workbook (default: combined.xlsx)
    """
    if len(sheet_paths) != 13:
        raise ValueError("Exactly 13 sheet paths must be provided.")

    # Create a new workbook
    new_wb = Workbook()
    # Remove the default sheet
    default_sheet = new_wb.active
    new_wb.remove(default_sheet)

    # Add sheets from the given files
    for i, path in enumerate(sheet_paths):
        print(f"Processing: {path}")
        src_wb = load_workbook(path)
        src_sheet = src_wb.active

        # Use the same sheet title if available, or make a custom name
        sheet_title = src_sheet.title.strip() or f"Sheet{i+1}"
        new_sheet = new_wb.create_sheet(title=sheet_title)

        # Copy all cells with their values and formatting
        for row in src_sheet.iter_rows():
            for cell in row:
                new_cell = new_sheet[cell.coordinate]
                new_cell.value = cell.value
                
                # Copy cell formatting
                if cell.has_style:
                    new_cell.font = copy(cell.font)
                    new_cell.border = copy(cell.border)
                    new_cell.fill = copy(cell.fill)
                    new_cell.number_format = cell.number_format
                    new_cell.protection = copy(cell.protection)
                    new_cell.alignment = copy(cell.alignment)
        
        # Copy column dimensions
        for col in src_sheet.column_dimensions:
            new_sheet.column_dimensions[col].width = src_sheet.column_dimensions[col].width
        
        # Copy row dimensions
        for row in src_sheet.row_dimensions:
            new_sheet.row_dimensions[row].height = src_sheet.row_dimensions[row].height
        
        # Copy merged cells
        for merged_range in src_sheet.merged_cells.ranges:
            new_sheet.merge_cells(str(merged_range))
            
        src_wb.close()

    # Save the final combined file
    new_wb.save(output_path)
    print(f"✅ Combined file saved to: {output_path}")


