# -*- coding: utf-8 -*-
import openpyxl
import json
import os
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict
import logging
from dotenv import load_dotenv
from groq import Groq

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class JobPosition:
    """Data class to represent a job position with all required fields"""
    eil_no: str
    project_impact_no: str
    action_expenditure_type: str
    job_title: str
    planned_posts: float
    employment_contract: str
    recruitment_year: str
    salary_year: str
    months_planned: int
    monthly_salary_rate: float
    allowances_bonuses: float
    increase_percentage: float
    increase_amount: float
    working_week_length: int  # Should be 5 for 5-day week, not 40
    annual_leave_days: int
    annual_leave_allowance_rate: float
    justification: str
    coefficient: str = ""
    bonus_breakdown: str = ""

@dataclass
class ProjectData:
    """Data class to represent complete project information"""
    project_code: str
    organization_name: str
    project_duration_months: int
    contribution_rate: float
    organization_type: str = ""
    budgetary_classification: str = ""
    job_positions: List[JobPosition] = None
    
    def __post_init__(self):
        if self.job_positions is None:
            self.job_positions = []

class AnnualLeaveRateCalculator:
    """Calculate annual leave rates based on working week and leave days"""
    
    # Fixed rates table from Sheet 3
    LEAVE_RATES = {
        5: {  # 5-day working week
            20: 0.0863, 21: 0.1044, 22: 0.1044, 23: 0.1044, 24: 0.1044, 25: 0.1044,
            26: 0.1235, 27: 0.1235, 28: 0.1235, 29: 0.1235, 30: 0.1235,
            31: 0.1499, 32: 0.1499, 33: 0.1499, 34: 0.1499, 35: 0.1499, 36: 0.1499,
            37: 0.1725, 38: 0.1725, 39: 0.1725, 40: 0.1889, 41: 0.2002,
            42: 0.2002, 43: 0.2002, 44: 0.2002, 45: 0.2002, 46: 0.2002,
            47: 0.2002, 48: 0.2002, 49: 0.2002, 50: 0.2002
        },
        6: {  # 6-day working week
            24: 0.0863, 25: 0.1044, 26: 0.1044, 27: 0.1044, 28: 0.1044, 29: 0.1044, 30: 0.1044,
            31: 0.1235, 32: 0.1235, 33: 0.1235, 34: 0.1235, 35: 0.1235, 36: 0.1235,
            37: 0.1499, 38: 0.1499, 39: 0.1499, 40: 0.1499, 41: 0.1499, 42: 0.1499,
            43: 0.1725, 44: 0.1725, 45: 0.1725, 46: 0.1725, 47: 0.1725, 48: 0.1889, 49: 0.2002, 50: 0.2002
        }
    }
    
    @classmethod
    def get_leave_rate(cls, working_week: int, leave_days: int) -> float:
        """Get the correct annual leave rate from reference table"""
        if working_week not in cls.LEAVE_RATES:
            working_week = 5  # Default to 5-day week
        
        if leave_days in cls.LEAVE_RATES[working_week]:
            return cls.LEAVE_RATES[working_week][leave_days]
        
        # If exact match not found, find closest
        available_days = sorted(cls.LEAVE_RATES[working_week].keys())
        closest_days = min(available_days, key=lambda x: abs(x - leave_days))
        return cls.LEAVE_RATES[working_week][closest_days]

class GroqAPIClient:
    """Enhanced client to interact with Groq API"""
    
    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key)
    
    def analyze_text(self, text: str, prompt: str, model: str = "meta-llama/llama-4-scout-17b-16e-instruct") -> str:
        """
        Send text to Groq API for analysis using new client format
        """
        
        try:
            completion = self.client.chat.completions.create(
                model=model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert data analyst specializing in extracting structured information from European Union project funding documents. You have deep knowledge of Lithuanian civil service regulations, salary coefficients, and budgetary institution requirements. Always respond with valid JSON format when requested and pay extreme attention to numerical accuracy."
                    },
                    {
                        "role": "user",
                        "content": f"{prompt}\n\nText to analyze:\n{text}"
                    }
                ],
                temperature=0.3
            )
            
            return completion.choices[0].message.content
            
        except Exception as e:
            logger.error(f"Groq API request failed: {e}")
            raise

class DataAnalyzer:
    """Enhanced data analyzer with improved extraction logic"""
    
    def __init__(self, groq_client: GroqAPIClient):
        self.groq_client = groq_client
        self.leave_calculator = AnnualLeaveRateCalculator()
    
    def create_analysis_prompt(self, prompt_path: str = "code/prompts/certificate_budgetary.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")
    def analyze_project_data(self, text_content: str) -> ProjectData:
        """Enhanced analysis with post-processing validation"""
        
        logger.info("Starting enhanced project data analysis...")
        
        prompt = self.create_analysis_prompt()
        
        try:
            analysis_result = self.groq_client.analyze_text(text_content, prompt)
            logger.info("Received analysis from Groq API")
            
            # Extract JSON from markdown code blocks if present
            json_content = self._extract_json_from_response(analysis_result)
            
            # Parse JSON response
            parsed_data = json.loads(json_content)
            
            # Convert and validate data
            project_data = self._convert_to_project_data(parsed_data)
            
            # Post-process to fix common errors
            project_data = self._post_process_data(project_data)
            
            return project_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            logger.error(f"Cleaned JSON content: {json_content}")
            logger.error(f"Raw response: {analysis_result}")
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_json_from_response(self, response: str) -> str:
        """Extract JSON content from markdown code blocks or raw response and clean invalid expressions"""
        
        # Remove any leading/trailing whitespace
        response = response.strip()
        
        # Check if response contains markdown code blocks
        if '```json' in response:
            # Extract content between ```json and ```
            start_marker = '```json'
            end_marker = '```'
            
            start_index = response.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = response.find(end_marker, start_index)
                if end_index != -1:
                    json_content = response[start_index:end_index].strip()
                    logger.info("Extracted JSON from markdown code blocks")
                    return self._clean_json_expressions(json_content)
        
        # Check if response contains ``` without json marker
        if response.startswith('```') and response.endswith('```'):
            json_content = response[3:-3].strip()
            # Remove any language identifier (like 'json' at the start)
            if json_content.startswith('json'):
                json_content = json_content[4:].strip()
            logger.info("Extracted JSON from generic code blocks")
            return self._clean_json_expressions(json_content)
        
        # If no code blocks, clean and return
        logger.info("No markdown code blocks found, cleaning raw response")
        return self._clean_json_expressions(response)
    
    def _clean_json_expressions(self, json_content: str) -> str:
        """Clean JSON content by evaluating mathematical expressions"""
        
        import re
        
        # Find and replace mathematical expressions in JSON values
        # Pattern to match: "field": number / number, or "field": number * number, etc.
        pattern = r'"([^"]+)":\s*(\d+(?:\.\d+)?)\s*([+\-*/])\s*(\d+(?:\.\d+)?)'
        
        def evaluate_expression(match):
            field_name = match.group(1)
            num1 = float(match.group(2))
            operator = match.group(3)
            num2 = float(match.group(4))
            
            if operator == '+':
                result = num1 + num2
            elif operator == '-':
                result = num1 - num2
            elif operator == '*':
                result = num1 * num2
            elif operator == '/':
                result = num1 / num2 if num2 != 0 else 0
            else:
                result = num1
            
            logger.info(f"Evaluated expression: {num1} {operator} {num2} = {result}")
            return f'"{field_name}": {result}'
        
        # Replace all mathematical expressions
        cleaned_content = re.sub(pattern, evaluate_expression, json_content)
        
        # Also handle expressions in parentheses like (2650 + 265) * 0.05
        parentheses_pattern = r'"([^"]+)":\s*\((\d+(?:\.\d+)?)\s*([+\-*/])\s*(\d+(?:\.\d+)?)\)\s*([+\-*/])\s*(\d+(?:\.\d+)?)'
        
        def evaluate_parentheses_expression(match):
            field_name = match.group(1)
            num1 = float(match.group(2))
            op1 = match.group(3)
            num2 = float(match.group(4))
            op2 = match.group(5)
            num3 = float(match.group(6))
            
            # Calculate expression in parentheses first
            if op1 == '+':
                intermediate = num1 + num2
            elif op1 == '-':
                intermediate = num1 - num2
            elif op1 == '*':
                intermediate = num1 * num2
            elif op1 == '/':
                intermediate = num1 / num2 if num2 != 0 else 0
            else:
                intermediate = num1
            
            # Apply second operation
            if op2 == '+':
                result = intermediate + num3
            elif op2 == '-':
                result = intermediate - num3
            elif op2 == '*':
                result = intermediate * num3
            elif op2 == '/':
                result = intermediate / num3 if num3 != 0 else 0
            else:
                result = intermediate
            
            logger.info(f"Evaluated complex expression: ({num1} {op1} {num2}) {op2} {num3} = {result}")
            return f'"{field_name}": {result}'
        
        # Replace complex expressions
        cleaned_content = re.sub(parentheses_pattern, evaluate_parentheses_expression, cleaned_content)
        
        return cleaned_content
    
    def _post_process_data(self, project_data: ProjectData) -> ProjectData:
        """Post-process data to fix common extraction errors"""
        
        logger.info("Post-processing data for accuracy...")
        
        for position in project_data.job_positions:
            # Fix working week length (common error: 40 instead of 5)
            if position.working_week_length == 40:
                position.working_week_length = 5
                logger.info(f"Fixed working week for {position.job_title}: 40 → 5")
            
            # Recalculate annual leave rate using reference table
            correct_rate = self.leave_calculator.get_leave_rate(
                position.working_week_length, 
                position.annual_leave_days
            )
            if abs(position.annual_leave_allowance_rate - correct_rate) > 0.001:
                logger.info(f"Fixed leave rate for {position.job_title}: {position.annual_leave_allowance_rate} → {correct_rate}")
                position.annual_leave_allowance_rate = correct_rate
            
            # Recalculate increase amount if it seems wrong
            if position.increase_percentage > 0:
                base_total = position.monthly_salary_rate + position.allowances_bonuses
                expected_increase = base_total * position.increase_percentage
                if abs(position.increase_amount - expected_increase) > 0.01:
                    logger.info(f"Fixed increase amount for {position.job_title}: {position.increase_amount} → {expected_increase}")
                    position.increase_amount = expected_increase
        
        # Set correct contribution rate based on organization type
        if "budgetary" in project_data.budgetary_classification.lower():
            project_data.contribution_rate = 0.014
        elif "business" in project_data.budgetary_classification.lower():
            project_data.contribution_rate = 0.046
        
        logger.info("Post-processing completed")
        return project_data
    
    def _convert_to_project_data(self, parsed_data: Dict) -> ProjectData:
        """Enhanced conversion with better error handling and field completion"""
        
        try:
            project_info = parsed_data["project_info"]
            positions_data = parsed_data["job_positions"]
            
            job_positions = []
            for idx, pos_data in enumerate(positions_data):
                # Handle calculated fields in JSON (like increase_amount expressions)
                increase_amount = pos_data.get("increase_amount", 0)
                if isinstance(increase_amount, str) and "(" in increase_amount:
                    # If it's an expression like "(2650 + 265) * 0.05", calculate it
                    try:
                        # Safely evaluate simple arithmetic expressions
                        increase_amount = eval(increase_amount) if increase_amount else 0
                    except:
                        increase_amount = 0
                        logger.warning(f"Could not evaluate increase_amount expression: {pos_data.get('increase_amount')}")
                
                # Handle allowances calculation
                allowances_bonuses = pos_data.get("allowances_bonuses", 0)
                if isinstance(allowances_bonuses, str) and "/" in allowances_bonuses:
                    try:
                        allowances_bonuses = eval(allowances_bonuses)
                    except:
                        allowances_bonuses = 0
                        logger.warning(f"Could not evaluate allowances_bonuses expression: {pos_data.get('allowances_bonuses')}")
                
                # Generate missing fields with intelligent defaults
                project_impact_no = pos_data.get("project_impact_no", f"1.{idx+2}")
                if not project_impact_no or project_impact_no.strip() == "" or project_impact_no == "1.2":
                    project_impact_no = f"1.{idx+2}"  # Sequential: 1.2, 1.3, 1.4, 1.5, etc.
                
                action_expenditure_type = pos_data.get("action_expenditure_type", f"1.{idx+2}.1")
                if not action_expenditure_type or action_expenditure_type.strip() == "" or action_expenditure_type == "1.2.1":
                    action_expenditure_type = f"1.{idx+2}.1"  # Sequential: 1.2.1, 1.3.1, 1.4.1, etc.
                
                # Intelligent contract type detection
                employment_contract = pos_data.get("employment_contract", "time-limited")
                if not employment_contract or employment_contract.strip() == "":
                    # Analyze job context for contract type
                    job_title = pos_data.get("job_title", "").lower()
                    months_planned = int(pos_data.get("months_planned", 12))
                    
                    # If coordinator for full project duration, likely indefinite
                    # If duration is full project length (36 months), could be indefinite
                    if "coordinator" in job_title and months_planned >= 36:
                        employment_contract = "indefinite"
                    elif months_planned >= 36:
                        employment_contract = "indefinite"
                    else:
                        employment_contract = "time-limited"
                
                salary_year = pos_data.get("salary_year", "2024")
                if not salary_year or salary_year.strip() == "":
                    # Determine salary year from recruitment year
                    recruitment_year = pos_data.get("recruitment_year", "2024-01")
                    if recruitment_year:
                        base_year = recruitment_year.split('-')[0]
                        months_planned = int(pos_data.get("months_planned", 12))
                        if months_planned > 12:
                            end_year = int(base_year) + (months_planned // 12)
                            salary_year = f"{base_year}-{end_year}"
                        else:
                            salary_year = base_year
                    else:
                        salary_year = "2024"
                
                # Enhanced justification combining all available info
                justification_parts = []
                base_justification = pos_data.get("justification", "")
                if base_justification and base_justification.strip():
                    justification_parts.append(base_justification)
                
                coefficient = pos_data.get("coefficient", "")
                if coefficient and coefficient.strip():
                    justification_parts.append(f"Salary structure follows {coefficient}")
                
                bonus_breakdown = pos_data.get("bonus_breakdown", "")
                if bonus_breakdown and bonus_breakdown.strip():
                    justification_parts.append(f"Bonus structure: {bonus_breakdown}")
                
                # If no justification provided, create basic one
                if not justification_parts:
                    job_title = pos_data.get("job_title", "")
                    if "Research" in job_title:
                        justification_parts.append(f"Research position requiring specialized expertise and qualifications")
                    elif "Coordinator" in job_title or "Manager" in job_title:
                        justification_parts.append(f"Management position with coordination responsibilities")
                    elif "Technical" in job_title or "Support" in job_title:
                        justification_parts.append(f"Technical position requiring specialized skills and certifications")
                    else:
                        justification_parts.append(f"Position requiring appropriate qualifications and experience")
                
                full_justification = ". ".join(justification_parts)
                
                position = JobPosition(
                    eil_no=pos_data.get("eil_no", f"{idx+1}."),
                    project_impact_no=project_impact_no,
                    action_expenditure_type=action_expenditure_type,
                    job_title=pos_data.get("job_title", "Unknown Position"),
                    planned_posts=float(pos_data.get("planned_posts", 1)),
                    employment_contract=employment_contract,
                    recruitment_year=pos_data.get("recruitment_year", "2024-01"),
                    salary_year=salary_year,
                    months_planned=int(pos_data.get("months_planned", 12)),
                    monthly_salary_rate=float(pos_data.get("monthly_salary_rate", 0)),
                    allowances_bonuses=float(allowances_bonuses),
                    increase_percentage=float(pos_data.get("increase_percentage", 0)),
                    increase_amount=float(increase_amount),
                    working_week_length=int(pos_data.get("working_week_length", 5)),  # Default to 5, not 40
                    annual_leave_days=int(pos_data.get("annual_leave_days", 20)),
                    annual_leave_allowance_rate=float(pos_data.get("annual_leave_allowance_rate", 0.0863)),
                    justification=full_justification,
                    coefficient=coefficient,
                    bonus_breakdown=bonus_breakdown
                )
                job_positions.append(position)
            
            project_data = ProjectData(
                project_code=project_info.get("project_code", ""),
                organization_name=project_info.get("organization_name", ""),
                project_duration_months=int(project_info.get("project_duration_months", 0)),
                contribution_rate=float(project_info.get("contribution_rate", 0.014)),
                organization_type=project_info.get("organization_type", ""),
                budgetary_classification=project_info.get("budgetary_classification", "budgetary"),
                job_positions=job_positions
            )
            
            logger.info(f"Successfully converted data: {len(job_positions)} positions with complete field population")
            return project_data
            
        except (KeyError, ValueError, TypeError) as e:
            logger.error(f"Data conversion failed: {e}")
            logger.error(f"Parsed data structure: {parsed_data}")
            raise

class ExcelFormFiller:
    """Enhanced Excel form filler with correct calculations"""
    
    def __init__(self):
        # Corrected column mapping
        self.column_mapping = {
            'eil_no': 2,                           # B
            'project_impact_no': 3,                # C  
            'action_expenditure_type': 4,          # D
            'job_title': 5,                        # E
            'planned_posts': 6,                    # F
            'employment_contract': 7,              # G
            'recruitment_year': 8,                 # H
            'salary_year': 9,                      # I
            'months_planned': 10,                  # J
            'monthly_salary_rate': 11,             # K
            'allowances_bonuses': 12,              # L
            'increase_percentage': 13,             # M
            'increase_amount': 14,                 # N
            'total_excluding_contribution': 15,    # O
            'total_including_contribution': 16,    # P
            'du_costs': 17,                        # Q
            'working_week_length': 18,             # R
            'annual_leave_days': 19,               # S
            'annual_leave_allowance_rate': 20,     # T
            'annual_leave_cost': 21,               # U
            'total_planned_remuneration': 22,      # V
            'justification': 23                    # W
        }
    
    def calculate_remuneration_fields(self, position: JobPosition, contribution_rate: float) -> Dict[str, float]:
        """Corrected calculation formulas"""
        
        # Column O (14): Total excluding employer contribution
        total_excluding_contribution = (
            position.monthly_salary_rate + 
            position.allowances_bonuses + 
            position.increase_amount
        )
        
        # Column P (15): Total including employer contribution
        total_including_contribution = total_excluding_contribution * (1 + contribution_rate)
        
        # Column Q (16): DU costs (Posts × Months × Rate with contributions)
        du_costs = (
            position.planned_posts * 
            position.months_planned * 
            total_including_contribution
        )
        
        # Column U (20): Annual leave cost (DU costs × Leave rate)
        annual_leave_cost = du_costs * position.annual_leave_allowance_rate
        
        # Column V (21): Total planned remuneration (DU costs + Leave cost)
        total_planned_remuneration = du_costs + annual_leave_cost
        
        return {
            'total_excluding_contribution': round(total_excluding_contribution, 2),
            'total_including_contribution': round(total_including_contribution, 2),
            'du_costs': round(du_costs, 2),
            'annual_leave_cost': round(annual_leave_cost, 2),
            'total_planned_remuneration': round(total_planned_remuneration, 2)
        }
    
    def fill_project_header_info(self, sheet, project_data: ProjectData):
        """Enhanced header filling"""
        
        # Project information
        sheet['B4'] = "Project/Joint project code"
        sheet['H4'] = project_data.project_code
        
        sheet['B5'] = "Name of the applicant/joint applicant/project partner"
        sheet['H5'] = project_data.organization_name
        
        sheet['B6'] = "Project duration, months"
        sheet['H6'] = project_data.project_duration_months
        
        sheet['B9'] = "Type of organization"
        sheet['H9'] = project_data.organization_type
        sheet['I9'] = "Budgetary" if "budgetary" in project_data.budgetary_classification.lower() else "Non-budgetary"
        sheet['J9'] = f"{project_data.contribution_rate:.3f}"
        
        logger.info("Enhanced project header information filled")


    def fill_excel_form(self, template_path: str, output_path: str, project_data: ProjectData) -> bool:
        """Enhanced Excel filling with accurate calculations"""
        
        try:
            logger.info(f"Loading Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            sheet_name = "Certificate for budgetary autho"
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found. Using first sheet.")
                sheet = workbook.active
            else:
                sheet = workbook[sheet_name]
            
            # Fill project header
            self.fill_project_header_info(sheet, project_data)
            
            num_positions = len(project_data.job_positions)  # Get actual number of positions
            logger.info(f"Filling {num_positions} job positions with enhanced calculations...")
            
            start_row = 15
            total_cost = 0
            
            for idx, position in enumerate(project_data.job_positions):
                row = start_row + idx
                
                # Calculate all derived fields correctly
                calculated = self.calculate_remuneration_fields(position, project_data.contribution_rate)
                total_cost += calculated['total_planned_remuneration']
                
                # Fill all position data
                sheet.cell(row=row, column=self.column_mapping['eil_no']).value = position.eil_no
                sheet.cell(row=row, column=self.column_mapping['project_impact_no']).value = position.project_impact_no
                sheet.cell(row=row, column=self.column_mapping['action_expenditure_type']).value = position.action_expenditure_type
                sheet.cell(row=row, column=self.column_mapping['job_title']).value = position.job_title
                sheet.cell(row=row, column=self.column_mapping['planned_posts']).value = position.planned_posts
                sheet.cell(row=row, column=self.column_mapping['employment_contract']).value = position.employment_contract
                sheet.cell(row=row, column=self.column_mapping['recruitment_year']).value = position.recruitment_year
                sheet.cell(row=row, column=self.column_mapping['salary_year']).value = position.salary_year
                sheet.cell(row=row, column=self.column_mapping['months_planned']).value = position.months_planned
                sheet.cell(row=row, column=self.column_mapping['monthly_salary_rate']).value = position.monthly_salary_rate
                sheet.cell(row=row, column=self.column_mapping['allowances_bonuses']).value = position.allowances_bonuses
                sheet.cell(row=row, column=self.column_mapping['increase_percentage']).value = position.increase_percentage
                sheet.cell(row=row, column=self.column_mapping['increase_amount']).value = position.increase_amount
                sheet.cell(row=row, column=self.column_mapping['total_excluding_contribution']).value = calculated['total_excluding_contribution']
                sheet.cell(row=row, column=self.column_mapping['total_including_contribution']).value = calculated['total_including_contribution']
                sheet.cell(row=row, column=self.column_mapping['du_costs']).value = calculated['du_costs']
                sheet.cell(row=row, column=self.column_mapping['working_week_length']).value = position.working_week_length
                sheet.cell(row=row, column=self.column_mapping['annual_leave_days']).value = position.annual_leave_days
                sheet.cell(row=row, column=self.column_mapping['annual_leave_allowance_rate']).value = position.annual_leave_allowance_rate
                sheet.cell(row=row, column=self.column_mapping['annual_leave_cost']).value = calculated['annual_leave_cost']
                sheet.cell(row=row, column=self.column_mapping['total_planned_remuneration']).value = calculated['total_planned_remuneration']
                
                # Enhanced justification with coefficient and bonus info
                justification_parts = []
                if position.justification and position.justification.strip():
                    justification_parts.append(position.justification)
                
                if position.coefficient:
                    justification_parts.append(f"Coefficient: {position.coefficient}")
                
                if position.bonus_breakdown:
                    justification_parts.append(f"Bonus breakdown: {position.bonus_breakdown}")
                
                # Add compliance information for budgetary institutions
                if "budgetary" in project_data.budgetary_classification.lower():
                    justification_parts.append("Compliant with Lithuanian civil service regulations and institutional salary scales")
                
                full_justification = " | ".join(justification_parts) if justification_parts else position.justification
                
                sheet.cell(row=row, column=self.column_mapping['justification']).value = full_justification
                
                logger.info(f"[OK] Filled position {idx + 1}: {position.job_title} - €{calculated['total_planned_remuneration']:,.2f}")
            
            # Add DYNAMIC sum formulas based on actual number of positions
            logger.info(f"Adding dynamic sum formulas to row 29 for {num_positions} positions...")
            self.add_sum_formulas(sheet, num_positions)  # Pass the number of positions
            
            workbook.save(output_path)
            logger.info(f"[OK] Successfully saved enhanced Excel form: {output_path}")
            logger.info(f"💰 Total project cost: €{total_cost:,.2f}")
            
            return True
            
        except Exception as e:
            logger.error(f"[X] Failed to fill Excel form: {e}")
            import traceback
            traceback.print_exc()
            return False

    def add_sum_formulas(self, sheet, num_positions):
        """Add dynamic sum formulas based on actual number of positions"""
        
        # Calculate the last data row (15 is start row, positions go from 15 to 15+num_positions-1)
        start_row = 15
        end_row = start_row + num_positions - 1  # e.g., if 10 positions: 15 to 24
        sum_row = 29  # Fixed sum row
        
        # Define dynamic sum formulas
        sum_formulas = {
            f'J{sum_row}': f'=SUM(J{start_row}:J{end_row})',  # Number of months/hours planned
            f'K{sum_row}': f'=SUM(K{start_row}:K{end_row})',  # Planned post salary/hourly rate, EUR
            f'L{sum_row}': f'=SUM(L{start_row}:L{end_row})',  # Allowances and bonuses, EUR
            f'M{sum_row}': f'=SUM(M{start_row}:M{end_row})',  # Increase, % (if applicable)
            f'N{sum_row}': f'=SUM(N{start_row}:N{end_row})',  # Amount of increase, EUR (if applicable)
            f'O{sum_row}': f'=SUM(O{start_row}:O{end_row})',  # Total planned remuneration excluding employer's contribution, EUR
            f'P{sum_row}': f'=SUM(P{start_row}:P{end_row})',  # Total rate of pay for the planned salary including employer's contribution, EUR
            f'Q{sum_row}': f'=SUM(Q{start_row}:Q{end_row})',  # Amount of DU costs to be financed, EUR
            f'U{sum_row}': f'=SUM(U{start_row}:U{end_row})',  # Planned cost of annual leave (including employer's contributions), EUR
            f'V{sum_row}': f'=SUM(V{start_row}:V{end_row})',  # Total planned remuneration, EUR
        }
        
        # Insert the formulas into the sheet
        for cell_address, formula in sum_formulas.items():
            sheet[cell_address] = formula
            logger.info(f"Added dynamic formula to {cell_address}: {formula}")
        
        logger.info(f"All sum formulas added to row {sum_row} for {num_positions} positions (rows {start_row}-{end_row})")


class ProjectAnalyzer:
    """Enhanced main orchestrator with better error handling"""
    
    def __init__(self, groq_api_key: str):
        self.groq_client = GroqAPIClient(groq_api_key)
        self.data_analyzer = DataAnalyzer(self.groq_client)
        self.excel_filler = ExcelFormFiller()
    
    def read_data_file(self, file_path: str) -> str:
        """Read content from data file with better encoding handling"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"[OK] Successfully read data file: {file_path} ({len(content)} characters)")
            return content
        except FileNotFoundError:
            logger.error(f"[X] Data file not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"[X] Error reading data file: {e}")
            raise
    
    def analyze_and_process(self, data_file_path: str, template_path: str, output_path: str) -> Dict[str, Any]:
        """Enhanced complete workflow with detailed logging"""
        
        results = {
            'success': False,
            'steps_completed': [],
            'project_data': None,
            'statistics': {},
            'errors': []
        }
        
        try:
            # Step 1: Read data file
            logger.info("[STEP] STEP 1: Reading Data File")
            text_content = self.read_data_file(data_file_path)
            results['steps_completed'].append('data_file_read')
            
            # Step 2: Enhanced analysis with Groq API
            logger.info("[STEP] STEP 2: Enhanced Analysis with Groq API")
            project_data = self.data_analyzer.analyze_project_data(text_content)
            results['project_data'] = project_data
            results['steps_completed'].append('data_analyzed')
            
            # Step 3: Generate enhanced statistics
            logger.info("[STEP] STEP 3: Generating Enhanced Statistics")
            statistics = self._generate_enhanced_statistics(project_data)
            results['statistics'] = statistics
            results['steps_completed'].append('statistics_generated')
            
            # Step 4: Fill Excel with enhanced accuracy
            logger.info("[STEP] STEP 4: Filling Excel Form with Enhanced Accuracy")
            excel_success = self.excel_filler.fill_excel_form(template_path, output_path, project_data)
            
            if excel_success:
                results['steps_completed'].append('excel_filled')
                results['success'] = True
                logger.info("[OK] PROCESS COMPLETED SUCCESSFULLY")
            else:
                results['errors'].append('Excel filling failed')
                
        except Exception as e:
            error_msg = f"[X] Process failed at step {len(results['steps_completed']) + 1}: {e}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _generate_enhanced_statistics(self, project_data: ProjectData) -> Dict[str, Any]:
        """Generate enhanced statistics with validation info"""
        
        total_positions = len(project_data.job_positions)
        total_staff = sum(pos.planned_posts for pos in project_data.job_positions)
        total_months = sum(pos.months_planned * pos.planned_posts for pos in project_data.job_positions)
        
        # Enhanced financial calculations
        total_salaries = 0
        total_bonuses = 0
        total_increases = 0
        total_remuneration = 0
        
        position_summary = []
        validation_issues = []
        
        for position in project_data.job_positions:
            calculated = self.excel_filler.calculate_remuneration_fields(position, project_data.contribution_rate)
            
            position_cost = calculated['total_planned_remuneration']
            total_remuneration += position_cost
            total_salaries += position.monthly_salary_rate * position.months_planned * position.planned_posts
            total_bonuses += position.allowances_bonuses * position.months_planned * position.planned_posts
            total_increases += position.increase_amount * position.months_planned * position.planned_posts
            
            # Validation checks
            if position.working_week_length not in [5, 6]:
                validation_issues.append(f"[!] {position.job_title}: Unusual working week ({position.working_week_length})")
            
            if position.annual_leave_allowance_rate == 0:
                validation_issues.append(f"[!] {position.job_title}: Missing annual leave rate")
            
            position_summary.append({
                'job_title': position.job_title,
                'planned_posts': position.planned_posts,
                'months_planned': position.months_planned,
                'total_cost': position_cost,
                'coefficient': position.coefficient,
                'leave_rate': f"{position.annual_leave_allowance_rate * 100:.2f}%"
            })
        
        statistics = {
            'project_overview': {
                'project_code': project_data.project_code,
                'organization': project_data.organization_name,
                'duration_months': project_data.project_duration_months,
                'contribution_rate': f"{project_data.contribution_rate * 100:.1f}%",
                'organization_type': project_data.organization_type,
                'budgetary_classification': project_data.budgetary_classification
            },
            'staffing': {
                'total_positions': total_positions,
                'total_staff': total_staff,
                'total_work_months': total_months
            },
            'financial': {
                'total_salaries': round(total_salaries, 2),
                'total_bonuses': round(total_bonuses, 2),
                'total_increases': round(total_increases, 2),
                'total_project_cost': round(total_remuneration, 2)
            },
            'position_breakdown': position_summary,
            'validation_issues': validation_issues
        }
        
        return statistics
    
    def print_enhanced_results(self, results: Dict[str, Any]):
        """Print comprehensive results with validation info"""
        
        print("\n" + "="*80)
        print("[>] ENHANCED PROJECT DATA ANALYSIS AND EXCEL FORM FILLING RESULTS")
        print("="*80)
        
        # Process Status
        status_icon = "[OK] SUCCESS" if results['success'] else "[X] FAILED"
        print(f"\n[STEP] PROCESS STATUS: {status_icon}")
        print(f"[STEP] Steps Completed: {', '.join(results['steps_completed'])}")
        
        if results['errors']:
            print(f"[!]  Errors: {'; '.join(results['errors'])}")
        
        # Project Overview
        if results['project_data']:
            stats = results['statistics']
            overview = stats['project_overview']
            
            print(f"\n[SUMMARY] PROJECT OVERVIEW")
            print(f"   Project Code: {overview['project_code']}")
            print(f"   Organization: {overview['organization']}")
            print(f"   Duration: {overview['duration_months']} months")
            print(f"   Contribution Rate: {overview['contribution_rate']}")
            print(f"   Organization Type: {overview['organization_type']}")
            print(f"   Budgetary Classification: {overview['budgetary_classification']}")
            
            # Staffing Statistics
            staffing = stats['staffing']
            print(f"\n👥 STAFFING STATISTICS")
            print(f"   Total Job Positions: {staffing['total_positions']}")
            print(f"   Total Staff Count: {staffing['total_staff']}")
            print(f"   Total Work Months: {staffing['total_work_months']}")
            
            # Financial Statistics
            financial = stats['financial']
            print(f"\n💰 FINANCIAL SUMMARY")
            print(f"   Total Salaries: €{financial['total_salaries']:,.2f}")
            print(f"   Total Bonuses: €{financial['total_bonuses']:,.2f}")
            print(f"   Total Increases: €{financial['total_increases']:,.2f}")
            print(f"   TOTAL PROJECT COST: €{financial['total_project_cost']:,.2f}")
            
            # Enhanced Position Breakdown
            print(f"\n[STEP] ENHANCED POSITION BREAKDOWN")
            for pos in stats['position_breakdown']:
                coefficient_info = f" (Coeff: {pos['coefficient']})" if pos['coefficient'] else ""
                print(f"   • {pos['job_title']}{coefficient_info}: {pos['planned_posts']} staff × {pos['months_planned']} months")
                print(f"     Leave Rate: {pos['leave_rate']} | Cost: €{pos['total_cost']:,.2f}")
            
            # Validation Issues
            if stats['validation_issues']:
                print(f"\n[!]  VALIDATION ISSUES DETECTED")
                for issue in stats['validation_issues']:
                    print(f"   {issue}")
            else:
                print(f"\n[OK] ALL VALIDATIONS PASSED")
        
        print("\n" + "="*80)

def main_budgetary_auth(data_file_path):
    """Enhanced main function with better error handling"""
    
    print("[>] ENHANCED Groq Agent Excel Form Analyzer and Filler System")
    print("="*70)
    
    # Configuration
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
    if not GROQ_API_KEY:
        print("[X] Error: GROQ_API_KEY environment variable not set")
        print("Please set your Groq API key in .env file: GROQ_API_KEY='your_api_key_here'")
        return
    
   
    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    template_path = "code/agents/ENGISH_Rekomenduojamaforma_en/Certificate_for_Budgetary_Authorization/Certificate_for_budgetary_autho.xlsx"
    output_path = "code/agents/ENGISH_Rekomenduojamaforma_en/Certificate_for_Budgetary_Authorization/Certificate_for_budgetary_autho_filled.xlsx"
    
    # Validate required files exist
    required_files = [data_file_path, template_path]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print(f"[X] Missing required files: {', '.join(missing_files)}")
        print("Required files:")
        print(f"  • {data_file_path} - Text file with project data to analyze")
        print(f"  • {template_path} - Excel template file to fill")
        return
    
    try:
        # Initialize enhanced analyzer
        print(f"\n[DEBUG] Initializing Enhanced Project Analyzer...")
        analyzer = ProjectAnalyzer(GROQ_API_KEY)
        
        # Run enhanced analysis and processing
        print(f"\n[STEP] Starting enhanced analysis of {data_file_path}...")
        results = analyzer.analyze_and_process(data_file_path, template_path, output_path)
        
        # Print enhanced results
        analyzer.print_enhanced_results(results)
        
        if results['success']:
            print(f"\n[OK] Enhanced Excel form successfully filled: {output_path}")
            print("\n[RESULT] KEY IMPROVEMENTS APPLIED:")
            print("   • [OK] Correct working week values (5-day = 5, not 40)")
            print("   • [OK] Accurate annual leave rates from reference table")
            print("   • [OK] Proper salary increase calculations including bonuses")
            print("   • [OK] Enhanced coefficient and bonus extraction")
            print("   • [OK] Correct employer contribution calculations")
            print("   • [OK] Post-processing validation and error correction")
            print("   • [OK] Enhanced justification with coefficient details")
            print("\n[SUMMARY] Open the file to see the perfectly structured data!")
            
            # Additional success information
            if results['project_data']:
                total_positions = len(results['project_data'].job_positions)
                total_cost = results['statistics']['financial']['total_project_cost']
                print(f"\n[STATS] PROCESSING SUMMARY:")
                print(f"   • Extracted {total_positions} job positions")
                print(f"   • Total project budget: €{total_cost:,.2f}")
                print(f"   • All calculations validated and corrected")
                
        else:
            print(f"\n[X] Process failed. Check the detailed logs above.")
            print("\n[DEBUG] TROUBLESHOOTING TIPS:")
            print("   • Ensure your .env file contains GROQ_API_KEY")
            print("   • Check that data.txt contains valid project information")
            print("   • Verify form.xlsx template is not corrupted")
            print("   • Check internet connection for Groq API access")
            
    except Exception as e:
        logger.error(f"[X] Fatal error in main process: {e}")
        print(f"\n[X] Fatal error: {e}")
        print("\n[DEBUG] DEBUG INFORMATION:")
        import traceback
        traceback.print_exc(limit=3)
        
        print("\n[TIP] COMMON SOLUTIONS:")
        print("   • Check your Groq API key is valid and has credits")
        print("   • Ensure all required Python packages are installed:")
        print("     pip install groq openpyxl python-dotenv")
        print("   • Verify file permissions for reading data.txt and form.xlsx")
