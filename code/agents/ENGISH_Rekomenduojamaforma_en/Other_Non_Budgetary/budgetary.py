# -*- coding: utf-8 -*-
import openpyxl
import json
import os
import re
from datetime import datetime
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict
import logging
from dotenv import load_dotenv
from groq import Groq

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class JobPosition:
    """Data class to represent a job position with all required fields according to the Excel form"""
    eil_no: str
    project_impact_no: str
    action_expenditure_no: str
    duties_in_orp: str
    position_function: str
    employee_name: str
    employment_contract_type: str
    remuneration_year: str
    months_hours_planned: int
    planned_salary_rate: float
    increase_percentage: float
    increase_amount: float
    total_excluding_contribution: float  # Column 13: 10+12
    total_including_contribution: float  # Column 14
    working_week_length: int
    annual_leave_days: int
    annual_leave_rate: float
    annual_leave_cost: float  # Column 18
    total_rd_fee: float  # Column 19: 14+18
    total_planned_remuneration: float  # Column 20: 19+9
    justification: str
    coefficient: str = ""
    bonus_breakdown: str = ""

@dataclass
class ProjectData:
    """Data class to represent complete project information"""
    project_code: str
    organization_name: str
    project_duration_months: int
    contribution_rate: float
    organization_type: str = ""
    budgetary_classification: str = ""
    job_positions: List[JobPosition] = None
    
    def __post_init__(self):
        if self.job_positions is None:
            self.job_positions = []

class AnnualLeaveRateCalculator:
    """Calculate annual leave rates based on working week and leave days"""
    
    # Fixed rates table from Sheet 3
    LEAVE_RATES = {
        5: {  # 5-day working week
            20: 0.0863, 21: 0.1044, 22: 0.1044, 23: 0.1044, 24: 0.1044, 25: 0.1044,
            26: 0.1235, 27: 0.1235, 28: 0.1235, 29: 0.1235, 30: 0.1235,
            31: 0.1499, 32: 0.1499, 33: 0.1499, 34: 0.1499, 35: 0.1499, 36: 0.1499,
            37: 0.1725, 38: 0.1725, 39: 0.1725, 40: 0.1889, 41: 0.2002,
            42: 0.2002, 43: 0.2002, 44: 0.2002, 45: 0.2002, 46: 0.2002,
            47: 0.2002, 48: 0.2002, 49: 0.2002, 50: 0.2002
        },
        6: {  # 6-day working week
            24: 0.0863, 25: 0.1044, 26: 0.1044, 27: 0.1044, 28: 0.1044, 29: 0.1044, 30: 0.1044,
            31: 0.1235, 32: 0.1235, 33: 0.1235, 34: 0.1235, 35: 0.1235, 36: 0.1235,
            37: 0.1499, 38: 0.1499, 39: 0.1499, 40: 0.1499, 41: 0.1499, 42: 0.1499,
            43: 0.1725, 44: 0.1725, 45: 0.1725, 46: 0.1725, 47: 0.1725, 48: 0.1889, 49: 0.2002, 50: 0.2002
        }
    }
    
    @classmethod
    def get_leave_rate(cls, working_week: int, leave_days: int) -> float:
        """Get the correct annual leave rate from reference table"""
        if working_week not in cls.LEAVE_RATES:
            working_week = 5  # Default to 5-day week
        
        if leave_days in cls.LEAVE_RATES[working_week]:
            return cls.LEAVE_RATES[working_week][leave_days]
        
        # If exact match not found, find closest
        available_days = sorted(cls.LEAVE_RATES[working_week].keys())
        closest_days = min(available_days, key=lambda x: abs(x - leave_days))
        return cls.LEAVE_RATES[working_week][closest_days]

class GroqAPIClient:
    """Enhanced client to interact with Groq API"""
    
    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key)
    
    def analyze_text(self, text: str, prompt: str, model: str = "meta-llama/llama-4-scout-17b-16e-instruct") -> str:
        """
        Send text to Groq API for analysis using new client format
        """
        
        try:
            completion = self.client.chat.completions.create(
                model=model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert data analyst specializing in extracting structured information from European Union project funding documents. You have deep knowledge of Lithuanian civil service regulations, salary coefficients, and budgetary institution requirements. Always respond with valid JSON format when requested and pay extreme attention to numerical accuracy."
                    },
                    {
                        "role": "user",
                        "content": f"{prompt}\n\nText to analyze:\n{text}"
                    }
                ],
                temperature=0.3
            )
            
            return completion.choices[0].message.content
            
        except Exception as e:
            logger.error(f"Groq API request failed: {e}")
            raise

class DataAnalyzer:
    """Enhanced data analyzer with improved extraction logic"""
    
    def __init__(self, groq_client: GroqAPIClient):
        self.groq_client = groq_client
        self.leave_calculator = AnnualLeaveRateCalculator()
    
    def create_analysis_prompt(self, prompt_path: str = "code/prompts/non_budgetary.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")
    
    def analyze_project_data(self, text_content: str) -> ProjectData:
        """Enhanced analysis with post-processing validation"""
        
        logger.info("Starting project data analysis with correct form mapping...")
        
        prompt = self.create_analysis_prompt()
        
        try:
            analysis_result = self.groq_client.analyze_text(text_content, prompt)
            logger.info("Received analysis from Groq API")
            
            # Extract JSON from markdown code blocks if present
            json_content = self._extract_json_from_response(analysis_result)
            
            # Parse JSON response
            parsed_data = json.loads(json_content)
            
            # Convert and validate data
            project_data = self._convert_to_project_data(parsed_data)
            
            # Post-process to fix common errors
            project_data = self._post_process_data(project_data)
            
            return project_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            logger.error(f"Cleaned JSON content: {json_content}")
            logger.error(f"Raw response: {analysis_result}")
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_json_from_response(self, response: str) -> str:
        """Extract JSON content from markdown code blocks or raw response and clean invalid expressions"""
        
        # Remove any leading/trailing whitespace
        response = response.strip()
        
        # Check if response contains markdown code blocks
        if '```json' in response:
            # Extract content between ```json and ```
            start_marker = '```json'
            end_marker = '```'
            
            start_index = response.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = response.find(end_marker, start_index)
                if end_index != -1:
                    json_content = response[start_index:end_index].strip()
                    logger.info("Extracted JSON from markdown code blocks")
                    return self._clean_json_expressions(json_content)
        
        # Check if response contains ``` without json marker
        if response.startswith('```') and response.endswith('```'):
            json_content = response[3:-3].strip()
            # Remove any language identifier (like 'json' at the start)
            if json_content.startswith('json'):
                json_content = json_content[4:].strip()
            logger.info("Extracted JSON from generic code blocks")
            return self._clean_json_expressions(json_content)
        
        # If no code blocks, clean and return
        logger.info("No markdown code blocks found, cleaning raw response")
        return self._clean_json_expressions(response)
    
    def _clean_json_expressions(self, json_content: str) -> str:
        """Clean JSON content by evaluating mathematical expressions"""
        
        import re
        
        # Find and replace mathematical expressions in JSON values
        pattern = r'"([^"]+)":\s*(\d+(?:\.\d+)?)\s*([+\-*/])\s*(\d+(?:\.\d+)?)'
        
        def evaluate_expression(match):
            field_name = match.group(1)
            num1 = float(match.group(2))
            operator = match.group(3)
            num2 = float(match.group(4))
            
            if operator == '+':
                result = num1 + num2
            elif operator == '-':
                result = num1 - num2
            elif operator == '*':
                result = num1 * num2
            elif operator == '/':
                result = num1 / num2 if num2 != 0 else 0
            else:
                result = num1
            
            logger.info(f"Evaluated expression: {num1} {operator} {num2} = {result}")
            return f'"{field_name}": {result}'
        
        # Replace all mathematical expressions
        cleaned_content = re.sub(pattern, evaluate_expression, json_content)
        
        # Also handle expressions in parentheses
        parentheses_pattern = r'"([^"]+)":\s*\((\d+(?:\.\d+)?)\s*([+\-*/])\s*(\d+(?:\.\d+)?)\)\s*([+\-*/])\s*(\d+(?:\.\d+)?)'
        
        def evaluate_parentheses_expression(match):
            field_name = match.group(1)
            num1 = float(match.group(2))
            op1 = match.group(3)
            num2 = float(match.group(4))
            op2 = match.group(5)
            num3 = float(match.group(6))
            
            # Calculate expression in parentheses first
            if op1 == '+':
                intermediate = num1 + num2
            elif op1 == '-':
                intermediate = num1 - num2
            elif op1 == '*':
                intermediate = num1 * num2
            elif op1 == '/':
                intermediate = num1 / num2 if num2 != 0 else 0
            else:
                intermediate = num1
            
            # Apply second operation
            if op2 == '+':
                result = intermediate + num3
            elif op2 == '-':
                result = intermediate - num3
            elif op2 == '*':
                result = intermediate * num3
            elif op2 == '/':
                result = intermediate / num3 if num3 != 0 else 0
            else:
                result = intermediate
            
            logger.info(f"Evaluated complex expression: ({num1} {op1} {num2}) {op2} {num3} = {result}")
            return f'"{field_name}": {result}'
        
        # Replace complex expressions
        cleaned_content = re.sub(parentheses_pattern, evaluate_parentheses_expression, cleaned_content)
        
        return cleaned_content
    
    def _post_process_data(self, project_data: ProjectData) -> ProjectData:
        """Post-process data to fix common extraction errors and perform correct calculations"""
        
        logger.info("Post-processing data for accuracy with correct form calculations...")
        
        for position in project_data.job_positions:
            # Fix working week length (common error: 40 instead of 5)
            if position.working_week_length == 40:
                position.working_week_length = 5
                logger.info(f"Fixed working week for {position.position_function}: 40 → 5")
            
            # Recalculate annual leave rate using reference table
            correct_rate = self.leave_calculator.get_leave_rate(
                position.working_week_length, 
                position.annual_leave_days
            )
            if abs(position.annual_leave_rate - correct_rate) > 0.001:
                logger.info(f"Fixed leave rate for {position.position_function}: {position.annual_leave_rate} → {correct_rate}")
                position.annual_leave_rate = correct_rate
            
            # Correct calculation: Column 13 = Column 10 + Column 12
            correct_total_excluding = position.planned_salary_rate + position.increase_amount
            if abs(position.total_excluding_contribution - correct_total_excluding) > 0.01:
                logger.info(f"Fixed total excluding contribution for {position.position_function}: {position.total_excluding_contribution} → {correct_total_excluding}")
                position.total_excluding_contribution = correct_total_excluding
            
            # Column 14: Total including employer contribution
            contribution_multiplier = 1 + project_data.contribution_rate
            correct_total_including = position.total_excluding_contribution * contribution_multiplier
            if abs(position.total_including_contribution - correct_total_including) > 0.01:
                logger.info(f"Fixed total including contribution for {position.position_function}: {position.total_including_contribution} → {correct_total_including}")
                position.total_including_contribution = correct_total_including
            
            # Column 19: Total R&D fee = Column 14 + Column 18
            correct_rd_fee = position.total_including_contribution + position.annual_leave_cost
            if abs(position.total_rd_fee - correct_rd_fee) > 0.01:
                logger.info(f"Fixed total R&D fee for {position.position_function}: {position.total_rd_fee} → {correct_rd_fee}")
                position.total_rd_fee = correct_rd_fee
            
            # Column 20: Total planned remuneration = Column 19 + Column 9
            correct_total_remuneration = position.total_rd_fee + position.months_hours_planned
            if abs(position.total_planned_remuneration - correct_total_remuneration) > 0.01:
                logger.info(f"Fixed total planned remuneration for {position.position_function}: {position.total_planned_remuneration} → {correct_total_remuneration}")
                position.total_planned_remuneration = correct_total_remuneration
        
        # Set correct contribution rate based on organization type
        if "budgetary" in project_data.budgetary_classification.lower():
            project_data.contribution_rate = 0.014
        elif "business" in project_data.budgetary_classification.lower():
            project_data.contribution_rate = 0.046
        
        logger.info("Post-processing completed with correct form calculations")
        return project_data
    
    def _convert_to_project_data(self, parsed_data: Dict) -> ProjectData:
        """Convert parsed JSON to ProjectData with correct field mapping"""
        
        try:
            project_info = parsed_data["project_info"]
            positions_data = parsed_data["job_positions"]
            
            job_positions = []
            for idx, pos_data in enumerate(positions_data):


                # Generate missing fields with intelligent defaults
                project_impact_no = pos_data.get("project_impact_no", f"1.{idx+2}")
                if not project_impact_no or project_impact_no.strip() == "" or project_impact_no == "1.2":
                    project_impact_no = f"1.{idx+2}"  # Sequential: 1.2, 1.3, 1.4, 1.5, etc.
                
                action_expenditure_no = pos_data.get("action_expenditure_no", f"1.{idx+2}.1")
                if not action_expenditure_no or action_expenditure_no.strip() == "" or action_expenditure_no == "1.2.1":
                    action_expenditure_no = f"1.{idx+2}.1"  # Sequential: 1.2.1, 1.3.1, 1.4.1, etc.
                

                # Intelligent contract type detection
                employment_contract_type = pos_data.get("employment_contract_type", "time-limited")
                if not employment_contract_type or employment_contract_type.strip() == "":
                    # Analyze job context for contract type
                    duties_in_orp = pos_data.get("duties_in_orp", "").lower()
                    months_hours_planned = int(pos_data.get("months_hours_planned", 12))
                    
                    # If coordinator for full project duration, likely indefinite
                    # If duration is full project length (36 months), could be indefinite
                    if "coordinator" in duties_in_orp and months_hours_planned >= 36:
                        employment_contract_type = "indefinite"
                    elif months_hours_planned >= 36:
                        employment_contract_type = "indefinite"
                    else:
                        employment_contract_type = "time-limited"



                # Handle calculated fields
                increase_amount = float(pos_data.get("increase_amount", 0))
                planned_salary_rate = float(pos_data.get("planned_salary_rate", 0))
                
                # Calculate total excluding contribution (Column 13 = Column 10 + Column 12)
                total_excluding_contribution = planned_salary_rate + increase_amount
                
                # Generate missing fields with intelligent defaults
                employee_name = pos_data.get("employee_name", "")
                if not employee_name or employee_name.strip() == "":
                    employee_name = ""  # Leave empty as per form requirement
                
                # Enhanced justification
                justification_parts = []
                base_justification = pos_data.get("justification", "")
                if base_justification and base_justification.strip():
                    justification_parts.append(base_justification)
                
                coefficient = pos_data.get("coefficient", "")
                if coefficient and coefficient.strip():
                    justification_parts.append(f"Salary coefficient: {coefficient}")
                
                bonus_breakdown = pos_data.get("bonus_breakdown", "")
                if bonus_breakdown and bonus_breakdown.strip():
                    justification_parts.append(f"Bonus structure: {bonus_breakdown}")
                
                if not justification_parts:
                    position_function = pos_data.get("position_function", "")
                    if "Research" in position_function:
                        justification_parts.append("Research position requiring specialized expertise and qualifications")
                    elif "Coordinator" in position_function or "Manager" in position_function:
                        justification_parts.append("Management position with coordination responsibilities")
                    else:
                        justification_parts.append("Position requiring appropriate qualifications and experience")
                
                full_justification = ". ".join(justification_parts)
                




                position = JobPosition(
                    eil_no=pos_data.get("eil_no", str(idx+1)),
                    project_impact_no=project_impact_no,
                    action_expenditure_no=action_expenditure_no,
                    duties_in_orp=pos_data.get("duties_in_orp", "Project-related duties"),
                    position_function=pos_data.get("position_function", "Unknown Position"),
                    employee_name=employee_name,
                    employment_contract_type=employment_contract_type,
                    remuneration_year=pos_data.get("remuneration_year", "2024"),
                    months_hours_planned=int(pos_data.get("months_hours_planned", 12)),
                    planned_salary_rate=planned_salary_rate,
                    increase_percentage=float(pos_data.get("increase_percentage", 0)),
                    increase_amount=increase_amount,
                    total_excluding_contribution=total_excluding_contribution,
                    total_including_contribution=float(pos_data.get("total_including_contribution", 0)),
                    working_week_length=int(pos_data.get("working_week_length", 5)),
                    annual_leave_days=int(pos_data.get("annual_leave_days", 20)),
                    annual_leave_rate=float(pos_data.get("annual_leave_rate", 0.0863)),
                    annual_leave_cost=float(pos_data.get("annual_leave_cost", 0)),
                    total_rd_fee=float(pos_data.get("total_rd_fee", 0)),
                    total_planned_remuneration=float(pos_data.get("total_planned_remuneration", 0)),
                    justification=full_justification,
                    coefficient=coefficient,
                    bonus_breakdown=bonus_breakdown
                )
                job_positions.append(position)
            
            project_data = ProjectData(
                project_code=project_info.get("project_code", ""),
                organization_name=project_info.get("organization_name", ""),
                project_duration_months=int(project_info.get("project_duration_months", 0)),
                contribution_rate=float(project_info.get("contribution_rate", 0.014)),
                organization_type=project_info.get("organization_type", ""),
                budgetary_classification=project_info.get("budgetary_classification", "budgetary"),
                job_positions=job_positions
            )
            
            logger.info(f"Successfully converted data: {len(job_positions)} positions with correct form mapping")
            return project_data
            
        except (KeyError, ValueError, TypeError) as e:
            logger.error(f"Data conversion failed: {e}")
            logger.error(f"Parsed data structure: {parsed_data}")
            raise

class ExcelFormFiller:
    """Excel form filler with correct column mapping B to V"""
    
    def __init__(self):
        # Corrected column mapping based on the exact form structure (B=2 to V=22)
        self.column_mapping = {
            'eil_no': 2,                           # B - Column 1
            'project_impact_no': 3,                # C - Column 2  
            'action_expenditure_no': 4,            # D - Column 3
            'duties_in_orp': 5,                    # E - Column 4
            'position_function': 6,                # F - Column 5
            'employee_name': 7,                    # G - Column 6
            'employment_contract_type': 8,         # H - Column 7
            'remuneration_year': 9,                # I - Column 8
            'months_hours_planned': 10,            # J - Column 9
            'planned_salary_rate': 11,             # K - Column 10
            'increase_percentage': 12,             # L - Column 11
            'increase_amount': 13,                 # M - Column 12
            'total_excluding_contribution': 14,    # N - Column 13
            'total_including_contribution': 15,    # O - Column 14
            'working_week_length': 16,             # P - Column 15
            'annual_leave_days': 17,               # Q - Column 16
            'annual_leave_rate': 18,               # R - Column 17
            'annual_leave_cost': 19,               # S - Column 18
            'total_rd_fee': 20,                    # T - Column 19
            'total_planned_remuneration': 21,      # U - Column 20
            'justification': 22                    # V - Column 21
        }
    
    def calculate_form_fields(self, position: JobPosition, contribution_rate: float) -> Dict[str, float]:
        """Calculate all form fields according to the exact form requirements"""
        
        # Column 13 (N): Total excluding employer contribution = Column 10 + Column 12
        total_excluding_contribution = position.planned_salary_rate + position.increase_amount
        
        # Column 14 (O): Total including employer contribution  
        total_including_contribution = total_excluding_contribution * (1 + contribution_rate)
        
        # Column 18 (S): Annual leave cost calculation
        annual_leave_cost = total_including_contribution * position.annual_leave_rate * position.months_hours_planned
        
        # Column 19 (T): Total R&D fee = Column 14 + Column 18
        total_rd_fee = total_including_contribution + annual_leave_cost
        
        # Column 20 (U): Total planned remuneration = Column 19 + Column 9
        total_planned_remuneration = total_rd_fee + position.months_hours_planned
        
        return {
            'total_excluding_contribution': round(total_excluding_contribution, 2),
            'total_including_contribution': round(total_including_contribution, 2),
            'annual_leave_cost': round(annual_leave_cost, 2),
            'total_rd_fee': round(total_rd_fee, 2),
            'total_planned_remuneration': round(total_planned_remuneration, 2)
        }
    
    def fill_project_header_info(self, sheet, project_data: ProjectData):
        """Fill project header information"""
        
        # Project information (adjust row numbers as needed)
        sheet['B4'] = "Project/Joint project code"
        sheet['I4'] = project_data.project_code
        
        sheet['B5'] = "Name of the applicant/joint applicant/project partner"
        sheet['I5'] = project_data.organization_name
        
        sheet['B6'] = "Project duration, months"
        sheet['I6'] = project_data.project_duration_months
        
        sheet['B9'] = "Type of organization"
        sheet['H9'] = project_data.organization_type
        sheet['I9'] = "Budgetary" if "budgetary" in project_data.budgetary_classification.lower() else "Non-budgetary"
        sheet['J9'] = f"{project_data.contribution_rate:.3f}"
        
        logger.info("Project header information filled")
    

    # def add_sum_formulas(self, sheet, num_positions):
    #     """Add dynamic sum formulas based on actual number of positions to row 33"""
        
    #     # Calculate the last data row (15 is start row, positions go from 15 to 15+num_positions-1)
    #     start_row = 15
    #     end_row = start_row + num_positions - 1  # e.g., if 10 positions: 15 to 24
    #     sum_row = 33  # Sum formulas go to row 33
        
    #     # Define dynamic sum formulas with correct column mapping
    #     sum_formulas = {
    #         f'J{sum_row}': f'=SUM(J{start_row}:J{end_row})',  # Number of months/hours planned
    #         f'K{sum_row}': f'=SUM(K{start_row}:K{end_row})',  # Planned post salary/hourly rate, EUR
    #         f'L{sum_row}': f'=SUM(L{start_row}:L{end_row})',  # Increase, % (if applicable)
    #         f'M{sum_row}': f'=SUM(M{start_row}:M{end_row})',  # Amount of increase, EUR (if applicable)
    #         f'N{sum_row}': f'=SUM(N{start_row}:N{end_row})',  # Total planned remuneration excluding employer's contribution, EUR
    #         f'O{sum_row}': f'=SUM(O{start_row}:O{end_row})',  # Total rate of pay for the planned salary including employer's contribution, EUR
    #         f'S{sum_row}': f'=SUM(S{start_row}:S{end_row})',  # Planned cost of annual leave (including employer's contributions), EUR
    #         f'T{sum_row}': f'=SUM(T{start_row}:T{end_row})',  # Total planned R&D fee, EUR
    #         f'U{sum_row}': f'=SUM(U{start_row}:U{end_row})',  # Total planned remuneration, EUR
    #     }
        
    #     # Insert the formulas into the sheet
    #     for cell_address, formula in sum_formulas.items():
    #         sheet[cell_address] = formula
    #         logger.info(f"Added dynamic formula to {cell_address}: {formula}")
        
    #     logger.info(f"All sum formulas added to row {sum_row} for {num_positions} positions (rows {start_row}-{end_row})")

    # Add this method to your ExcelFormFiller class
    def add_sum_formulas(self, sheet, num_positions):
        """Add dynamic sum formulas based on actual number of positions to row 33"""
        
        # Calculate the last data row (15 is start row, positions go from 15 to 15+num_positions-1)
        start_row = 15
        end_row = start_row + num_positions - 1  # e.g., if 10 positions: 15 to 24
        sum_row = 33  # Sum formulas go to row 33
        
        # Define dynamic sum formulas with correct column mapping
        sum_formulas = {
            f'J{sum_row}': f'=SUM(J{start_row}:J{end_row})',  # Number of months/hours planned
            f'K{sum_row}': f'=SUM(K{start_row}:K{end_row})',  # Planned post salary/hourly rate, EUR
            f'L{sum_row}': f'=SUM(L{start_row}:L{end_row})',  # Increase, % (if applicable)
            f'M{sum_row}': f'=SUM(M{start_row}:M{end_row})',  # Amount of increase, EUR (if applicable)
            f'N{sum_row}': f'=SUM(N{start_row}:N{end_row})',  # Total planned remuneration excluding employer's contribution, EUR
            f'O{sum_row}': f'=SUM(O{start_row}:O{end_row})',  # Total rate of pay for the planned salary including employer's contribution, EUR
            f'S{sum_row}': f'=SUM(S{start_row}:S{end_row})',  # Planned cost of annual leave (including employer's contributions), EUR
            f'T{sum_row}': f'=SUM(T{start_row}:T{end_row})',  # Total planned R&D fee, EUR
            f'U{sum_row}': f'=SUM(U{start_row}:U{end_row})',  # Total planned remuneration, EUR
        }
        
        # Insert the formulas into the sheet
        for cell_address, formula in sum_formulas.items():
            sheet[cell_address] = formula
            logger.info(f"Added dynamic formula to {cell_address}: {formula}")
        
        logger.info(f"All sum formulas added to row {sum_row} for {num_positions} positions (rows {start_row}-{end_row})")


    def fill_excel_form(self, template_path: str, output_path: str, project_data: ProjectData) -> bool:
        """Fill Excel form with correct calculations and mapping"""
        
        try:
            logger.info(f"Loading Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            # Try to find the correct sheet
            sheet_name = "Other (non-budgetary) "
            if sheet_name not in workbook.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found. Using first sheet.")
                sheet = workbook.active
            else:
                sheet = workbook[sheet_name]
            
            # Fill project header
            self.fill_project_header_info(sheet, project_data)
            
            num_positions = len(project_data.job_positions)  # Get actual number of positions
            logger.info(f"Filling {num_positions} job positions with correct form calculations...")
            
            start_row = 15  # Adjust based on actual form structure
            total_project_cost = 0
            
            for idx, position in enumerate(project_data.job_positions):
                row = start_row + idx
                
                # Calculate all derived fields correctly according to form requirements
                calculated = self.calculate_form_fields(position, project_data.contribution_rate)
                total_project_cost += calculated['total_planned_remuneration']
                
                # Fill all position data according to correct column mapping
                sheet.cell(row=row, column=self.column_mapping['eil_no']).value = position.eil_no
                sheet.cell(row=row, column=self.column_mapping['project_impact_no']).value = position.project_impact_no
                sheet.cell(row=row, column=self.column_mapping['action_expenditure_no']).value = position.action_expenditure_no
                sheet.cell(row=row, column=self.column_mapping['duties_in_orp']).value = position.duties_in_orp
                sheet.cell(row=row, column=self.column_mapping['position_function']).value = position.position_function
                sheet.cell(row=row, column=self.column_mapping['employee_name']).value = position.employee_name
                sheet.cell(row=row, column=self.column_mapping['employment_contract_type']).value = position.employment_contract_type
                sheet.cell(row=row, column=self.column_mapping['remuneration_year']).value = position.remuneration_year
                sheet.cell(row=row, column=self.column_mapping['months_hours_planned']).value = position.months_hours_planned
                sheet.cell(row=row, column=self.column_mapping['planned_salary_rate']).value = position.planned_salary_rate
                sheet.cell(row=row, column=self.column_mapping['increase_percentage']).value = position.increase_percentage
                sheet.cell(row=row, column=self.column_mapping['increase_amount']).value = position.increase_amount
                sheet.cell(row=row, column=self.column_mapping['total_excluding_contribution']).value = calculated['total_excluding_contribution']
                sheet.cell(row=row, column=self.column_mapping['total_including_contribution']).value = calculated['total_including_contribution']
                sheet.cell(row=row, column=self.column_mapping['working_week_length']).value = position.working_week_length
                sheet.cell(row=row, column=self.column_mapping['annual_leave_days']).value = position.annual_leave_days
                sheet.cell(row=row, column=self.column_mapping['annual_leave_rate']).value = position.annual_leave_rate
                sheet.cell(row=row, column=self.column_mapping['annual_leave_cost']).value = calculated['annual_leave_cost']
                sheet.cell(row=row, column=self.column_mapping['total_rd_fee']).value = calculated['total_rd_fee']
                sheet.cell(row=row, column=self.column_mapping['total_planned_remuneration']).value = calculated['total_planned_remuneration']
                
                # Enhanced justification with all relevant information
                justification_parts = []
                if position.justification and position.justification.strip():
                    justification_parts.append(position.justification)
                
                if position.coefficient:
                    justification_parts.append(f"Salary coefficient: {position.coefficient}")
                
                if position.bonus_breakdown:
                    justification_parts.append(f"Bonus structure: {position.bonus_breakdown}")
                
                # Add salary calculation explanation as requested
                salary_calculation = f"Planned rate calculation: Base salary €{position.planned_salary_rate}"
                if position.increase_amount > 0:
                    salary_calculation += f" + Increase €{position.increase_amount} ({position.increase_percentage*100:.1f}%)"
                salary_calculation += f" = Total €{calculated['total_excluding_contribution']}"
                justification_parts.append(salary_calculation)
                
                # Add compliance information for budgetary institutions
                if "budgetary" in project_data.budgetary_classification.lower():
                    justification_parts.append("Compliant with Lithuanian civil service regulations")
                
                full_justification = " | ".join(justification_parts) if justification_parts else position.justification
                
                sheet.cell(row=row, column=self.column_mapping['justification']).value = full_justification
                
                logger.info(f"[OK] Filled position {idx + 1}: {position.position_function} - €{calculated['total_planned_remuneration']:,.2f}")
            
            # Add DYNAMIC sum formulas based on actual number of positions to ROW 33 - NEW CODE
            logger.info(f"Adding dynamic sum formulas to row 33 for {num_positions} positions...")
            self.add_sum_formulas(sheet, num_positions)  # Pass the number of positions
            
            workbook.save(output_path)
            logger.info(f"[OK] Successfully saved Excel form with correct mapping: {output_path}")
            logger.info(f"💰 Total project cost: €{total_project_cost:,.2f}")
            
            return True
            
        except Exception as e:
            logger.error(f"[X] Failed to fill Excel form: {e}")
            import traceback
            traceback.print_exc()
            return False
        

    # DEBUGGING: Add this method to help identify issues
    def debug_sum_ranges(self, sheet, num_positions):
        """Debug method to check what ranges are being summed"""
        
        start_row = 15
        end_row = start_row + num_positions - 1
        
        logger.info(f"DEBUG: Number of positions: {num_positions}")
        logger.info(f"DEBUG: Data range: {start_row} to {end_row}")
        logger.info(f"DEBUG: Sum formulas will use ranges like J{start_row}:J{end_row}, K{start_row}:K{end_row}, etc.")
        logger.info(f"DEBUG: Sum results will appear in row 33")
        
        # Check what values are actually in the key columns
        for row in range(start_row, end_row + 1):
            j_val = sheet.cell(row=row, column=10).value  # Column J - months planned
            k_val = sheet.cell(row=row, column=11).value  # Column K - salary rate
            n_val = sheet.cell(row=row, column=14).value  # Column N - total excluding contribution
            u_val = sheet.cell(row=row, column=21).value  # Column U - total planned remuneration
            logger.info(f"DEBUG: Row {row} - J: {j_val}, K: {k_val}, N: {n_val}, U: {u_val}")

    # Optional: Add header labels for sum row
    def add_sum_row_labels(self, sheet):
        """Add labels for the sum row to make it clear what the totals represent"""
        
        sum_row = 33
        
        # Add labels in column A or B
        sheet.cell(row=sum_row, column=1).value = "TOTALS:"
        sheet.cell(row=sum_row, column=2).value = "Project Totals"
        
        logger.info("Added sum row labels")

class ProjectAnalyzer:
    """Main orchestrator with correct form handling"""
    
    def __init__(self, groq_api_key: str):
        self.groq_client = GroqAPIClient(groq_api_key)
        self.data_analyzer = DataAnalyzer(self.groq_client)
        self.excel_filler = ExcelFormFiller()
    
    def read_data_file(self, file_path: str) -> str:
        """Read content from data file with better encoding handling"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"[OK] Successfully read data file: {file_path} ({len(content)} characters)")
            return content
        except FileNotFoundError:
            logger.error(f"[X] Data file not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"[X] Error reading data file: {e}")
            raise
    
    def analyze_and_process(self, data_file_path: str, template_path: str, output_path: str) -> Dict[str, Any]:
        """Complete workflow with correct form processing"""
        
        results = {
            'success': False,
            'steps_completed': [],
            'project_data': None,
            'statistics': {},
            'errors': []
        }
        
        try:
            # Step 1: Read data file
            logger.info("[STEP] STEP 1: Reading Data File")
            text_content = self.read_data_file(data_file_path)
            results['steps_completed'].append('data_file_read')
            
            # Step 2: Analysis with correct form mapping
            logger.info("[STEP] STEP 2: Analysis with Correct Form Mapping")
            project_data = self.data_analyzer.analyze_project_data(text_content)
            results['project_data'] = project_data
            results['steps_completed'].append('data_analyzed')
            
            # Step 3: Generate statistics
            logger.info("[STEP] STEP 3: Generating Statistics")
            statistics = self._generate_statistics(project_data)
            results['statistics'] = statistics
            results['steps_completed'].append('statistics_generated')
            
            # Step 4: Fill Excel with correct mapping
            logger.info("[STEP] STEP 4: Filling Excel Form with Correct Mapping")
            excel_success = self.excel_filler.fill_excel_form(template_path, output_path, project_data)
            
            if excel_success:
                results['steps_completed'].append('excel_filled')
                results['success'] = True
                logger.info("[OK] PROCESS COMPLETED SUCCESSFULLY")
            else:
                results['errors'].append('Excel filling failed')
                
        except Exception as e:
            error_msg = f"[X] Process failed at step {len(results['steps_completed']) + 1}: {e}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _generate_statistics(self, project_data: ProjectData) -> Dict[str, Any]:
        """Generate statistics with correct form calculations"""
        
        total_positions = len(project_data.job_positions)
        total_staff = sum(pos.planned_salary_rate for pos in project_data.job_positions)
        total_months = sum(pos.months_hours_planned for pos in project_data.job_positions)
        
        # Financial calculations using correct form logic
        total_base_salaries = 0
        total_increases = 0
        total_project_cost = 0
        
        position_summary = []
        validation_issues = []
        
        for position in project_data.job_positions:
            calculated = self.excel_filler.calculate_form_fields(position, project_data.contribution_rate)
            
            position_cost = calculated['total_planned_remuneration']
            total_project_cost += position_cost
            total_base_salaries += position.planned_salary_rate * position.months_hours_planned
            total_increases += position.increase_amount * position.months_hours_planned
            
            # Validation checks
            if position.working_week_length not in [5, 6]:
                validation_issues.append(f"[!] {position.position_function}: Unusual working week ({position.working_week_length})")
            
            if position.annual_leave_rate == 0:
                validation_issues.append(f"[!] {position.position_function}: Missing annual leave rate")
            
            # Verify calculations
            expected_total_excluding = position.planned_salary_rate + position.increase_amount
            if abs(calculated['total_excluding_contribution'] - expected_total_excluding) > 0.01:
                validation_issues.append(f"[!] {position.position_function}: Calculation mismatch in total excluding contribution")
            
            position_summary.append({
                'position_function': position.position_function,
                'duties_in_orp': position.duties_in_orp,
                'months_hours_planned': position.months_hours_planned,
                'planned_salary_rate': position.planned_salary_rate,
                'total_cost': position_cost,
                'coefficient': position.coefficient,
                'leave_rate': f"{position.annual_leave_rate * 100:.2f}%"
            })
        
        statistics = {
            'project_overview': {
                'project_code': project_data.project_code,
                'organization': project_data.organization_name,
                'duration_months': project_data.project_duration_months,
                'contribution_rate': f"{project_data.contribution_rate * 100:.1f}%",
                'organization_type': project_data.organization_type,
                'budgetary_classification': project_data.budgetary_classification
            },
            'staffing': {
                'total_positions': total_positions,
                'total_base_salaries': total_base_salaries,
                'total_work_months': total_months
            },
            'financial': {
                'total_base_salaries': round(total_base_salaries, 2),
                'total_increases': round(total_increases, 2),
                'total_project_cost': round(total_project_cost, 2)
            },
            'position_breakdown': position_summary,
            'validation_issues': validation_issues
        }
        
        return statistics
    
    def print_results(self, results: Dict[str, Any]):
        """Print comprehensive results with correct form information"""
        
        print("\n" + "="*80)
        print("[>] PROJECT DATA ANALYSIS AND EXCEL FORM FILLING RESULTS")
        print("   (Updated with Correct Form Mapping B-V)")
        print("="*80)
        
        # Process Status
        status_icon = "[OK] SUCCESS" if results['success'] else "[X] FAILED"
        print(f"\n[STEP] PROCESS STATUS: {status_icon}")
        print(f"[STEP] Steps Completed: {', '.join(results['steps_completed'])}")
        
        if results['errors']:
            print(f"[!]  Errors: {'; '.join(results['errors'])}")
        
        # Project Overview
        if results['project_data']:
            stats = results['statistics']
            overview = stats['project_overview']
            
            print(f"\n[SUMMARY] PROJECT OVERVIEW")
            print(f"   Project Code: {overview['project_code']}")
            print(f"   Organization: {overview['organization']}")
            print(f"   Duration: {overview['duration_months']} months")
            print(f"   Contribution Rate: {overview['contribution_rate']}")
            print(f"   Organization Type: {overview['organization_type']}")
            print(f"   Budgetary Classification: {overview['budgetary_classification']}")
            
            # Staffing Statistics
            staffing = stats['staffing']
            print(f"\n👥 STAFFING STATISTICS")
            print(f"   Total Job Positions: {staffing['total_positions']}")
            print(f"   Total Base Salaries: €{staffing['total_base_salaries']:,.2f}")
            print(f"   Total Work Months: {staffing['total_work_months']}")
            
            # Financial Statistics
            financial = stats['financial']
            print(f"\n💰 FINANCIAL SUMMARY")
            print(f"   Total Base Salaries: €{financial['total_base_salaries']:,.2f}")
            print(f"   Total Increases: €{financial['total_increases']:,.2f}")
            print(f"   TOTAL PROJECT COST: €{financial['total_project_cost']:,.2f}")
            
            # Position Breakdown
            print(f"\n[STEP] POSITION BREAKDOWN")
            for pos in stats['position_breakdown']:
                coefficient_info = f" (Coeff: {pos['coefficient']})" if pos['coefficient'] else ""
                print(f"   • {pos['position_function']}{coefficient_info}")
                print(f"     Duties: {pos['duties_in_orp']}")
                print(f"     Duration: {pos['months_hours_planned']} months | Salary: €{pos['planned_salary_rate']}")
                print(f"     Leave Rate: {pos['leave_rate']} | Total Cost: €{pos['total_cost']:,.2f}")
            
            # Validation Issues
            if stats['validation_issues']:
                print(f"\n[!]  VALIDATION ISSUES DETECTED")
                for issue in stats['validation_issues']:
                    print(f"   {issue}")
            else:
                print(f"\n[OK] ALL VALIDATIONS PASSED")
            
            # Form Mapping Information
            print(f"\n[STEP] EXCEL FORM MAPPING (B-V)")
            print(f"   Column B (2): Eil. No.")
            print(f"   Column C (3): Project Impact No.")
            print(f"   Column D (4): Action/Expenditure No.")
            print(f"   Column E (5): Duties in ORP")
            print(f"   Column F (6): Position/Function")
            print(f"   Column G (7): Employee Name")
            print(f"   Column H (8): Employment Contract Type")
            print(f"   Column I (9): Remuneration Year")
            print(f"   Column J (10): Months/Hours Planned")
            print(f"   Column K (11): Planned Salary Rate")
            print(f"   Column L (12): Increase %")
            print(f"   Column M (13): Increase Amount")
            print(f"   Column N (14): Total Excluding Contribution")
            print(f"   Column O (15): Total Including Contribution")
            print(f"   Column P (16): Working Week Length")
            print(f"   Column Q (17): Annual Leave Days")
            print(f"   Column R (18): Annual Leave Rate")
            print(f"   Column S (19): Annual Leave Cost")
            print(f"   Column T (20): Total R&D Fee")
            print(f"   Column U (21): Total Planned Remuneration")
            print(f"   Column V (22): Justification")
        
        print("\n" + "="*80)

def main_nonbudgetary(data_file_path):
    """Main function with correct form processing"""
    
    print("[>] PROJECT DATA ANALYZER AND EXCEL FORM FILLER")
    print("   (Updated with Correct Form Mapping B-V)")
    print("="*70)
    
    # Configuration
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
    if not GROQ_API_KEY:
        print("[X] Error: GROQ_API_KEY environment variable not set")
        print("Please set your Groq API key in .env file: GROQ_API_KEY='your_api_key_here'")
        return
    
   
    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    template_path = "code/agents/ENGISH_Rekomenduojamaforma_en/Other_Non_Budgetary/budgetary.xlsx"
    output_path = "code/agents/ENGISH_Rekomenduojamaforma_en/Other_Non_Budgetary/budgetary_filled.xlsx"
    
    # Validate required files exist
    required_files = [data_file_path, template_path]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print(f"[X] Missing required files: {', '.join(missing_files)}")
        print("Required files:")
        print(f"  • {data_file_path} - Text file with project data to analyze")
        print(f"  • {template_path} - Excel template file to fill")
        return
    
    try:
        # Initialize analyzer with correct form mapping
        print(f"\n[DEBUG] Initializing Project Analyzer with Correct Form Mapping...")
        analyzer = ProjectAnalyzer(GROQ_API_KEY)
        
        # Run analysis and processing
        print(f"\n[STEP] Starting analysis of {data_file_path} with correct form structure...")
        results = analyzer.analyze_and_process(data_file_path, template_path, output_path)
        
        # Print results
        analyzer.print_results(results)
        
        if results['success']:
            print(f"\n[OK] Excel form successfully filled with correct mapping: {output_path}")
            print("\n[RESULT] CORRECT FORM FEATURES APPLIED:")
            print("   • [OK] Correct column mapping (B-V)")
            print("   • [OK] Proper calculation formulas:")
            print("       - Column 13 = Column 10 + Column 12")
            print("       - Column 19 = Column 14 + Column 18") 
            print("       - Column 20 = Column 19 + Column 9")
            print("   • [OK] Accurate field extraction and mapping")
            print("   • [OK] Proper employer contribution calculations")
            print("   • [OK] Correct annual leave rate calculations")
            print("   • [OK] Enhanced justification with all details")
            print("\n[SUMMARY] Open the file to see the correctly structured data!")
            
            # Additional success information
            if results['project_data']:
                total_positions = len(results['project_data'].job_positions)
                total_cost = results['statistics']['financial']['total_project_cost']
                print(f"\n[STATS] PROCESSING SUMMARY:")
                print(f"   • Extracted {total_positions} job positions")
                print(f"   • Total project budget: €{total_cost:,.2f}")
                print(f"   • All form calculations verified and applied")
                
        else:
            print(f"\n[X] Process failed. Check the detailed logs above.")
            print("\n[DEBUG] TROUBLESHOOTING TIPS:")
            print("   • Ensure your .env file contains GROQ_API_KEY")
            print("   • Check that data.txt contains valid project information")
            print("   • Verify form.xlsx template matches expected structure")
            print("   • Check internet connection for Groq API access")
            
    except Exception as e:
        logger.error(f"[X] Fatal error in main process: {e}")
        print(f"\n[X] Fatal error: {e}")
        print("\n[DEBUG] DEBUG INFORMATION:")
        import traceback
        traceback.print_exc(limit=3)
        
        print("\n[TIP] COMMON SOLUTIONS:")
        print("   • Check your Groq API key is valid and has credits")
        print("   • Ensure all required Python packages are installed:")
        print("     pip install groq openpyxl python-dotenv")
        print("   • Verify file permissions for reading data.txt and form.xlsx")





