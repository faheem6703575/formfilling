import openpyxl
import json
import os
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from dotenv import load_dotenv
from groq import Groq

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class ProductSalesData:
    """Data class for individual product sales information"""
    product_name: str
    sales_quantity_during: int
    sales_quantity_n1: int
    sales_quantity_n2: int
    sales_quantity_n3: int
    unit_price_during: float
    unit_price_n1: float
    unit_price_n2: float
    unit_price_n3: float
    total_revenue_during: float
    total_revenue_n1: float
    total_revenue_n2: float
    total_revenue_n3: float

@dataclass
class RevenueProjectionData:
    """Complete revenue projection data structure"""
    total_project_income_during: float
    total_project_income_n1: float
    total_project_income_n2: float
    total_project_income_n3: float
    products: List[ProductSalesData]
    
    def __post_init__(self):
        # Ensure maximum 16 products
        self.products = self.products[:16]

class GroqRevenueAnalyzer:
    """Analyzes revenue projection data using Groq LLM with streaming"""
    
    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key)
    
    def create_analysis_prompt(self, prompt_path: str = "code/prompts/revenue_forecast.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")
   
    def analyze_revenue_data_streaming(self, text_content: str) -> RevenueProjectionData:
        """Analyze text using Groq streaming and extract revenue projection data"""
        
        logger.info("Starting revenue projection data analysis with streaming...")
        
        prompt = self.create_analysis_prompt()
        full_content = f"{prompt}\n\nText to analyze:\n{text_content}"
        
        try:
            # Use streaming as specified in requirements
            completion = self.client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[
                    {
                        "role": "user",
                        "content": full_content
                    }
                ],
                temperature=0.3,
                stream=True  
            )
            
            # Collect streaming response
            analysis_result = ""
            logger.info("Receiving streaming response from Groq API...")
            
            for chunk in completion:
                if chunk.choices[0].delta.content:
                    content = chunk.choices[0].delta.content
                    analysis_result += content
            
            logger.info(f"Received complete analysis ({len(analysis_result)} characters)")
            
            # Clean and parse JSON
            json_content = self._extract_json_from_response(analysis_result)
            parsed_data = json.loads(json_content)
            
            # Convert to RevenueProjectionData object
            revenue_data = self._convert_to_revenue_data(parsed_data)
            
            return revenue_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            logger.error(f"Raw response: {analysis_result[:500]}...")
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_json_from_response(self, response: str) -> str:
        """Extract JSON from response, handling markdown code blocks"""
        
        response = response.strip()
        
        # Remove markdown code blocks if present
        if '```json' in response:
            start_marker = '```json'
            end_marker = '```'
            start_index = response.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = response.find(end_marker, start_index)
                if end_index != -1:
                    return response[start_index:end_index].strip()
        
        if response.startswith('```') and response.endswith('```'):
            return response[3:-3].strip()
        
        return response
    
    def _convert_to_revenue_data(self, parsed_data: Dict) -> RevenueProjectionData:
        """Convert parsed JSON to RevenueProjectionData object"""
        
        products = []
        for product_data in parsed_data.get('products', [])[:16]:  # Max 16 products
            # Calculate total revenues if not provided
            total_revenue_during = product_data.get('total_revenue_during', 
                product_data.get('sales_quantity_during', 0) * product_data.get('unit_price_during', 0))
            total_revenue_n1 = product_data.get('total_revenue_n1',
                product_data.get('sales_quantity_n1', 0) * product_data.get('unit_price_n1', 0))
            total_revenue_n2 = product_data.get('total_revenue_n2',
                product_data.get('sales_quantity_n2', 0) * product_data.get('unit_price_n2', 0))
            total_revenue_n3 = product_data.get('total_revenue_n3',
                product_data.get('sales_quantity_n3', 0) * product_data.get('unit_price_n3', 0))
            
            product = ProductSalesData(
                product_name=product_data.get('product_name', 'Product'),
                sales_quantity_during=int(product_data.get('sales_quantity_during', 0)),
                sales_quantity_n1=int(product_data.get('sales_quantity_n1', 0)),
                sales_quantity_n2=int(product_data.get('sales_quantity_n2', 0)),
                sales_quantity_n3=int(product_data.get('sales_quantity_n3', 0)),
                unit_price_during=float(product_data.get('unit_price_during', 0.0)),
                unit_price_n1=float(product_data.get('unit_price_n1', 0.0)),
                unit_price_n2=float(product_data.get('unit_price_n2', 0.0)),
                unit_price_n3=float(product_data.get('unit_price_n3', 0.0)),
                total_revenue_during=float(total_revenue_during),
                total_revenue_n1=float(total_revenue_n1),
                total_revenue_n2=float(total_revenue_n2),
                total_revenue_n3=float(total_revenue_n3)
            )
            products.append(product)
        
        revenue_data = RevenueProjectionData(
            total_project_income_during=float(parsed_data.get('total_project_income_during', 0.0)),
            total_project_income_n1=float(parsed_data.get('total_project_income_n1', 0.0)),
            total_project_income_n2=float(parsed_data.get('total_project_income_n2', 0.0)),
            total_project_income_n3=float(parsed_data.get('total_project_income_n3', 0.0)),
            products=products
        )
        
        return revenue_data

class RevenueProjectionExcelFiller:
    """Fills Excel form with revenue projection data according to exact cell mappings"""
    
    def __init__(self):
        # Product row mappings for each product (16 products max)
        self.product_mappings = [
            {'name_cell': 'A6', 'quantity_row': 7, 'price_row': 8, 'revenue_row': 9},      # Product 1
            {'name_cell': 'A11', 'quantity_row': 12, 'price_row': 13, 'revenue_row': 14},  # Product 2
            {'name_cell': 'A16', 'quantity_row': 17, 'price_row': 18, 'revenue_row': 19},  # Product 3
            {'name_cell': 'A21', 'quantity_row': 22, 'price_row': 23, 'revenue_row': 24},  # Product 4
            {'name_cell': 'A26', 'quantity_row': 27, 'price_row': 28, 'revenue_row': 29},  # Product 5
            {'name_cell': 'A31', 'quantity_row': 32, 'price_row': 33, 'revenue_row': 34},  # Product 6
            {'name_cell': 'A36', 'quantity_row': 37, 'price_row': 38, 'revenue_row': 39},  # Product 7
            {'name_cell': 'A41', 'quantity_row': 42, 'price_row': 43, 'revenue_row': 44},  # Product 8
            {'name_cell': 'A46', 'quantity_row': 47, 'price_row': 48, 'revenue_row': 49},  # Product 9
            {'name_cell': 'A51', 'quantity_row': 52, 'price_row': 53, 'revenue_row': 54},  # Product 10
            {'name_cell': 'A56', 'quantity_row': 57, 'price_row': 58, 'revenue_row': 59},  # Product 11
            {'name_cell': 'A61', 'quantity_row': 62, 'price_row': 63, 'revenue_row': 64},  # Product 12
            {'name_cell': 'A66', 'quantity_row': 67, 'price_row': 68, 'revenue_row': 69},  # Product 13
            {'name_cell': 'A71', 'quantity_row': 72, 'price_row': 73, 'revenue_row': 74},  # Product 14
            {'name_cell': 'A76', 'quantity_row': 77, 'price_row': 78, 'revenue_row': 79},  # Product 15
            {'name_cell': 'A81', 'quantity_row': 82, 'price_row': 83, 'revenue_row': 84},  # Product 16
        ]
        
        # Time period columns
        self.time_columns = {
            'during': 'B',  # During project
            'n1': 'C',      # N+1 year
            'n2': 'D',      # N+2 year
            'n3': 'E'       # N+3 year
        }
    
    def fill_excel_form(self, template_path: str, output_path: str, revenue_data: RevenueProjectionData) -> bool:
        """Fill Excel form with revenue projection data"""
        
        try:
            logger.info(f"Loading Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            # Use first sheet or find revenue-related sheet
            sheet_name = workbook.sheetnames[0]
            for name in workbook.sheetnames:
                if any(keyword in name.lower() for keyword in ['revenue', 'financial', 'sales', 'projection']):
                    sheet_name = name
                    break
            
            sheet = workbook[sheet_name]
            logger.info(f"Using sheet: {sheet_name}")
            
            # Fill Table 1: Total Project Income
            self._fill_total_income(sheet, revenue_data)
            
            # Fill Table 2: Individual Product Sales Data
            self._fill_product_sales(sheet, revenue_data.products)
            
            # Save the filled workbook
            workbook.save(output_path)
            logger.info(f"✅ Successfully saved Excel form: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to fill Excel form: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _fill_total_income(self, sheet, revenue_data: RevenueProjectionData):
        """Fill Table 1: Total Project Income"""
        
        logger.info("Filling Table 1: Total Project Income...")
        
        # Total project income row (B3, C3, D3, E3)
        sheet['B4'] = revenue_data.total_project_income_during
        sheet['C4'] = revenue_data.total_project_income_n1
        sheet['D4'] = revenue_data.total_project_income_n2
        sheet['E4'] = revenue_data.total_project_income_n3
        
        logger.info(f"✅ Total Income - During: €{revenue_data.total_project_income_during:,.2f}")
        logger.info(f"✅ Total Income - N+1: €{revenue_data.total_project_income_n1:,.2f}")
        logger.info(f"✅ Total Income - N+2: €{revenue_data.total_project_income_n2:,.2f}")
        logger.info(f"✅ Total Income - N+3: €{revenue_data.total_project_income_n3:,.2f}")
    
    def _fill_product_sales(self, sheet, products: List[ProductSalesData]):
        """Fill Table 2: Individual Product Sales Data"""
        
        logger.info(f"Filling Table 2: {len(products)} Product Sales...")
        
        for i, product in enumerate(products):
            if i < len(self.product_mappings):
                mapping = self.product_mappings[i]
                
                # Fill product name
                sheet[mapping['name_cell']] = product.product_name
                
                # Fill sales quantities (B, C, D, E columns)
                quantity_row = mapping['quantity_row']
                sheet[f"B{quantity_row}"] = product.sales_quantity_during
                sheet[f"C{quantity_row}"] = product.sales_quantity_n1
                sheet[f"D{quantity_row}"] = product.sales_quantity_n2
                sheet[f"E{quantity_row}"] = product.sales_quantity_n3
                
                # Fill unit prices (B, C, D, E columns)
                price_row = mapping['price_row']
                sheet[f"B{price_row}"] = product.unit_price_during
                sheet[f"C{price_row}"] = product.unit_price_n1
                sheet[f"D{price_row}"] = product.unit_price_n2
                sheet[f"E{price_row}"] = product.unit_price_n3
                
                # Fill total revenues (B, C, D, E columns)
                revenue_row = mapping['revenue_row']
                sheet[f"B{revenue_row}"] = product.total_revenue_during
                sheet[f"C{revenue_row}"] = product.total_revenue_n1
                sheet[f"D{revenue_row}"] = product.total_revenue_n2
                sheet[f"E{revenue_row}"] = product.total_revenue_n3
                
                logger.info(f"✅ Product {i+1}: {product.product_name}")
                logger.info(f"   Revenue - During: €{product.total_revenue_during:,.2f}")
                logger.info(f"   Revenue - N+1: €{product.total_revenue_n1:,.2f}")
                logger.info(f"   Revenue - N+2: €{product.total_revenue_n2:,.2f}")
                logger.info(f"   Revenue - N+3: €{product.total_revenue_n3:,.2f}")

class RevenueProjectionAgent:
    """Main orchestrator for revenue projection processing with streaming"""
    
    def __init__(self, groq_api_key: str):
        self.analyzer = GroqRevenueAnalyzer(groq_api_key)
        self.excel_filler = RevenueProjectionExcelFiller()
    
    def read_data_file(self, file_path: str) -> str:
        """Read content from data file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"✅ Successfully read data file: {file_path} ({len(content)} characters)")
            return content
        except FileNotFoundError:
            logger.error(f"❌ Data file not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"❌ Error reading data file: {e}")
            raise
    
    def process_revenue_projections(self, data_file_path: str, template_path: str, output_path: str) -> Dict[str, Any]:
        """Complete workflow for processing revenue projections"""
        
        results = {
            'success': False,
            'steps_completed': [],
            'revenue_data': None,
            'statistics': {},
            'errors': []
        }
        
        try:
            # Step 1: Read data file
            logger.info("🔄 STEP 1: Reading Data File")
            text_content = self.read_data_file(data_file_path)
            results['steps_completed'].append('data_file_read')
            
            # Step 2: Analyze with Groq API (using streaming)
            logger.info("🔄 STEP 2: Analyzing Revenue Projections with Groq Streaming API")
            revenue_data = self.analyzer.analyze_revenue_data_streaming(text_content)
            results['revenue_data'] = revenue_data
            results['steps_completed'].append('data_analyzed')
            
            # Step 3: Generate statistics
            logger.info("🔄 STEP 3: Generating Statistics")
            statistics = self._generate_statistics(revenue_data)
            results['statistics'] = statistics
            results['steps_completed'].append('statistics_generated')
            
            # Step 4: Fill Excel form
            logger.info("🔄 STEP 4: Filling Excel Form")
            excel_success = self.excel_filler.fill_excel_form(template_path, output_path, revenue_data)
            
            if excel_success:
                results['steps_completed'].append('excel_filled')
                results['success'] = True
                logger.info("✅ PROCESS COMPLETED SUCCESSFULLY")
            else:
                results['errors'].append('Excel filling failed')
                
        except Exception as e:
            error_msg = f"❌ Process failed at step {len(results['steps_completed']) + 1}: {e}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _generate_statistics(self, revenue_data: RevenueProjectionData) -> Dict[str, Any]:
        """Generate comprehensive statistics"""
        
        total_products = len(revenue_data.products)
        
        # Calculate totals per period
        product_revenue_during = sum(p.total_revenue_during for p in revenue_data.products)
        product_revenue_n1 = sum(p.total_revenue_n1 for p in revenue_data.products)
        product_revenue_n2 = sum(p.total_revenue_n2 for p in revenue_data.products)
        product_revenue_n3 = sum(p.total_revenue_n3 for p in revenue_data.products)
        
        # Calculate total quantities per period
        total_quantity_during = sum(p.sales_quantity_during for p in revenue_data.products)
        total_quantity_n1 = sum(p.sales_quantity_n1 for p in revenue_data.products)
        total_quantity_n2 = sum(p.sales_quantity_n2 for p in revenue_data.products)
        total_quantity_n3 = sum(p.sales_quantity_n3 for p in revenue_data.products)
        
        # Calculate average unit prices per period
        avg_price_during = product_revenue_during / total_quantity_during if total_quantity_during > 0 else 0
        avg_price_n1 = product_revenue_n1 / total_quantity_n1 if total_quantity_n1 > 0 else 0
        avg_price_n2 = product_revenue_n2 / total_quantity_n2 if total_quantity_n2 > 0 else 0
        avg_price_n3 = product_revenue_n3 / total_quantity_n3 if total_quantity_n3 > 0 else 0
        
        # Growth rates
        revenue_growth_n1 = ((product_revenue_n1 - product_revenue_during) / product_revenue_during * 100) if product_revenue_during > 0 else 0
        revenue_growth_n2 = ((product_revenue_n2 - product_revenue_n1) / product_revenue_n1 * 100) if product_revenue_n1 > 0 else 0
        revenue_growth_n3 = ((product_revenue_n3 - product_revenue_n2) / product_revenue_n2 * 100) if product_revenue_n2 > 0 else 0
        
        statistics = {
            'project_summary': {
                'total_products': total_products,
                'total_project_income_during': revenue_data.total_project_income_during,
                'total_project_income_n1': revenue_data.total_project_income_n1,
                'total_project_income_n2': revenue_data.total_project_income_n2,
                'total_project_income_n3': revenue_data.total_project_income_n3,
                'total_revenue_4_years': (revenue_data.total_project_income_during + 
                                        revenue_data.total_project_income_n1 + 
                                        revenue_data.total_project_income_n2 + 
                                        revenue_data.total_project_income_n3)
            },
            'product_totals': {
                'total_quantity_during': total_quantity_during,
                'total_quantity_n1': total_quantity_n1,
                'total_quantity_n2': total_quantity_n2,
                'total_quantity_n3': total_quantity_n3,
                'product_revenue_during': product_revenue_during,
                'product_revenue_n1': product_revenue_n1,
                'product_revenue_n2': product_revenue_n2,
                'product_revenue_n3': product_revenue_n3
            },
            'averages': {
                'avg_price_during': avg_price_during,
                'avg_price_n1': avg_price_n1,
                'avg_price_n2': avg_price_n2,
                'avg_price_n3': avg_price_n3
            },
            'growth_rates': {
                'revenue_growth_n1': revenue_growth_n1,
                'revenue_growth_n2': revenue_growth_n2,
                'revenue_growth_n3': revenue_growth_n3
            },
            'top_products': sorted(revenue_data.products, 
                                 key=lambda p: p.total_revenue_during + p.total_revenue_n1 + p.total_revenue_n2 + p.total_revenue_n3, 
                                 reverse=True)[:5]
        }
        
        return statistics
    
    def print_results(self, results: Dict[str, Any]):
        """Print comprehensive results"""
        
        print("\n" + "="*80)
        print("🚀 REVENUE PROJECTION EXCEL FORM FILLING RESULTS (STREAMING)")
        print("="*80)
        
        # Process Status
        status_icon = "✅ SUCCESS" if results['success'] else "❌ FAILED"
        print(f"\n🔄 PROCESS STATUS: {status_icon}")
        print(f"📋 Steps Completed: {', '.join(results['steps_completed'])}")
        
        if results['errors']:
            print(f"⚠️  Errors: {'; '.join(results['errors'])}")
        
        # Statistics
        if results['revenue_data'] and results['statistics']:
            stats = results['statistics']
            
            print(f"\n📊 PROJECT SUMMARY")
            summary = stats['project_summary']
            print(f"   Total Products: {summary['total_products']}")
            print(f"   Total Revenue (4 years): €{summary['total_revenue_4_years']:,.2f}")
            print(f"   During Project: €{summary['total_project_income_during']:,.2f}")
            print(f"   Year N+1: €{summary['total_project_income_n1']:,.2f}")
            print(f"   Year N+2: €{summary['total_project_income_n2']:,.2f}")
            print(f"   Year N+3: €{summary['total_project_income_n3']:,.2f}")
            
            print(f"\n💰 PRODUCT TOTALS")
            totals = stats['product_totals']
            print(f"   During Project: {totals['total_quantity_during']:,} units → €{totals['product_revenue_during']:,.2f}")
            print(f"   Year N+1: {totals['total_quantity_n1']:,} units → €{totals['product_revenue_n1']:,.2f}")
            print(f"   Year N+2: {totals['total_quantity_n2']:,} units → €{totals['product_revenue_n2']:,.2f}")
            print(f"   Year N+3: {totals['total_quantity_n3']:,} units → €{totals['product_revenue_n3']:,.2f}")
            
            print(f"\n📈 GROWTH RATES")
            growth = stats['growth_rates']
            print(f"   Revenue Growth N+1: {growth['revenue_growth_n1']:+.1f}%")
            print(f"   Revenue Growth N+2: {growth['revenue_growth_n2']:+.1f}%")
            print(f"   Revenue Growth N+3: {growth['revenue_growth_n3']:+.1f}%")
            
            print(f"\n🏆 TOP 5 PRODUCTS (by total 4-year revenue)")
            for i, product in enumerate(stats['top_products'], 1):
                total_4yr = product.total_revenue_during + product.total_revenue_n1 + product.total_revenue_n2 + product.total_revenue_n3
                print(f"   {i}. {product.product_name}: €{total_4yr:,.2f}")
        
        print("\n" + "="*80)

def main_forecast(data_file_path):
    """Main function to run the revenue projection agent"""
    
    print("🚀 REVENUE PROJECTION EXCEL FORM FILLER (GROQ STREAMING)")
    print("="*65)
    
    # Configuration
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
    if not GROQ_API_KEY:
        print("❌ Error: GROQ_API_KEY environment variable not set")
        print("Please set your Groq API key in .env file: GROQ_API_KEY='your_api_key_here'")
        return
    
    
    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    template_path = "code/agents/ENGLISH_Finansinis_planas_en/revenue_forecast.xlsx"  # Your Excel template
    output_path = "code/output/ENGLISH_Finansinis planas en.xlsx"
    
    # Validate required files exist
    required_files = [data_file_path, template_path]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print(f"❌ Missing required files: {', '.join(missing_files)}")
        print("Required files:")
        print(f"  • {data_file_path} - Text file with revenue projection data")
        print(f"  • {template_path} - Excel template file (Revenue Forecast Form)")
        return
    
    try:
        # Initialize agent
        print(f"\n🔧 Initializing Revenue Projection Agent with Groq Streaming...")
        agent = RevenueProjectionAgent(GROQ_API_KEY)
        
        # Process revenue projections
        print(f"\n🔄 Processing revenue projections from {data_file_path}...")
        results = agent.process_revenue_projections(data_file_path, template_path, output_path)
        
        # Print results
        agent.print_results(results)
        
        if results['success']:
            print(f"\n✅ Revenue projection Excel form successfully filled: {output_path}")
            print(f"\n🎯 Key Features Implemented:")
            print(f"   • ✅ Groq streaming API with meta-llama/llama-4-scout-17b-16e-instruct")
            print(f"   • ✅ Total project income across 4 time periods")
            print(f"   • ✅ Up to 16 individual products with full sales data")
            print(f"   • ✅ Accurate cell mapping (B3:E3, A6-A81 products)")
            print(f"   • ✅ Automatic revenue calculations (quantity × price)")
            print(f"   • ✅ JSON validation and comprehensive error handling")
            print(f"   • ✅ Complete revenue projection workflow")
            
            if results['statistics']:
                total_revenue = results['statistics']['project_summary']['total_revenue_4_years']
                total_products = results['statistics']['project_summary']['total_products']
                print(f"\n📊 Project Overview:")
                print(f"   • Total 4-Year Revenue: €{total_revenue:,.2f}")
                print(f"   • Products Analyzed: {total_products}")
        else:
            print(f"\n❌ Process failed. Check the logs above for details.")
            
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        print(f"\n❌ Fatal error: {e}")
