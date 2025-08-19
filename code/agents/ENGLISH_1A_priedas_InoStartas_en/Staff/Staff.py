import openpyxl
import json
import os
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from dotenv import load_dotenv
from groq import Groq

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class StaffMember:
    """Data class for individual staff member information"""
    name_surname: str
    responsibilities: str
    total_hours: int
    hourly_rate: float
    cost_justification: str
    organization: str  # 'Applicant', 'Partner No 1', 'Partner No 2', 'Partner No 3'

@dataclass
class StaffWagesData:
    """Complete staff wages data structure"""
    applicant_staff: List[StaffMember]      # Up to 20 staff members (C7:G26)
    partner1_staff: List[StaffMember]       # Up to 20 staff members (C27:G46)
    partner2_staff: List[StaffMember]       # Up to 20 staff members (C47:G66)
    partner3_staff: List[StaffMember]       # Up to 20 staff members (C67:G86)
    
    def __post_init__(self):
        # Ensure maximum 20 staff members per organization
        self.applicant_staff = self.applicant_staff[:20]
        self.partner1_staff = self.partner1_staff[:20]
        self.partner2_staff = self.partner2_staff[:20]
        self.partner3_staff = self.partner3_staff[:20]

class GroqStaffAnalyzer:
    """Analyzes staff wages data using Groq LLM with streaming"""
    
    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key)
    
    def create_analysis_prompt(self, prompt_path: str = "code/prompts/staff.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")
    
    def analyze_staff_data_streaming(self, text_content: str) -> StaffWagesData:
        """Analyze text using Groq streaming and extract staff wages data"""
        
        logger.info("Starting staff wages data analysis with streaming...")
        
        prompt = self.create_analysis_prompt()
        full_content = f"{prompt}\n\nText to analyze:\n{text_content}"
        
        try:
            # Use streaming as specified in requirements
            completion = self.client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[
                    {
                        "role": "user",
                        "content": full_content
                    }
                ],
                temperature=0.3,
                stream=True  
            )
            
            # Collect streaming response
            analysis_result = ""
            logger.info("Receiving streaming response from Groq API...")
            
            for chunk in completion:
                if chunk.choices[0].delta.content:
                    content = chunk.choices[0].delta.content
                    analysis_result += content
            
            logger.info(f"Received complete analysis ({len(analysis_result)} characters)")
            
            # Clean and parse JSON
            json_content = self._extract_json_from_response(analysis_result)
            parsed_data = json.loads(json_content)
            
            # Convert to StaffWagesData object
            staff_data = self._convert_to_staff_data(parsed_data)
            
            return staff_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            logger.error(f"Raw response: {analysis_result[:500]}...")
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_json_from_response(self, response: str) -> str:
        """Extract JSON from response, handling markdown code blocks"""
        
        response = response.strip()
        
        # Remove markdown code blocks if present
        if '```json' in response:
            start_marker = '```json'
            end_marker = '```'
            start_index = response.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = response.find(end_marker, start_index)
                if end_index != -1:
                    return response[start_index:end_index].strip()
        
        if response.startswith('```') and response.endswith('```'):
            return response[3:-3].strip()
        
        return response
    
    def _convert_to_staff_data(self, parsed_data: Dict) -> StaffWagesData:
        """Convert parsed JSON to StaffWagesData object"""
        
        def create_staff_list(staff_data_list: List[Dict], organization: str) -> List[StaffMember]:
            staff_members = []
            for staff_data in staff_data_list[:20]:  # Max 20 staff members
                staff_member = StaffMember(
                    name_surname=staff_data.get('name_surname', 'Not specified'),
                    responsibilities=staff_data.get('responsibilities', 'Not specified'),
                    total_hours=int(staff_data.get('total_hours', 0)),
                    hourly_rate=float(staff_data.get('hourly_rate', 0.0)),
                    cost_justification=staff_data.get('cost_justification', 'To be provided'),
                    organization=organization
                )
                staff_members.append(staff_member)
            return staff_members
        
        staff_data = StaffWagesData(
            applicant_staff=create_staff_list(
                parsed_data.get('applicant_staff', []), 'Applicant'
            ),
            partner1_staff=create_staff_list(
                parsed_data.get('partner1_staff', []), 'Partner No 1'
            ),
            partner2_staff=create_staff_list(
                parsed_data.get('partner2_staff', []), 'Partner No 2'
            ),
            partner3_staff=create_staff_list(
                parsed_data.get('partner3_staff', []), 'Partner No 3'
            )
        )
        
        return staff_data

class StaffWagesExcelFiller:
    """Fills Excel form with staff wages data according to exact cell mappings"""
    
    def __init__(self):
        # Cell mappings for each organization
        self.organization_mappings = {
            'Applicant': {
                'start_row': 7,
                'end_row': 26,
                'name_col': 'C',
                'responsibilities_col': 'D',
                'hours_col': 'E',
                'rate_col': 'F',
                'justification_col': 'G'
            },
            'Partner No 1': {
                'start_row': 27,
                'end_row': 46,
                'name_col': 'C',
                'responsibilities_col': 'D',
                'hours_col': 'E',
                'rate_col': 'F',
                'justification_col': 'G'
            },
            'Partner No 2': {
                'start_row': 47,
                'end_row': 66,
                'name_col': 'C',
                'responsibilities_col': 'D',
                'hours_col': 'E',
                'rate_col': 'F',
                'justification_col': 'G'
            },
            'Partner No 3': {
                'start_row': 67,
                'end_row': 86,
                'name_col': 'C',
                'responsibilities_col': 'D',
                'hours_col': 'E',
                'rate_col': 'F',
                'justification_col': 'G'
            }
        }
    
    def fill_excel_form(self, template_path: str, output_path: str, staff_data: StaffWagesData) -> bool:
        """Fill Excel form with staff wages data"""
        
        try:
            logger.info(f"Loading Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            # Use first sheet or find staff/wages-related sheet
            sheet_name = workbook.sheetnames[0]
            for name in workbook.sheetnames:
                if any(keyword in name.lower() for keyword in ['staff', 'wages', 'salary', 'personnel']):
                    sheet_name = name
                    break
            
            sheet = workbook[sheet_name]
            logger.info(f"Using sheet: {sheet_name}")
            
            # Fill staff data for each organization
            self._fill_organization_staff(sheet, staff_data.applicant_staff, 'Applicant')
            self._fill_organization_staff(sheet, staff_data.partner1_staff, 'Partner No 1')
            self._fill_organization_staff(sheet, staff_data.partner2_staff, 'Partner No 2')
            self._fill_organization_staff(sheet, staff_data.partner3_staff, 'Partner No 3')
            
            # Save the filled workbook
            workbook.save(output_path)
            logger.info(f"✅ Successfully saved Excel form: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to fill Excel form: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _fill_organization_staff(self, sheet, staff_members: List[StaffMember], organization: str):
        """Fill staff data for a specific organization"""
        
        if not staff_members:
            logger.info(f"No staff members for {organization}, skipping...")
            return
        
        mapping = self.organization_mappings[organization]
        logger.info(f"Filling {len(staff_members)} staff members for {organization}...")
        
        total_hours = 0
        total_cost = 0.0
        
        for i, staff_member in enumerate(staff_members):
            if i < 20:  # Maximum 20 staff members per organization
                row = mapping['start_row'] + i
                
                # Fill staff member data
                sheet[f"{mapping['name_col']}{row}"] = staff_member.name_surname
                sheet[f"{mapping['responsibilities_col']}{row}"] = staff_member.responsibilities
                sheet[f"{mapping['hours_col']}{row}"] = staff_member.total_hours
                sheet[f"{mapping['rate_col']}{row}"] = staff_member.hourly_rate
                sheet[f"{mapping['justification_col']}{row}"] = staff_member.cost_justification
                
                # Calculate totals
                total_hours += staff_member.total_hours
                total_cost += staff_member.total_hours * staff_member.hourly_rate
                
                logger.info(f"✅ {organization} Row {row}: {staff_member.name_surname} - {staff_member.total_hours}h @ €{staff_member.hourly_rate}/h")
        
        logger.info(f"✅ {organization} Summary: {len(staff_members)} staff, {total_hours} total hours, €{total_cost:,.2f} total cost")

class StaffWagesAgent:
    """Main orchestrator for staff wages processing with streaming"""
    
    def __init__(self, groq_api_key: str):
        self.analyzer = GroqStaffAnalyzer(groq_api_key)
        self.excel_filler = StaffWagesExcelFiller()
    
    def read_data_file(self, file_path: str) -> str:
        """Read content from data file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"✅ Successfully read data file: {file_path} ({len(content)} characters)")
            return content
        except FileNotFoundError:
            logger.error(f"❌ Data file not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"❌ Error reading data file: {e}")
            raise
    
    def process_staff_wages(self, data_file_path: str, template_path: str, output_path: str) -> Dict[str, Any]:
        """Complete workflow for processing staff wages"""
        
        results = {
            'success': False,
            'steps_completed': [],
            'staff_data': None,
            'statistics': {},
            'errors': []
        }
        
        try:
            # Step 1: Read data file
            logger.info("🔄 STEP 1: Reading Data File")
            text_content = self.read_data_file(data_file_path)
            results['steps_completed'].append('data_file_read')
            
            # Step 2: Analyze with Groq API (using streaming)
            logger.info("🔄 STEP 2: Analyzing Staff Wages with Groq Streaming API")
            staff_data = self.analyzer.analyze_staff_data_streaming(text_content)
            results['staff_data'] = staff_data
            results['steps_completed'].append('data_analyzed')
            
            # Step 3: Generate statistics
            logger.info("🔄 STEP 3: Generating Statistics")
            statistics = self._generate_statistics(staff_data)
            results['statistics'] = statistics
            results['steps_completed'].append('statistics_generated')
            
            # Step 4: Fill Excel form
            logger.info("🔄 STEP 4: Filling Excel Form")
            excel_success = self.excel_filler.fill_excel_form(template_path, output_path, staff_data)
            
            if excel_success:
                results['steps_completed'].append('excel_filled')
                results['success'] = True
                logger.info("✅ PROCESS COMPLETED SUCCESSFULLY")
            else:
                results['errors'].append('Excel filling failed')
                
        except Exception as e:
            error_msg = f"❌ Process failed at step {len(results['steps_completed']) + 1}: {e}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _generate_statistics(self, staff_data: StaffWagesData) -> Dict[str, Any]:
        """Generate comprehensive statistics"""
        
        def calculate_org_stats(staff_list: List[StaffMember], org_name: str):
            if not staff_list:
                return {
                    'staff_count': 0,
                    'total_hours': 0,
                    'total_cost': 0.0,
                    'avg_hourly_rate': 0.0,
                    'min_rate': 0.0,
                    'max_rate': 0.0
                }
            
            total_hours = sum(s.total_hours for s in staff_list)
            total_cost = sum(s.total_hours * s.hourly_rate for s in staff_list)
            rates = [s.hourly_rate for s in staff_list]
            
            return {
                'staff_count': len(staff_list),
                'total_hours': total_hours,
                'total_cost': total_cost,
                'avg_hourly_rate': sum(rates) / len(rates),
                'min_rate': min(rates),
                'max_rate': max(rates)
            }
        
        # Calculate statistics for each organization
        applicant_stats = calculate_org_stats(staff_data.applicant_staff, 'Applicant')
        partner1_stats = calculate_org_stats(staff_data.partner1_staff, 'Partner No 1')
        partner2_stats = calculate_org_stats(staff_data.partner2_staff, 'Partner No 2')
        partner3_stats = calculate_org_stats(staff_data.partner3_staff, 'Partner No 3')
        
        # Calculate overall project statistics
        all_staff = (staff_data.applicant_staff + staff_data.partner1_staff + 
                    staff_data.partner2_staff + staff_data.partner3_staff)
        
        total_project_hours = sum(s.total_hours for s in all_staff)
        total_project_cost = sum(s.total_hours * s.hourly_rate for s in all_staff)
        
        # Top 5 highest paid staff members
        top_paid_staff = sorted(all_staff, key=lambda s: s.hourly_rate, reverse=True)[:5]
        
        # Cost distribution by organization
        org_costs = {
            'Applicant': applicant_stats['total_cost'],
            'Partner No 1': partner1_stats['total_cost'],
            'Partner No 2': partner2_stats['total_cost'],
            'Partner No 3': partner3_stats['total_cost']
        }
        
        active_organizations = len([cost for cost in org_costs.values() if cost > 0])
        
        statistics = {
            'project_overview': {
                'total_staff': len(all_staff),
                'active_organizations': active_organizations,
                'total_project_hours': total_project_hours,
                'total_project_cost': total_project_cost,
                'average_hourly_rate': total_project_cost / total_project_hours if total_project_hours > 0 else 0
            },
            'organization_breakdown': {
                'applicant': applicant_stats,
                'partner1': partner1_stats,
                'partner2': partner2_stats,
                'partner3': partner3_stats
            },
            'cost_distribution': {
                'applicant_percentage': (applicant_stats['total_cost'] / total_project_cost * 100) if total_project_cost > 0 else 0,
                'partner1_percentage': (partner1_stats['total_cost'] / total_project_cost * 100) if total_project_cost > 0 else 0,
                'partner2_percentage': (partner2_stats['total_cost'] / total_project_cost * 100) if total_project_cost > 0 else 0,
                'partner3_percentage': (partner3_stats['total_cost'] / total_project_cost * 100) if total_project_cost > 0 else 0
            },
            'top_paid_staff': [
                {
                    'name': staff.name_surname,
                    'organization': staff.organization,
                    'hourly_rate': staff.hourly_rate,
                    'total_hours': staff.total_hours,
                    'total_cost': staff.total_hours * staff.hourly_rate
                } for staff in top_paid_staff
            ]
        }
        
        return statistics
    
    def print_results(self, results: Dict[str, Any]):
        """Print comprehensive results"""
        
        print("\n" + "="*80)
        print("🚀 STAFF WAGES EXCEL FORM FILLING RESULTS (STREAMING)")
        print("="*80)
        
        # Process Status
        status_icon = "✅ SUCCESS" if results['success'] else "❌ FAILED"
        print(f"\n🔄 PROCESS STATUS: {status_icon}")
        print(f"📋 Steps Completed: {', '.join(results['steps_completed'])}")
        
        if results['errors']:
            print(f"⚠️  Errors: {'; '.join(results['errors'])}")
        
        # Statistics
        if results['staff_data'] and results['statistics']:
            stats = results['statistics']
            
            print(f"\n📊 PROJECT OVERVIEW")
            overview = stats['project_overview']
            print(f"   Total Staff Members: {overview['total_staff']}")
            print(f"   Active Organizations: {overview['active_organizations']}")
            print(f"   Total Project Hours: {overview['total_project_hours']:,}")
            print(f"   Total Project Cost: €{overview['total_project_cost']:,.2f}")
            print(f"   Average Hourly Rate: €{overview['average_hourly_rate']:.2f}/hour")
            
            print(f"\n🏢 ORGANIZATION BREAKDOWN")
            org_breakdown = stats['organization_breakdown']
            cost_dist = stats['cost_distribution']
            
            organizations = [
                ('Applicant', org_breakdown['applicant'], cost_dist['applicant_percentage']),
                ('Partner No 1', org_breakdown['partner1'], cost_dist['partner1_percentage']),
                ('Partner No 2', org_breakdown['partner2'], cost_dist['partner2_percentage']),
                ('Partner No 3', org_breakdown['partner3'], cost_dist['partner3_percentage'])
            ]
            
            for org_name, org_stats, percentage in organizations:
                if org_stats['staff_count'] > 0:
                    print(f"   {org_name}:")
                    print(f"     • Staff: {org_stats['staff_count']} members")
                    print(f"     • Hours: {org_stats['total_hours']:,} ({percentage:.1f}% of project)")
                    print(f"     • Cost: €{org_stats['total_cost']:,.2f}")
                    print(f"     • Rate Range: €{org_stats['min_rate']:.2f} - €{org_stats['max_rate']:.2f}/hour")
            
            print(f"\n💰 TOP 5 HIGHEST PAID STAFF")
            for i, staff in enumerate(stats['top_paid_staff'], 1):
                print(f"   {i}. {staff['name']} ({staff['organization']})")
                print(f"      €{staff['hourly_rate']:.2f}/hr × {staff['total_hours']}h = €{staff['total_cost']:,.2f}")
        
        print("\n" + "="*80)

def mainstaff(data_file_path):
    """Main function to run the staff wages agent"""
    
    print("🚀 STAFF WAGES EXCEL FORM FILLER (GROQ STREAMING)")
    print("="*60)
    
    # Configuration
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
    if not GROQ_API_KEY:
        print("❌ Error: GROQ_API_KEY environment variable not set")
        print("Please set your Groq API key in .env file: GROQ_API_KEY='your_api_key_here'")
        return
    

    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    template_path = "code/agents/ENGLISH_1A_priedas_InoStartas_en/Staff/inputStaff.xlsx"  # Your Excel template
    output_path = "code/agents/ENGLISH_1A_priedas_InoStartas_en/Staff/Staff.xlsx"
    
    # Validate required files exist
    required_files = [data_file_path, template_path]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print(f"❌ Missing required files: {', '.join(missing_files)}")
        print("Required files:")
        print(f"  • {data_file_path} - Text file with staff wages data")
        print(f"  • {template_path} - Excel template file (Staff Wages Form)")
        return
    
    try:
        # Initialize agent
        print(f"\n🔧 Initializing Staff Wages Agent with Groq Streaming...")
        agent = StaffWagesAgent(GROQ_API_KEY)
        
        # Process staff wages
        print(f"\n🔄 Processing staff wages from {data_file_path}...")
        results = agent.process_staff_wages(data_file_path, template_path, output_path)
        
        # Print results
        agent.print_results(results)
        
        if results['success']:
            print(f"\n✅ Staff wages Excel form successfully filled: {output_path}")
            print(f"\n🎯 Key Features Implemented:")
            print(f"   • ✅ Groq streaming API with meta-llama/llama-4-scout-17b-16e-instruct")
            print(f"   • ✅ Applicant staff data (C7:G26)")
            print(f"   • ✅ Partner No 1 staff data (C27:G46)")
            print(f"   • ✅ Partner No 2 staff data (C47:G66)")
            print(f"   • ✅ Partner No 3 staff data (C67:G86)")
            print(f"   • ✅ Up to 20 staff members per organization")
            print(f"   • ✅ Detailed cost justification documentation")
            print(f"   • ✅ Comprehensive statistics and analysis")
            
            if results['statistics']:
                total_staff = results['statistics']['project_overview']['total_staff']
                total_cost = results['statistics']['project_overview']['total_project_cost']
                print(f"\n📊 Project Overview:")
                print(f"   • Total Staff: {total_staff} members")
                print(f"   • Total Labor Cost: €{total_cost:,.2f}")
        else:
            print(f"\n❌ Process failed. Check the logs above for details.")
            
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        print(f"\n❌ Fatal error: {e}")

