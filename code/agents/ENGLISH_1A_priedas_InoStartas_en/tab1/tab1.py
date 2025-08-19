# import argparse
# import json
# import logging
# import sys
# from typing import Dict, List, Any, Optional, Tuple
# from pathlib import Path
# import pandas as pd
# from groq import Groq
# import openpyxl
# from datetime import datetime
# import os
# from dotenv import load_dotenv

# # Load environment variables
# load_dotenv()

# # Configure logging
# logging.basicConfig(
#     level=logging.INFO,
#     format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
#     handlers=[
#         logging.FileHandler('eu_form_agent.log'),
#         logging.StreamHandler(sys.stdout)
#     ]
# )
# logger = logging.getLogger(__name__)


# class ValidationError(Exception):
#     """Custom exception for validation errors"""
#     pass


# class FormFillingError(Exception):
#     """Custom exception for form filling errors"""
#     pass


# class EnhancedEUFormAgent:
#     """
#     Enhanced AI-powered agent for filling EU project forms
#     """
    
#     def __init__(self, groq_api_key: str = None):
#         """Initialize the enhanced agent"""
#         self.api_key = groq_api_key or os.getenv("GROQ_API_KEY")
#         if not self.api_key:
#             raise ValueError("Groq API key not provided. Set GROQ_API_KEY environment variable.")
        
#         self.client = Groq(api_key=self.api_key)
#         self.model = "meta-llama/llama-4-scout-17b-16e-instruct"
        
#         # Statistics tracking
#         self.processing_stats = {
#             "start_time": None,
#             "end_time": None,
#             "input_length": 0,
#             "extracted_entries": 0,
#             "filled_cells": 0,
#             "validation_errors": 0
#         }
#     def create_enhanced_extraction_prompt(self, prompt_path: str = "C://Users//USER//Documents//eu excel form//excel form filling//prompts//1.txt") -> str:
#         """Read detailed prompt for market development service extraction from a file"""
#         try:
#             with open(prompt_path, 'r', encoding='utf-8') as file:
#                 prompt = file.read()
#             return prompt
#         except FileNotFoundError:
#             raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
#         except Exception as e:
#             raise RuntimeError(f"Failed to read prompt file: {e}")
        

# #     def create_enhanced_extraction_prompt(self, input_text: str) -> str:
# #         """Create an enhanced prompt with better instruction clarity"""
# #         prompt = f"""
# # You are a specialized AI assistant for European Union project documentation. Your task is to extract and structure financial and project information from the provided text into a standardized JSON format that will be used to automatically fill official EU forms.

# # CRITICAL INSTRUCTIONS:
# # 1. Extract ONLY explicitly mentioned information - never guess or estimate
# # 2. All monetary values must be precise numbers (no currency symbols)
# # 3. All dates must be in YYYY-MM-DD format
# # 4. All percentages as decimal numbers (75 for 75%, not 0.75)
# # 5. Use null for any missing information
# # 6. Validate all calculations where possible

# # INPUT TEXT TO ANALYZE:
# # {input_text}

# # REQUIRED OUTPUT FORMAT (JSON only):
# # {{
# #     "project_info": {{
# #         "type_of_rd": "Applied Research|Experimental Development|Industrial Research",
# #         "project_impact_no": "string",
# #         "impact_name": "string",
# #         "legal_entity_name": "string", 
# #         "applicant_partner": "Applicant|Partner 1|Partner 2|Partner 3",
# #         "funding_intensity": number
# #     }},
# #     "staff_costs": [
# #         {{
# #             "position": "string",
# #             "person_name": "string",
# #             "monthly_salary": number,
# #             "employer_costs": number,
# #             "duration_months": number,
# #             "total_salary_costs": number,
# #             "hourly_rate": number,
# #             "eligible_costs": number,
# #             "funding_requested": number,
# #             "supporting_docs": "string"
# #         }}
# #     ],
# #     "mission_expenses": {{
# #         "total_eligible_costs": number,
# #         "total_funding_requested": number,
# #         "missions": [
# #             {{
# #                 "mission_name": "string",
# #                 "destination_country": "string",
# #                 "duration_days": number,
# #                 "travelers_count": number,
# #                 "per_diem_rate": number,
# #                 "accommodation_rate": number,
# #                 "travel_costs": number,
# #                 "participation_fee": number,
# #                 "total_mission_cost": number
# #             }}
# #         ]
# #     }},
# #     "equipment_depreciation": [
# #         {{
# #             "equipment_name": "string",
# #             "equipment_description": "string",
# #             "acquisition_date": "YYYY-MM-DD",
# #             "acquisition_value": number,
# #             "depreciation_period_months": number,
# #             "residual_value": number,
# #             "monthly_depreciation": number,
# #             "project_usage_months": number,
# #             "usage_percentage": number,
# #             "project_depreciation_amount": number
# #         }}
# #     ],
# #     "rd_services": [
# #         {{
# #             "service_name": "string",
# #             "service_description": "string",
# #             "service_provider": "string",
# #             "units": "string",
# #             "quantity": number,
# #             "unit_price": number,
# #             "total_cost": number,
# #             "eligible_costs": number,
# #             "funding_requested": number,
# #             "supporting_docs": "string"
# #         }}
# #     ],
# #     "materials_supplies": [
# #         {{
# #             "item_name": "string",
# #             "item_description": "string",
# #             "category": "string",
# #             "units": "string",
# #             "quantity": number,
# #             "unit_price": number,
# #             "total_cost": number,
# #             "eligible_costs": number,
# #             "funding_requested": number,
# #             "supporting_docs": "string"
# #         }}
# #     ],
# #     "equipment_rental": [
# #         {{
# #             "equipment_name": "string",
# #             "rental_provider": "string",
# #             "monthly_rental_cost": number,
# #             "usage_duration_months": number,
# #             "total_rental_cost": number,
# #             "units": "string",
# #             "quantity": number,
# #             "unit_price": number,
# #             "eligible_costs": number,
# #             "funding_requested": number
# #         }}
# #     ],
# #     "premises_rental": [
# #         {{
# #             "premises_address": "string",
# #             "premises_description": "string",
# #             "monthly_rental_cost": number,
# #             "usage_duration_months": number,
# #             "total_rental_cost": number,
# #             "units": "string",
# #             "quantity": number,
# #             "unit_price": number,
# #             "eligible_costs": number,
# #             "funding_requested": number
# #         }}
# #     ],
# #     "project_totals": {{
# #         "total_eligible_costs": number,
# #         "total_funding_requested": number,
# #         "own_contribution": number,
# #         "project_duration_months": number
# #     }}
# # }}

# # VALIDATION RULES:
# # - All financial calculations must be consistent
# # - Funding requested should not exceed eligible costs
# # - Percentages must be between 0 and 100
# # - Dates must be valid and logical
# # - Duration must be positive numbers

# # Return ONLY the JSON object - no explanations, no additional text:
# # """
# #         return prompt

#     def extract_data_with_enhanced_llm(self, input_text: str) -> Dict[str, Any]:
#         """Enhanced LLM extraction with retry logic and validation"""
#         max_retries = 3
        
#         for attempt in range(max_retries):
#             try:
#                 # prompt = self.create_enhanced_extraction_prompt(input_text)
#                 prompt = self.create_enhanced_extraction_prompt()
#                 logger.info(f"Attempt {attempt + 1}: Sending request to Groq LLM...")
#                 completion = self.client.chat.completions.create(
#                     model=self.model,
#                     messages=[
#                         {
#                             "role": "system",
#                             "content": "You are a precise data extraction specialist for EU forms. Return only valid JSON."
#                         },
#                         {
#                             "role": "user", 
#                             "content": prompt
#                         }
#                     ],
#                     temperature=0.3
#                 )
                
#                 response_content = completion.choices[0].message.content.strip()
#                 logger.info("Received response from Groq LLM")
                
#                 # Clean the response (remove any markdown formatting)
#                 if response_content.startswith("```json"):
#                     response_content = response_content[7:-3]
#                 elif response_content.startswith("```"):
#                     response_content = response_content[3:-3]
                
#                 # Parse JSON response
#                 extracted_data = json.loads(response_content)
#                 logger.info("Successfully parsed JSON response")
                
#                 # Validate the extracted data
#                 validation_result = self.validate_extracted_data_enhanced(extracted_data)
#                 if validation_result[0]:
#                     return extracted_data
#                 else:
#                     logger.warning(f"Validation failed on attempt {attempt + 1}: {validation_result[1]}")
#                     if attempt == max_retries - 1:
#                         raise ValidationError(f"Data validation failed after {max_retries} attempts: {validation_result[1]}")
                    
#             except json.JSONDecodeError as e:
#                 logger.error(f"JSON parsing failed on attempt {attempt + 1}: {str(e)}")
#                 if attempt == max_retries - 1:
#                     logger.error(f"Raw response: {response_content}")
#                     raise
#             except Exception as e:
#                 logger.error(f"LLM extraction failed on attempt {attempt + 1}: {str(e)}")
#                 if attempt == max_retries - 1:
#                     raise
        
#         raise Exception("Failed to extract data after maximum retries")

#     def validate_extracted_data_enhanced(self, data: Dict[str, Any]) -> Tuple[bool, str]:
#         """Enhanced validation with detailed error reporting"""
#         errors = []
        
#         # Check required top-level keys
#         required_keys = ["project_info", "staff_costs", "mission_expenses", 
#                         "equipment_depreciation", "rd_services", "materials_supplies",
#                         "equipment_rental", "premises_rental"]
        
#         for key in required_keys:
#             if key not in data:
#                 errors.append(f"Missing required section: {key}")
        
#         # Validate project_info
#         if "project_info" in data:
#             project_info = data["project_info"]
#             if not isinstance(project_info, dict):
#                 errors.append("project_info must be a dictionary")
#             else:
#                 # Check funding intensity
#                 if "funding_intensity" in project_info:
#                     intensity = project_info["funding_intensity"]
#                     if isinstance(intensity, (int, float)) and not (0 <= intensity <= 100):
#                         errors.append(f"Funding intensity must be between 0-100, got {intensity}")
        
#         # Validate staff costs
#         if "staff_costs" in data and isinstance(data["staff_costs"], list):
#             for i, staff in enumerate(data["staff_costs"]):
#                 if not isinstance(staff, dict):
#                     errors.append(f"Staff cost entry {i} must be a dictionary")
#                     continue
                
#                 # Validate financial consistency
#                 if all(k in staff for k in ["monthly_salary", "employer_costs", "duration_months"]):
#                     expected_total = (staff["monthly_salary"] + staff["employer_costs"]) * staff["duration_months"]
#                     if "total_salary_costs" in staff:
#                         actual_total = staff["total_salary_costs"]
#                         if abs(expected_total - actual_total) > 0.01:
#                             errors.append(f"Staff {i}: Inconsistent salary calculation. Expected {expected_total}, got {actual_total}")
        
#         # Validate mission expenses
#         if "mission_expenses" in data and isinstance(data["mission_expenses"], dict):
#             missions = data["mission_expenses"].get("missions", [])
#             if isinstance(missions, list):
#                 for i, mission in enumerate(missions):
#                     if "duration_days" in mission and isinstance(mission["duration_days"], (int, float)):
#                         if mission["duration_days"] <= 0:
#                             errors.append(f"Mission {i}: Duration must be positive")
        
#         # Validate equipment depreciation
#         if "equipment_depreciation" in data and isinstance(data["equipment_depreciation"], list):
#             for i, equipment in enumerate(data["equipment_depreciation"]):
#                 if isinstance(equipment, dict):
#                     # Check depreciation calculation
#                     if all(k in equipment for k in ["acquisition_value", "residual_value", "depreciation_period_months"]):
#                         expected_monthly = (equipment["acquisition_value"] - equipment["residual_value"]) / equipment["depreciation_period_months"]
#                         if "monthly_depreciation" in equipment:
#                             actual_monthly = equipment["monthly_depreciation"]
#                             if abs(expected_monthly - actual_monthly) > 0.01:
#                                 errors.append(f"Equipment {i}: Inconsistent depreciation calculation")
        
#         # Count validation errors for statistics
#         self.processing_stats["validation_errors"] = len(errors)
        
#         if errors:
#             return False, "; ".join(errors)
        
#         logger.info("Enhanced data validation passed")
#         return True, "Validation successful"

#     def generate_summary_report(self, data: Dict[str, Any], output_path: str) -> str:
#         """Generate a summary report of the extracted data"""
#         report_path = output_path.replace('.xlsx', '_summary_report.txt')
        
#         try:
#             with open(report_path, 'w', encoding='utf-8') as f:
#                 f.write("EU FORM FILLING AGENT - PROCESSING SUMMARY REPORT\n")
#                 f.write("=" * 55 + "\n\n")
                
#                 # Processing statistics
#                 f.write("PROCESSING STATISTICS:\n")
#                 f.write("-" * 22 + "\n")
#                 duration = (self.processing_stats["end_time"] - self.processing_stats["start_time"]).total_seconds()
#                 f.write(f"Processing Duration: {duration:.2f} seconds\n")
#                 f.write(f"Input Text Length: {self.processing_stats['input_length']} characters\n")
#                 f.write(f"Extracted Entries: {self.processing_stats['extracted_entries']}\n")
#                 f.write(f"Filled Cells: {self.processing_stats['filled_cells']}\n")
#                 f.write(f"Validation Errors: {self.processing_stats['validation_errors']}\n\n")
                
#                 # Project overview
#                 project_info = data.get("project_info", {})
#                 f.write("PROJECT OVERVIEW:\n")
#                 f.write("-" * 17 + "\n")
#                 f.write(f"Legal Entity: {project_info.get('legal_entity_name', 'N/A')}\n")
#                 f.write(f"Impact Name: {project_info.get('impact_name', 'N/A')}\n")
#                 f.write(f"R&D Type: {project_info.get('type_of_rd', 'N/A')}\n")
#                 f.write(f"Funding Intensity: {project_info.get('funding_intensity', 'N/A')}%\n\n")
                
#                 # Financial summary
#                 project_totals = data.get("project_totals", {})
#                 f.write("FINANCIAL SUMMARY:\n")
#                 f.write("-" * 18 + "\n")
#                 f.write(f"Total Eligible Costs: {project_totals.get('total_eligible_costs', 'N/A')} EUR\n")
#                 f.write(f"Total Funding Requested: {project_totals.get('total_funding_requested', 'N/A')} EUR\n")
#                 f.write(f"Own Contribution: {project_totals.get('own_contribution', 'N/A')} EUR\n\n")
                
#                 # Section breakdown
#                 sections = [
#                     ("Staff Costs", "staff_costs"),
#                     ("Mission Expenses", "mission_expenses"),
#                     ("Equipment Depreciation", "equipment_depreciation"),
#                     ("R&D Services", "rd_services"),
#                     ("Materials & Supplies", "materials_supplies"),
#                     ("Equipment Rental", "equipment_rental"),
#                     ("Premises Rental", "premises_rental")
#                 ]
                
#                 f.write("SECTION BREAKDOWN:\n")
#                 f.write("-" * 18 + "\n")
#                 for section_name, section_key in sections:
#                     section_data = data.get(section_key, [])
#                     if isinstance(section_data, list):
#                         count = len(section_data)
#                     elif isinstance(section_data, dict):
#                         count = len(section_data.get("missions", [])) if section_key == "mission_expenses" else 1
#                     else:
#                         count = 0
#                     f.write(f"{section_name}: {count} entries\n")
                
#                 f.write(f"\nReport generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
#             logger.info(f"Summary report generated: {report_path}")
#             return report_path
            
#         except Exception as e:
#             logger.error(f"Error generating summary report: {str(e)}")
#             return None

#     def validate_excel_template(self, excel_path: str) -> bool:
#         """Validate that the Excel template has the expected structure"""
#         try:
#             workbook = openpyxl.load_workbook(excel_path)
#             worksheet = workbook.active
            
#             # Check for key cells that should exist in the template
#             test_cells = ["D1", "D2", "B10", "G30", "B102", "B123"]
#             for cell in test_cells:
#                 try:
#                     _ = worksheet[cell]
#                 except Exception:
#                     logger.error(f"Template validation failed: Cell {cell} not accessible")
#                     return False
            
#             logger.info("Excel template validation passed")
#             return True
            
#         except Exception as e:
#             logger.error(f"Error validating Excel template: {str(e)}")
#             return False

#     def process_with_enhanced_features(self, input_file: str, excel_template: str, 
#                                      output_file: str = None, validate_only: bool = False) -> str:
#         """Enhanced processing with additional features"""
#         try:
#             self.processing_stats["start_time"] = datetime.now()
#             logger.info("Starting enhanced form processing...")
            
#             # Validate inputs
#             if not Path(input_file).exists():
#                 raise FileNotFoundError(f"Input file not found: {input_file}")
            
#             if not Path(excel_template).exists():
#                 raise FileNotFoundError(f"Excel template not found: {excel_template}")
            
#             # Validate Excel template
#             if not self.validate_excel_template(excel_template):
#                 raise ValidationError("Excel template validation failed")
            
#             # Read input data
#             input_text = self.read_input_data(input_file)
#             self.processing_stats["input_length"] = len(input_text)
            
#             # Extract structured data using enhanced LLM
#             extracted_data = self.extract_data_with_enhanced_llm(input_text)
            
#             # Count extracted entries
#             total_entries = 0
#             for key, value in extracted_data.items():
#                 if isinstance(value, list):
#                     total_entries += len(value)
#                 elif isinstance(value, dict) and key == "mission_expenses":
#                     total_entries += len(value.get("missions", []))
#                 elif isinstance(value, dict):
#                     total_entries += 1
#             self.processing_stats["extracted_entries"] = total_entries
            
#             if validate_only:
#                 logger.info("Validation-only mode completed successfully")
#                 return "Validation completed successfully"
            
#             # Fill Excel form with enhanced error handling
#             output_path = self.fill_excel_form_enhanced(extracted_data, excel_template, output_file)
            
#             # Generate summary report
#             self.generate_summary_report(extracted_data, output_path)
            
#             self.processing_stats["end_time"] = datetime.now()
#             logger.info("Enhanced form processing completed successfully!")
            
#             return output_path
            
#         except Exception as e:
#             self.processing_stats["end_time"] = datetime.now()
#             logger.error(f"Error in enhanced form processing: {str(e)}")
#             raise

#     def fill_excel_form_enhanced(self, data: Dict[str, Any], excel_path: str, output_path: str = None) -> str:
#         """Enhanced Excel form filling with better error handling and cell tracking"""
#         try:
#             workbook = openpyxl.load_workbook(excel_path)
#             worksheet = workbook.active
            
#             filled_cells = 0
#             logger.info("Filling Excel form with extracted data...")
            
#             # Fill each section and count filled cells
#             filled_cells += self._fill_project_info_enhanced(worksheet, data.get("project_info", {}))
#             filled_cells += self._fill_staff_costs_enhanced(worksheet, data.get("staff_costs", []))
#             filled_cells += self._fill_mission_expenses_enhanced(worksheet, data.get("mission_expenses", {}))
#             filled_cells += self._fill_equipment_depreciation_enhanced(worksheet, data.get("equipment_depreciation", []))
#             filled_cells += self._fill_rd_services_enhanced(worksheet, data.get("rd_services", []))
#             filled_cells += self._fill_materials_supplies_enhanced(worksheet, data.get("materials_supplies", []))
#             filled_cells += self._fill_equipment_rental_enhanced(worksheet, data.get("equipment_rental", []))
#             filled_cells += self._fill_premises_rental_enhanced(worksheet, data.get("premises_rental", []))
            
#             self.processing_stats["filled_cells"] = filled_cells
            
#             # Generate output filename if not provided
#             if output_path is None:
#                 timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
#                 output_path = f"filled_eu_form_{timestamp}.xlsx"
            
#             # Save the workbook
#             workbook.save(output_path)
#             logger.info(f"Successfully saved filled form to {output_path}")
#             logger.info(f"Total cells filled: {filled_cells}")
            
#             return output_path
            
#         except Exception as e:
#             logger.error(f"Error in enhanced Excel form filling: {str(e)}")
#             raise FormFillingError(f"Failed to fill Excel form: {str(e)}")

#     def _fill_project_info_enhanced(self, worksheet, project_info: Dict[str, Any]) -> int:
#         """Enhanced project info filling with cell counting"""
#         filled_count = 0
#         try:
#             mappings = [
#                 ("type_of_rd", "D1"),
#                 ("project_impact_no", "D2"),
#                 ("impact_name", "D3"),
#                 ("legal_entity_name", "D4"),
#                 ("applicant_partner", "D5"),
#                 ("funding_intensity", "H5")
#             ]
            
#             for field, cell in mappings:
#                 if project_info.get(field) is not None:
#                     worksheet[cell] = project_info[field]
#                     filled_count += 1
                    
#             logger.info(f"Project information: {filled_count} cells filled")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling project info: {str(e)}")
#             return filled_count

#     def _fill_staff_costs_enhanced(self, worksheet, staff_costs: List[Dict[str, Any]]) -> int:
#         """Enhanced staff costs filling"""
#         filled_count = 0
#         try:
#             start_row = 10
#             for i, staff in enumerate(staff_costs[:20]):
#                 row = start_row + i
                
#                 mappings = [
#                     ("position", f"B{row}"),
#                     ("monthly_salary", f"C{row}"),
#                     ("employer_costs", f"D{row}"),
#                     ("duration_months", f"E{row}"),
#                     ("hourly_rate", f"F{row}"),
#                     ("eligible_costs", f"G{row}"),
#                     ("funding_requested", f"H{row}"),
#                     ("supporting_docs", f"I{row}")
#                 ]
                
#                 for field, cell in mappings:
#                     if staff.get(field) is not None:
#                         worksheet[cell] = staff[field]
#                         filled_count += 1
                        
#             logger.info(f"Staff costs: {filled_count} cells filled for {len(staff_costs)} staff members")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling staff costs: {str(e)}")
#             return filled_count

#     def _fill_mission_expenses_enhanced(self, worksheet, mission_expenses: Dict[str, Any]) -> int:
#         """Enhanced mission expenses filling"""
#         filled_count = 0
#         try:
#             # Fill totals
#             if mission_expenses.get("total_eligible_costs") is not None:
#                 worksheet["G30"] = mission_expenses["total_eligible_costs"]
#                 filled_count += 1
#             if mission_expenses.get("total_funding_requested") is not None:
#                 worksheet["H30"] = mission_expenses["total_funding_requested"]
#                 filled_count += 1
                
#             # Fill individual missions
#             missions = mission_expenses.get("missions", [])
#             start_row = 31
            
#             for i, mission in enumerate(missions[:10]):
#                 base_row = start_row + (i * 7)
                
#                 if mission.get("mission_name"):
#                     worksheet[f"B{base_row}"] = mission["mission_name"]
#                     filled_count += 1
#                 if mission.get("destination_country"):
#                     worksheet[f"J{base_row}"] = mission["destination_country"]
#                     filled_count += 1
#                 if mission.get("duration_days"):
#                     worksheet[f"J{base_row+1}"] = mission["duration_days"]
#                     filled_count += 1
#                 if mission.get("travelers_count"):
#                     worksheet[f"J{base_row+2}"] = mission["travelers_count"]
#                     filled_count += 1
                    
#             logger.info(f"Mission expenses: {filled_count} cells filled for {len(missions)} missions")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling mission expenses: {str(e)}")
#             return filled_count

#     def _fill_equipment_depreciation_enhanced(self, worksheet, equipment: List[Dict[str, Any]]) -> int:
#         """Enhanced equipment depreciation filling"""
#         filled_count = 0
#         try:
#             start_row = 102
#             for i, equip in enumerate(equipment[:20]):
#                 row = start_row + i
                
#                 mappings = [
#                     ("equipment_name", f"B{row}"),
#                     ("acquisition_date", f"K{row}"),
#                     ("acquisition_value", f"L{row}"),
#                     ("depreciation_period_months", f"M{row}"),
#                     ("residual_value", f"N{row}"),
#                     ("monthly_depreciation", f"O{row}"),
#                     ("project_usage_months", f"P{row}"),
#                     ("usage_percentage", f"Q{row}"),
#                     ("project_depreciation_amount", f"R{row}")
#                 ]
                
#                 for field, cell in mappings:
#                     if equip.get(field) is not None:
#                         worksheet[cell] = equip[field]
#                         filled_count += 1
                        
#             logger.info(f"Equipment depreciation: {filled_count} cells filled for {len(equipment)} items")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling equipment depreciation: {str(e)}")
#             return filled_count

#     def _fill_rd_services_enhanced(self, worksheet, rd_services: List[Dict[str, Any]]) -> int:
#         """Enhanced R&D services filling"""
#         filled_count = 0
#         try:
#             start_row = 123
#             for i, service in enumerate(rd_services[:10]):
#                 row = start_row + i
                
#                 mappings = [
#                     ("service_name", f"B{row}"),
#                     ("units", f"D{row}"),
#                     ("quantity", f"E{row}"),
#                     ("unit_price", f"F{row}"),
#                     ("eligible_costs", f"G{row}"),
#                     ("funding_requested", f"H{row}"),
#                     ("supporting_docs", f"I{row}")
#                 ]
                
#                 for field, cell in mappings:
#                     if service.get(field) is not None:
#                         worksheet[cell] = service[field]
#                         filled_count += 1
                        
#             logger.info(f"R&D services: {filled_count} cells filled for {len(rd_services)} services")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling R&D services: {str(e)}")
#             return filled_count

#     def _fill_materials_supplies_enhanced(self, worksheet, materials: List[Dict[str, Any]]) -> int:
#         """Enhanced materials and supplies filling"""
#         filled_count = 0
#         try:
#             start_row = 145
#             for i, material in enumerate(materials[:25]):
#                 row = start_row + i
                
#                 mappings = [
#                     ("item_name", f"B{row}"),
#                     ("units", f"D{row}"),
#                     ("quantity", f"E{row}"),
#                     ("unit_price", f"F{row}"),
#                     ("eligible_costs", f"G{row}"),
#                     ("funding_requested", f"H{row}"),
#                     ("supporting_docs", f"I{row}")
#                 ]
                
#                 for field, cell in mappings:
#                     if material.get(field) is not None:
#                         worksheet[cell] = material[field]
#                         filled_count += 1
                        
#             logger.info(f"Materials and supplies: {filled_count} cells filled for {len(materials)} items")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling materials and supplies: {str(e)}")
#             return filled_count

#     def _fill_equipment_rental_enhanced(self, worksheet, equipment_rental: List[Dict[str, Any]]) -> int:
#         """Enhanced equipment rental filling"""
#         filled_count = 0
#         try:
#             start_row = 171
#             for i, rental in enumerate(equipment_rental[:10]):
#                 row = start_row + i
                
#                 mappings = [
#                     ("equipment_name", f"B{row}"),
#                     ("units", f"D{row}"),
#                     ("quantity", f"E{row}"),
#                     ("unit_price", f"F{row}"),
#                     ("eligible_costs", f"G{row}"),
#                     ("funding_requested", f"H{row}"),
#                     ("monthly_rental_cost", f"K{row}"),
#                     ("usage_duration_months", f"L{row}")
#                 ]
                
#                 for field, cell in mappings:
#                     if rental.get(field) is not None:
#                         worksheet[cell] = rental[field]
#                         filled_count += 1
                        
#             logger.info(f"Equipment rental: {filled_count} cells filled for {len(equipment_rental)} items")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling equipment rental: {str(e)}")
#             return filled_count

#     def _fill_premises_rental_enhanced(self, worksheet, premises_rental: List[Dict[str, Any]]) -> int:
#         """Enhanced premises rental filling"""
#         filled_count = 0
#         try:
#             start_row = 182
#             for i, premises in enumerate(premises_rental[:5]):
#                 row = start_row + i
                
#                 mappings = [
#                     ("premises_address", f"B{row}"),
#                     ("units", f"D{row}"),
#                     ("quantity", f"E{row}"),
#                     ("unit_price", f"F{row}"),
#                     ("eligible_costs", f"G{row}"),
#                     ("funding_requested", f"H{row}"),
#                     ("monthly_rental_cost", f"K{row}"),
#                     ("usage_duration_months", f"L{row}")
#                 ]
                
#                 for field, cell in mappings:
#                     if premises.get(field) is not None:
#                         worksheet[cell] = premises[field]
#                         filled_count += 1
                        
#             logger.info(f"Premises rental: {filled_count} cells filled for {len(premises_rental)} premises")
#             return filled_count
            
#         except Exception as e:
#             logger.error(f"Error filling premises rental: {str(e)}")
#             return filled_count

#     def read_input_data(self, file_path: str) -> str:
#         """Read and return content from the input text file"""
#         try:
#             with open(file_path, 'r', encoding='utf-8') as file:
#                 content = file.read()
#             logger.info(f"Successfully read {len(content)} characters from {file_path}")
#             return content
#         except FileNotFoundError:
#             logger.error(f"File {file_path} not found")
#             raise
#         except Exception as e:
#             logger.error(f"Error reading file {file_path}: {str(e)}")
#             raise


# def create_cli_parser() -> argparse.ArgumentParser:
#     """Create command-line interface parser"""
#     parser = argparse.ArgumentParser(
#         description='Enhanced EU Form Filling Agent using AI',
#         formatter_class=argparse.RawDescriptionHelpFormatter,
#         epilog="""
# Examples:
#   python enhanced_agent.py --input data.txt --template form.xlsx
#   python enhanced_agent.py --input data.txt --template form.xlsx --output my_form.xlsx
#   python enhanced_agent.py --input data.txt --template form.xlsx --validate-only
#   python enhanced_agent.py --input data.txt --template form.xlsx --api-key YOUR_KEY
#         """
#     )
    
#     parser.add_argument(
#         '--input', '-i',
#         required=True,
#         help='Path to input text file containing project data'
#     )
    
#     parser.add_argument(
#         '--template', '-t',
#         required=True,
#         help='Path to Excel template file (ENGLISH_1A priedas_InoStartas en.xlsx)'
#     )
    
#     parser.add_argument(
#         '--output', '-o',
#         help='Path for output filled Excel file (default: auto-generated with timestamp)'
#     )
    
#     parser.add_argument(
#         '--api-key',
#         help='Groq API key (default: reads from GROQ_API_KEY environment variable)'
#     )
    
#     parser.add_argument(
#         '--validate-only',
#         action='store_true',
#         help='Only validate extracted data without filling the form'
#     )
    
#     parser.add_argument(
#         '--verbose', '-v',
#         action='store_true',
#         help='Enable verbose logging'
#     )
    
#     parser.add_argument(
#         '--version',
#         action='version',
#         version='Enhanced EU Form Agent v2.0'
#     )
    
#     return parser

# def main1():
#     """Main function with hardcoded input/output paths"""
#     try:
#         # # Define file paths manually
#         # input_file = "tab1.txt"  # <- Your input text file
#         # excel_template = "input1.xlsx"  # <- Your Excel template
#         # output_file = "1.xlsx"  # <- Desired output filename

#         input_file = r"C:\Users\USER\Documents\eu excel form\excel form filling\input\tab1.txt"  # <- Your input text file
#         excel_template = "agents/ENGLISH_1A_priedas_InoStartas_en/tab1/input1.xlsx"  # <- Your Excel template
#         output_file = "agents/ENGLISH_1A_priedas_InoStartas_en/tab1/1.xlsx"  # <- Desired output filename

#         # Optional: enable verbose logging
#         logging.getLogger().setLevel(logging.DEBUG)

#         # Initialize the agent (Groq API key will be read from .env)
#         agent = EnhancedEUFormAgent()

#         print("🚀 Enhanced EU Form Filling Agent v2.0")
#         print("=" * 45)

#         # Process the form
#         result = agent.process_with_enhanced_features(
#             input_file=input_file,
#             excel_template=excel_template,
#             output_file=output_file,
#             validate_only=False
#         )

#         print(f"✅ Form successfully filled and saved to: {result}")
#         print(f"📊 Processing Statistics:")
#         print(f"   • Extracted entries: {agent.processing_stats['extracted_entries']}")
#         print(f"   • Filled cells: {agent.processing_stats['filled_cells']}")
#         print(f"   • Validation errors: {agent.processing_stats['validation_errors']}")

#         summary_path = result.replace('.xlsx', '_summary_report.txt')
#         if Path(summary_path).exists():
#             print(f"📄 Summary report: {summary_path}")

#     except Exception as e:
#         print(f"❌ Error: {e}")
#         logger.exception("Unexpected error")
#         sys.exit(1)


# if __name__ == "__main__":
#     main1()



import argparse
import json
import logging
import sys
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import pandas as pd
from groq import Groq
import openpyxl
from datetime import datetime
import os
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler('eu_form_agent.log'),
        logging.StreamHandler(sys.stdout)
    ]
)
logger = logging.getLogger(__name__)


class ValidationError(Exception):
    """Custom exception for validation errors"""
    pass


class FormFillingError(Exception):
    """Custom exception for form filling errors"""
    pass


class EnhancedEUFormAgent:
    """
    Enhanced AI-powered agent for filling EU project forms
    """
    
    def __init__(self, groq_api_key: str = None):
        """Initialize the enhanced agent"""
        self.api_key = groq_api_key or os.getenv("GROQ_API_KEY")
        if not self.api_key:
            raise ValueError("Groq API key not provided. Set GROQ_API_KEY environment variable.")
        
        self.client = Groq(api_key=self.api_key)
        # Using a valid Groq model
        self.model = "meta-llama/llama-4-scout-17b-16e-instruct"
        
        # Statistics tracking
        self.processing_stats = {
            "start_time": None,
            "end_time": None,
            "input_length": 0,
            "extracted_entries": 0,
            "filled_cells": 0,
            "validation_errors": 0
        }

    def create_enhanced_extraction_prompt(self, input_text: str, prompt_file_path: str = None) -> str:
        """Create an enhanced prompt by reading from file and including the input text"""
        
        # Default prompt file path if not provided
        if prompt_file_path is None:
            prompt_file_path = os.path.join("prompts", "1.txt")
        
        # Read the prompt from file
        try:
            with open(prompt_file_path, 'r', encoding='utf-8') as file:
                base_prompt = file.read()
            logger.info(f"Successfully loaded prompt from {prompt_file_path}")
        except FileNotFoundError:
            logger.error(f"Prompt file not found at: {prompt_file_path}")
            raise FileNotFoundError(f"Prompt file not found at: {prompt_file_path}")
        except Exception as e:
            logger.error(f"Failed to read prompt file: {e}")
            raise RuntimeError(f"Failed to read prompt file: {e}")
        
        # Add input text to the prompt
        full_prompt = f"{base_prompt}\n\nINPUT TEXT TO ANALYZE:\n{input_text}\n\nReturn ONLY the JSON object - no explanations, no additional text:"
        
        return full_prompt

    def extract_data_with_enhanced_llm(self, input_text: str, prompt_file_path: str = None) -> Dict[str, Any]:
        """Enhanced LLM extraction with retry logic and validation"""
        max_retries = 3
        
        for attempt in range(max_retries):
            try:
                # Create prompt with input text
                prompt = self.create_enhanced_extraction_prompt(input_text, prompt_file_path)
                
                logger.info(f"Attempt {attempt + 1}: Sending request to Groq LLM...")
                completion = self.client.chat.completions.create(
                    model=self.model,
                    messages=[
                        {
                            "role": "system",
                            "content": "You are a precise data extraction specialist for EU forms. Return only valid JSON."
                        },
                        {
                            "role": "user", 
                            "content": prompt
                        }
                    ],
                    temperature=0.3
                )
                
                response_content = completion.choices[0].message.content.strip()
                logger.info("Received response from Groq LLM")
                
                # Clean the response (remove any markdown formatting)
                if response_content.startswith("```json"):
                    response_content = response_content[7:-3].strip()
                elif response_content.startswith("```"):
                    response_content = response_content[3:-3].strip()
                
                # Parse JSON response
                extracted_data = json.loads(response_content)
                logger.info("Successfully parsed JSON response")
                
                # Validate the extracted data
                validation_result = self.validate_extracted_data_enhanced(extracted_data)
                if validation_result[0]:
                    return extracted_data
                else:
                    logger.warning(f"Validation failed on attempt {attempt + 1}: {validation_result[1]}")
                    if attempt == max_retries - 1:
                        raise ValidationError(f"Data validation failed after {max_retries} attempts: {validation_result[1]}")
                    
            except json.JSONDecodeError as e:
                logger.error(f"JSON parsing failed on attempt {attempt + 1}: {str(e)}")
                if attempt == max_retries - 1:
                    logger.error(f"Raw response: {response_content}")
                    raise
            except Exception as e:
                logger.error(f"LLM extraction failed on attempt {attempt + 1}: {str(e)}")
                if attempt == max_retries - 1:
                    raise
        
        raise Exception("Failed to extract data after maximum retries")

    def validate_extracted_data_enhanced(self, data: Dict[str, Any]) -> Tuple[bool, str]:
        """Enhanced validation with detailed error reporting"""
        errors = []
        
        # Check required top-level keys
        required_keys = ["project_info", "staff_costs", "mission_expenses", 
                        "equipment_depreciation", "rd_services", "materials_supplies",
                        "equipment_rental", "premises_rental"]
        
        for key in required_keys:
            if key not in data:
                errors.append(f"Missing required section: {key}")
        
        # Validate project_info
        if "project_info" in data:
            project_info = data["project_info"]
            if not isinstance(project_info, dict):
                errors.append("project_info must be a dictionary")
            else:
                # Check funding intensity
                if "funding_intensity" in project_info:
                    intensity = project_info["funding_intensity"]
                    if isinstance(intensity, (int, float)) and not (0 <= intensity <= 100):
                        errors.append(f"Funding intensity must be between 0-100, got {intensity}")
        
        # Validate staff costs
        if "staff_costs" in data and isinstance(data["staff_costs"], list):
            for i, staff in enumerate(data["staff_costs"]):
                if not isinstance(staff, dict):
                    errors.append(f"Staff cost entry {i} must be a dictionary")
                    continue
                
                # Validate financial consistency
                if all(k in staff for k in ["monthly_salary", "employer_costs", "duration_months"]):
                    expected_total = (staff["monthly_salary"] + staff["employer_costs"]) * staff["duration_months"]
                    if "total_salary_costs" in staff:
                        actual_total = staff["total_salary_costs"]
                        if abs(expected_total - actual_total) > 0.01:
                            errors.append(f"Staff {i}: Inconsistent salary calculation. Expected {expected_total}, got {actual_total}")
        
        # Validate mission expenses
        if "mission_expenses" in data and isinstance(data["mission_expenses"], dict):
            missions = data["mission_expenses"].get("missions", [])
            if isinstance(missions, list):
                for i, mission in enumerate(missions):
                    if "duration_days" in mission and isinstance(mission["duration_days"], (int, float)):
                        if mission["duration_days"] <= 0:
                            errors.append(f"Mission {i}: Duration must be positive")
        
        # Validate equipment depreciation
        if "equipment_depreciation" in data and isinstance(data["equipment_depreciation"], list):
            for i, equipment in enumerate(data["equipment_depreciation"]):
                if isinstance(equipment, dict):
                    # Check depreciation calculation
                    if all(k in equipment for k in ["acquisition_value", "residual_value", "depreciation_period_months"]):
                        expected_monthly = (equipment["acquisition_value"] - equipment["residual_value"]) / equipment["depreciation_period_months"]
                        if "monthly_depreciation" in equipment:
                            actual_monthly = equipment["monthly_depreciation"]
                            if abs(expected_monthly - actual_monthly) > 0.01:
                                errors.append(f"Equipment {i}: Inconsistent depreciation calculation")
        
        # Count validation errors for statistics
        self.processing_stats["validation_errors"] = len(errors)
        
        if errors:
            return False, "; ".join(errors)
        
        logger.info("Enhanced data validation passed")
        return True, "Validation successful"

    def generate_summary_report(self, data: Dict[str, Any], output_path: str) -> str:
        """Generate a summary report of the extracted data"""
        report_path = output_path.replace('.xlsx', '_summary_report.txt')
        
        try:
            with open(report_path, 'w', encoding='utf-8') as f:
                f.write("EU FORM FILLING AGENT - PROCESSING SUMMARY REPORT\n")
                f.write("=" * 55 + "\n\n")
                
                # Processing statistics
                f.write("PROCESSING STATISTICS:\n")
                f.write("-" * 22 + "\n")
                duration = (self.processing_stats["end_time"] - self.processing_stats["start_time"]).total_seconds()
                f.write(f"Processing Duration: {duration:.2f} seconds\n")
                f.write(f"Input Text Length: {self.processing_stats['input_length']} characters\n")
                f.write(f"Extracted Entries: {self.processing_stats['extracted_entries']}\n")
                f.write(f"Filled Cells: {self.processing_stats['filled_cells']}\n")
                f.write(f"Validation Errors: {self.processing_stats['validation_errors']}\n\n")
                
                # Project overview
                project_info = data.get("project_info", {})
                f.write("PROJECT OVERVIEW:\n")
                f.write("-" * 17 + "\n")
                f.write(f"Legal Entity: {project_info.get('legal_entity_name', 'N/A')}\n")
                f.write(f"Impact Name: {project_info.get('impact_name', 'N/A')}\n")
                f.write(f"R&D Type: {project_info.get('type_of_rd', 'N/A')}\n")
                f.write(f"Funding Intensity: {project_info.get('funding_intensity', 'N/A')}%\n\n")
                
                # Financial summary
                project_totals = data.get("project_totals", {})
                f.write("FINANCIAL SUMMARY:\n")
                f.write("-" * 18 + "\n")
                f.write(f"Total Eligible Costs: {project_totals.get('total_eligible_costs', 'N/A')} EUR\n")
                f.write(f"Total Funding Requested: {project_totals.get('total_funding_requested', 'N/A')} EUR\n")
                f.write(f"Own Contribution: {project_totals.get('own_contribution', 'N/A')} EUR\n\n")
                
                # Section breakdown
                sections = [
                    ("Staff Costs", "staff_costs"),
                    ("Mission Expenses", "mission_expenses"),
                    ("Equipment Depreciation", "equipment_depreciation"),
                    ("R&D Services", "rd_services"),
                    ("Materials & Supplies", "materials_supplies"),
                    ("Equipment Rental", "equipment_rental"),
                    ("Premises Rental", "premises_rental")
                ]
                
                f.write("SECTION BREAKDOWN:\n")
                f.write("-" * 18 + "\n")
                for section_name, section_key in sections:
                    section_data = data.get(section_key, [])
                    if isinstance(section_data, list):
                        count = len(section_data)
                    elif isinstance(section_data, dict) and section_key == "mission_expenses":
                        count = len(section_data.get("missions", []))
                    elif isinstance(section_data, dict):
                        count = 1
                    else:
                        count = 0
                    f.write(f"{section_name}: {count} entries\n")
                
                f.write(f"\nReport generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
            
            logger.info(f"Summary report generated: {report_path}")
            return report_path
            
        except Exception as e:
            logger.error(f"Error generating summary report: {str(e)}")
            return None

    def validate_excel_template(self, excel_path: str) -> bool:
        """Validate that the Excel template has the expected structure"""
        try:
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            # Check for key cells that should exist in the template
            test_cells = ["D1", "D2", "B10", "G30", "B102", "B123"]
            for cell in test_cells:
                try:
                    _ = worksheet[cell]
                except Exception:
                    logger.error(f"Template validation failed: Cell {cell} not accessible")
                    return False
            
            logger.info("Excel template validation passed")
            return True
            
        except Exception as e:
            logger.error(f"Error validating Excel template: {str(e)}")
            return False

    def process_with_enhanced_features(self, input_file: str, excel_template: str, 
                                     output_file: str = None, validate_only: bool = False,
                                     prompt_file_path: str = None) -> str:
        """Enhanced processing with additional features"""
        try:
            self.processing_stats["start_time"] = datetime.now()
            logger.info("Starting enhanced form processing...")
            
            # Validate inputs
            if not Path(input_file).exists():
                raise FileNotFoundError(f"Input file not found: {input_file}")
            
            if not Path(excel_template).exists():
                raise FileNotFoundError(f"Excel template not found: {excel_template}")
            
            # Validate Excel template
            if not self.validate_excel_template(excel_template):
                raise ValidationError("Excel template validation failed")
            
            # Read input data
            input_text = self.read_input_data(input_file)
            self.processing_stats["input_length"] = len(input_text)
            
            # Extract structured data using enhanced LLM
            extracted_data = self.extract_data_with_enhanced_llm(input_text, prompt_file_path)
            
            # Count extracted entries
            total_entries = 0
            for key, value in extracted_data.items():
                if isinstance(value, list):
                    total_entries += len(value)
                elif isinstance(value, dict) and key == "mission_expenses":
                    total_entries += len(value.get("missions", []))
                elif isinstance(value, dict):
                    total_entries += 1
            self.processing_stats["extracted_entries"] = total_entries
            
            if validate_only:
                logger.info("Validation-only mode completed successfully")
                return "Validation completed successfully"
            
            # Fill Excel form with enhanced error handling
            output_path = self.fill_excel_form_enhanced(extracted_data, excel_template, output_file)
            
            # Generate summary report
            self.generate_summary_report(extracted_data, output_path)
            
            self.processing_stats["end_time"] = datetime.now()
            logger.info("Enhanced form processing completed successfully!")
            
            return output_path
            
        except Exception as e:
            self.processing_stats["end_time"] = datetime.now()
            logger.error(f"Error in enhanced form processing: {str(e)}")
            raise

    def fill_excel_form_enhanced(self, data: Dict[str, Any], excel_path: str, output_path: str = None) -> str:
        """Enhanced Excel form filling with better error handling and cell tracking"""
        try:
            workbook = openpyxl.load_workbook(excel_path)
            worksheet = workbook.active
            
            filled_cells = 0
            logger.info("Filling Excel form with extracted data...")
            
            # Fill each section and count filled cells
            filled_cells += self._fill_project_info_enhanced(worksheet, data.get("project_info", {}))
            filled_cells += self._fill_staff_costs_enhanced(worksheet, data.get("staff_costs", []))
            filled_cells += self._fill_mission_expenses_enhanced(worksheet, data.get("mission_expenses", {}))
            filled_cells += self._fill_equipment_depreciation_enhanced(worksheet, data.get("equipment_depreciation", []))
            filled_cells += self._fill_rd_services_enhanced(worksheet, data.get("rd_services", []))
            filled_cells += self._fill_materials_supplies_enhanced(worksheet, data.get("materials_supplies", []))
            filled_cells += self._fill_equipment_rental_enhanced(worksheet, data.get("equipment_rental", []))
            filled_cells += self._fill_premises_rental_enhanced(worksheet, data.get("premises_rental", []))
            
            self.processing_stats["filled_cells"] = filled_cells
            
            # Generate output filename if not provided
            if output_path is None:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_path = f"filled_eu_form_{timestamp}.xlsx"
            
            # Save the workbook
            workbook.save(output_path)
            logger.info(f"Successfully saved filled form to {output_path}")
            logger.info(f"Total cells filled: {filled_cells}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error in enhanced Excel form filling: {str(e)}")
            raise FormFillingError(f"Failed to fill Excel form: {str(e)}")

    def _fill_project_info_enhanced(self, worksheet, project_info: Dict[str, Any]) -> int:
        """Enhanced project info filling with cell counting"""
        filled_count = 0
        try:
            mappings = [
                ("type_of_rd", "D1"),
                ("project_impact_no", "D2"),
                ("impact_name", "D3"),
                ("legal_entity_name", "D4"),
                ("applicant_partner", "D5"),
                ("funding_intensity", "H5")
            ]
            
            for field, cell in mappings:
                if project_info.get(field) is not None:
                    worksheet[cell] = project_info[field]
                    filled_count += 1
                    
            logger.info(f"Project information: {filled_count} cells filled")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling project info: {str(e)}")
            return filled_count

    def _fill_staff_costs_enhanced(self, worksheet, staff_costs: List[Dict[str, Any]]) -> int:
        """Enhanced staff costs filling"""
        filled_count = 0
        try:
            start_row = 10
            for i, staff in enumerate(staff_costs[:20]):
                row = start_row + i
                
                mappings = [
                    ("position", f"B{row}"),
                    ("monthly_salary", f"C{row}"),
                    ("employer_costs", f"D{row}"),
                    ("duration_months", f"E{row}"),
                    ("hourly_rate", f"F{row}"),
                    ("eligible_costs", f"G{row}"),
                    ("funding_requested", f"H{row}"),
                    ("supporting_docs", f"I{row}")
                ]
                
                for field, cell in mappings:
                    if staff.get(field) is not None:
                        worksheet[cell] = staff[field]
                        filled_count += 1
                        
            logger.info(f"Staff costs: {filled_count} cells filled for {len(staff_costs)} staff members")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling staff costs: {str(e)}")
            return filled_count

    def _fill_mission_expenses_enhanced(self, worksheet, mission_expenses: Dict[str, Any]) -> int:
        """Enhanced mission expenses filling"""
        filled_count = 0
        try:
            # Fill totals
            if mission_expenses.get("total_eligible_costs") is not None:
                worksheet["G30"] = mission_expenses["total_eligible_costs"]
                filled_count += 1
            if mission_expenses.get("total_funding_requested") is not None:
                worksheet["H30"] = mission_expenses["total_funding_requested"]
                filled_count += 1
                
            # Fill individual missions
            missions = mission_expenses.get("missions", [])
            start_row = 31
            
            for i, mission in enumerate(missions[:10]):
                base_row = start_row + (i * 7)
                
                if mission.get("mission_name"):
                    worksheet[f"B{base_row}"] = mission["mission_name"]
                    filled_count += 1
                if mission.get("destination_country"):
                    worksheet[f"J{base_row}"] = mission["destination_country"]
                    filled_count += 1
                if mission.get("duration_days"):
                    worksheet[f"J{base_row+1}"] = mission["duration_days"]
                    filled_count += 1
                if mission.get("travelers_count"):
                    worksheet[f"J{base_row+2}"] = mission["travelers_count"]
                    filled_count += 1
                    
            logger.info(f"Mission expenses: {filled_count} cells filled for {len(missions)} missions")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling mission expenses: {str(e)}")
            return filled_count

    def _fill_equipment_depreciation_enhanced(self, worksheet, equipment: List[Dict[str, Any]]) -> int:
        """Enhanced equipment depreciation filling"""
        filled_count = 0
        try:
            start_row = 102
            for i, equip in enumerate(equipment[:20]):
                row = start_row + i
                
                mappings = [
                    ("equipment_name", f"B{row}"),
                    ("acquisition_date", f"K{row}"),
                    ("acquisition_value", f"L{row}"),
                    ("depreciation_period_months", f"M{row}"),
                    ("residual_value", f"N{row}"),
                    ("monthly_depreciation", f"O{row}"),
                    ("project_usage_months", f"P{row}"),
                    ("usage_percentage", f"Q{row}"),
                    ("project_depreciation_amount", f"R{row}")
                ]
                
                for field, cell in mappings:
                    if equip.get(field) is not None:
                        worksheet[cell] = equip[field]
                        filled_count += 1
                        
            logger.info(f"Equipment depreciation: {filled_count} cells filled for {len(equipment)} items")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling equipment depreciation: {str(e)}")
            return filled_count

    def _fill_rd_services_enhanced(self, worksheet, rd_services: List[Dict[str, Any]]) -> int:
        """Enhanced R&D services filling"""
        filled_count = 0
        try:
            start_row = 123
            for i, service in enumerate(rd_services[:10]):
                row = start_row + i
                
                mappings = [
                    ("service_name", f"B{row}"),
                    ("units", f"D{row}"),
                    ("quantity", f"E{row}"),
                    ("unit_price", f"F{row}"),
                    ("eligible_costs", f"G{row}"),
                    ("funding_requested", f"H{row}"),
                    ("supporting_docs", f"I{row}")
                ]
                
                for field, cell in mappings:
                    if service.get(field) is not None:
                        worksheet[cell] = service[field]
                        filled_count += 1
                        
            logger.info(f"R&D services: {filled_count} cells filled for {len(rd_services)} services")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling R&D services: {str(e)}")
            return filled_count

    def _fill_materials_supplies_enhanced(self, worksheet, materials: List[Dict[str, Any]]) -> int:
        """Enhanced materials and supplies filling"""
        filled_count = 0
        try:
            start_row = 145
            for i, material in enumerate(materials[:25]):
                row = start_row + i
                
                mappings = [
                    ("item_name", f"B{row}"),
                    ("units", f"D{row}"),
                    ("quantity", f"E{row}"),
                    ("unit_price", f"F{row}"),
                    ("eligible_costs", f"G{row}"),
                    ("funding_requested", f"H{row}"),
                    ("supporting_docs", f"I{row}")
                ]
                
                for field, cell in mappings:
                    if material.get(field) is not None:
                        worksheet[cell] = material[field]
                        filled_count += 1
                        
            logger.info(f"Materials and supplies: {filled_count} cells filled for {len(materials)} items")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling materials and supplies: {str(e)}")
            return filled_count

    def _fill_equipment_rental_enhanced(self, worksheet, equipment_rental: List[Dict[str, Any]]) -> int:
        """Enhanced equipment rental filling"""
        filled_count = 0
        try:
            start_row = 171
            for i, rental in enumerate(equipment_rental[:10]):
                row = start_row + i
                
                mappings = [
                    ("equipment_name", f"B{row}"),
                    ("units", f"D{row}"),
                    ("quantity", f"E{row}"),
                    ("unit_price", f"F{row}"),
                    ("eligible_costs", f"G{row}"),
                    ("funding_requested", f"H{row}"),
                    ("monthly_rental_cost", f"K{row}"),
                    ("usage_duration_months", f"L{row}")
                ]
                
                for field, cell in mappings:
                    if rental.get(field) is not None:
                        worksheet[cell] = rental[field]
                        filled_count += 1
                        
            logger.info(f"Equipment rental: {filled_count} cells filled for {len(equipment_rental)} items")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling equipment rental: {str(e)}")
            return filled_count

    def _fill_premises_rental_enhanced(self, worksheet, premises_rental: List[Dict[str, Any]]) -> int:
        """Enhanced premises rental filling"""
        filled_count = 0
        try:
            start_row = 182
            for i, premises in enumerate(premises_rental[:5]):
                row = start_row + i
                
                mappings = [
                    ("premises_address", f"B{row}"),
                    ("units", f"D{row}"),
                    ("quantity", f"E{row}"),
                    ("unit_price", f"F{row}"),
                    ("eligible_costs", f"G{row}"),
                    ("funding_requested", f"H{row}"),
                    ("monthly_rental_cost", f"K{row}"),
                    ("usage_duration_months", f"L{row}")
                ]
                
                for field, cell in mappings:
                    if premises.get(field) is not None:
                        worksheet[cell] = premises[field]
                        filled_count += 1
                        
            logger.info(f"Premises rental: {filled_count} cells filled for {len(premises_rental)} premises")
            return filled_count
            
        except Exception as e:
            logger.error(f"Error filling premises rental: {str(e)}")
            return filled_count

    def read_input_data(self, file_path: str) -> str:
        """Read and return content from the input text file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"Successfully read {len(content)} characters from {file_path}")
            return content
        except FileNotFoundError:
            logger.error(f"File {file_path} not found")
            raise
        except Exception as e:
            logger.error(f"Error reading file {file_path}: {str(e)}")
            raise


def create_cli_parser() -> argparse.ArgumentParser:
    """Create command-line interface parser"""
    parser = argparse.ArgumentParser(
        description='Enhanced EU Form Filling Agent using AI',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python enhanced_agent.py --input data.txt --template form.xlsx
  python enhanced_agent.py --input data.txt --template form.xlsx --output my_form.xlsx
  python enhanced_agent.py --input data.txt --template form.xlsx --validate-only
  python enhanced_agent.py --input data.txt --template form.xlsx --api-key YOUR_KEY
        """
    )
    
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='Path to input text file containing project data'
    )
    
    parser.add_argument(
        '--template', '-t',
        required=True,
        help='Path to Excel template file (ENGLISH_1A priedas_InoStartas en.xlsx)'
    )
    
    parser.add_argument(
        '--output', '-o',
        help='Path for output filled Excel file (default: auto-generated with timestamp)'
    )
    
    parser.add_argument(
        '--prompt-file', '-p',
        help='Path to custom prompt file (default: prompts/1.txt)'
    )
    
    parser.add_argument(
        '--api-key',
        help='Groq API key (default: reads from GROQ_API_KEY environment variable)'
    )
    
    parser.add_argument(
        '--validate-only',
        action='store_true',
        help='Only validate extracted data without filling the form'
    )
    
    parser.add_argument(
        '--verbose', '-v',
        action='store_true',
        help='Enable verbose logging'
    )
    
    parser.add_argument(
        '--version',
        action='version',
        version='Enhanced EU Form Agent v2.0'
    )
    
    return parser


def main():
    """Main function with command line argument parsing"""
    parser = create_cli_parser()
    args = parser.parse_args()
    
    # Set logging level
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)
    
    try:
        # Initialize the agent
        agent = EnhancedEUFormAgent(groq_api_key=args.api_key)
        
        print("🚀 Enhanced EU Form Filling Agent v2.0")
        print("=" * 45)
        
        # Process the form
        result = agent.process_with_enhanced_features(
            input_file=args.input,
            excel_template=args.template,
            output_file=args.output,
            validate_only=args.validate_only,
            prompt_file_path=args.prompt_file
        )
        
        if args.validate_only:
            print(f"✅ Validation completed: {result}")
        else:
            print(f"✅ Form successfully filled and saved to: {result}")
        
        print(f"📊 Processing Statistics:")
        print(f"   • Input length: {agent.processing_stats['input_length']} characters")
        print(f"   • Extracted entries: {agent.processing_stats['extracted_entries']}")
        print(f"   • Filled cells: {agent.processing_stats['filled_cells']}")
        print(f"   • Validation errors: {agent.processing_stats['validation_errors']}")
        
        if not args.validate_only:
            summary_path = result.replace('.xlsx', '_summary_report.txt')
            if Path(summary_path).exists():
                print(f"📄 Summary report: {summary_path}")
        
    except Exception as e:
        print(f"❌ Error: {e}")
        logger.exception("Unexpected error")
        sys.exit(1)


def main1(data_file_path):
    """Main function with hardcoded input/output paths for testing"""
    try:
        
        excel_template = "code/agents/ENGLISH_1A_priedas_InoStartas_en/tab1/input1.xlsx"
        output_file = "code/agents/ENGLISH_1A_priedas_InoStartas_en/tab1/1.xlsx"
        prompt_file = "code/prompts/1.txt"

        # Optional: enable verbose logging
        logging.getLogger().setLevel(logging.DEBUG)

        # Initialize the agent (Groq API key will be read from .env)
        agent = EnhancedEUFormAgent()

        print("🚀 Enhanced EU Form Filling Agent v2.0")
        print("=" * 45)

        # Process the form
        result = agent.process_with_enhanced_features(
            input_file=data_file_path,
            excel_template=excel_template,
            output_file=output_file,
            validate_only=False,
            prompt_file_path=prompt_file
        )

        print(f"✅ Form successfully filled and saved to: {result}")
        print(f"📊 Processing Statistics:")
        print(f"   • Input length: {agent.processing_stats['input_length']} characters")
        print(f"   • Extracted entries: {agent.processing_stats['extracted_entries']}")
        print(f"   • Filled cells: {agent.processing_stats['filled_cells']}")
        print(f"   • Validation errors: {agent.processing_stats['validation_errors']}")

        summary_path = result.replace('.xlsx', '_summary_report.txt')
        if Path(summary_path).exists():
            print(f"📄 Summary report: {summary_path}")

    except Exception as e:
        print(f"❌ Error: {e}")
        logger.exception("Unexpected error")
        sys.exit(1)


