import openpyxl
import json
import os
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass
from dotenv import load_dotenv
from groq import Groq

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class RDImpactEntry:
    """Data class for R&D activities by impact (Table 1)"""
    serial_no: int
    impact_name: str
    direct_costs: float
    indirect_costs: float
    total_costs: float
    funding_requested: float

@dataclass
class RDExpenditureCategory:
    """Data class for R&D expenditure categories (Table 2)"""
    category_id: int
    category_title: str
    direct_costs: float
    funding_requested: float

@dataclass
class PartnerBreakdown:
    """Data class for partner expenditure breakdown (Table 3)"""
    partner_name: str
    eligible_costs: float
    funding_requested: float
    percentage: float

@dataclass
class RDExpenditureData:
    """Complete R&D expenditure data structure"""
    # Table 1: R&D activities by impact (max 10 entries)
    impact_entries: List[RDImpactEntry]
    
    # Table 2: R&D expenditure categories (8 fixed categories)
    expenditure_categories: List[RDExpenditureCategory]
    indirect_cost_rate: float  # 0% or 7%
    
    # Table 3: Partner breakdown (applicant + up to 3 partners)
    partner_breakdown: List[PartnerBreakdown]
    
    # Additional calculations
    lines_4_5_amount: float
    lines_4_5_percentage: float
    heading_8_amount: float
    heading_8_percentage: float
    
    def __post_init__(self):
        # Ensure maximum 10 impact entries
        self.impact_entries = self.impact_entries[:10]
        # Ensure exactly 8 expenditure categories
        while len(self.expenditure_categories) < 8:
            self.expenditure_categories.append(RDExpenditureCategory(
                category_id=len(self.expenditure_categories) + 1,
                category_title="Not applicable",
                direct_costs=0.0,
                funding_requested=0.0
            ))
        self.expenditure_categories = self.expenditure_categories[:8]
        # Ensure maximum 4 partners (applicant + 3 partners)
        self.partner_breakdown = self.partner_breakdown[:4]

class GroqRDAnalyzer:
    """Analyzes R&D expenditure data using Groq LLM with streaming"""
    
    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key)
    
    def create_analysis_prompt(self, prompt_path: str = "code/prompts/summary.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")
    
    def analyze_rd_data_streaming(self, text_content: str) -> RDExpenditureData:
        """Analyze text using Groq streaming and extract R&D expenditure data"""
        
        logger.info("Starting R&D expenditure data analysis with streaming...")
        
        prompt = self.create_analysis_prompt()
        full_content = f"{prompt}\n\nText to analyze:\n{text_content}"
        
        try:
            # Use streaming as specified in requirements
            completion = self.client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[
                    {
                        "role": "user",
                        "content": full_content
                    }
                ],
                temperature=0.3,
                stream=True
            )
            
            # Collect streaming response
            analysis_result = ""
            logger.info("Receiving streaming response from Groq API...")
            
            for chunk in completion:
                if chunk.choices[0].delta.content:
                    content = chunk.choices[0].delta.content
                    analysis_result += content
            
            logger.info(f"Received complete analysis ({len(analysis_result)} characters)")
            
            # Clean and parse JSON
            json_content = self._extract_json_from_response(analysis_result)
            parsed_data = json.loads(json_content)
            
            # Convert to RDExpenditureData object
            rd_data = self._convert_to_rd_data(parsed_data)
            
            return rd_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            logger.error(f"Raw response: {analysis_result[:500]}...")
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_json_from_response(self, response: str) -> str:
        """Extract JSON from response, handling markdown code blocks"""
        
        response = response.strip()
        
        # Remove markdown code blocks if present
        if '```json' in response:
            start_marker = '```json'
            end_marker = '```'
            start_index = response.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = response.find(end_marker, start_index)
                if end_index != -1:
                    return response[start_index:end_index].strip()
        
        if response.startswith('```') and response.endswith('```'):
            return response[3:-3].strip()
        
        return response
    
    def _convert_to_rd_data(self, parsed_data: Dict) -> RDExpenditureData:
        """Convert parsed JSON to RDExpenditureData object"""
        
        # Convert impact entries
        impact_entries = []
        for entry_data in parsed_data.get('impact_entries', [])[:10]:  # Max 10 entries
            entry = RDImpactEntry(
                serial_no=int(entry_data.get('serial_no', 1)),
                impact_name=entry_data.get('impact_name', 'R&D Activity'),
                direct_costs=float(entry_data.get('direct_costs', 0.0)),
                indirect_costs=float(entry_data.get('indirect_costs', 0.0)),
                total_costs=float(entry_data.get('total_costs', 0.0)),
                funding_requested=float(entry_data.get('funding_requested', 0.0))
            )
            impact_entries.append(entry)
        
        # Convert expenditure categories
        expenditure_categories = []
        category_titles = [
            "Salaries and wages of project staff and employer's liability costs",
            "Mission expenses for project staff",
            "Depreciation costs for tools and equipment",
            "Expenditure for the acquisition of R&D services",
            "Costs of contractual research, know-how and patents purchased or licensed from external sources",
            "Materials, non-expendable inventories, stocks, etc., classified as current assets",
            "Rental costs of equipment",
            "Rental expenses for buildings or premises allocated to activities"
        ]
        
        for i, cat_data in enumerate(parsed_data.get('expenditure_categories', [])[:8]):
            category = RDExpenditureCategory(
                category_id=i + 1,
                category_title=category_titles[i] if i < len(category_titles) else f"Category {i+1}",
                direct_costs=float(cat_data.get('direct_costs', 0.0)),
                funding_requested=float(cat_data.get('funding_requested', 0.0))
            )
            expenditure_categories.append(category)
        
        # Ensure we have exactly 8 categories
        while len(expenditure_categories) < 8:
            i = len(expenditure_categories)
            category = RDExpenditureCategory(
                category_id=i + 1,
                category_title=category_titles[i] if i < len(category_titles) else f"Category {i+1}",
                direct_costs=0.0,
                funding_requested=0.0
            )
            expenditure_categories.append(category)
        
        # Convert partner breakdown
        partner_breakdown = []
        partner_names = ["Applicant", "Partner No 1", "Partner No 2", "Partner No 3"]
        
        for i, partner_data in enumerate(parsed_data.get('partner_breakdown', [])[:4]):
            partner = PartnerBreakdown(
                partner_name=partner_names[i] if i < len(partner_names) else f"Partner {i}",
                eligible_costs=float(partner_data.get('eligible_costs', 0.0)),
                funding_requested=float(partner_data.get('funding_requested', 0.0)),
                percentage=float(partner_data.get('percentage', 0.0))
            )
            partner_breakdown.append(partner)
        
        # Ensure we have exactly 4 partner entries
        while len(partner_breakdown) < 4:
            i = len(partner_breakdown)
            partner = PartnerBreakdown(
                partner_name=partner_names[i] if i < len(partner_names) else f"Partner {i}",
                eligible_costs=0.0,
                funding_requested=0.0,
                percentage=0.0
            )
            partner_breakdown.append(partner)
        
        rd_data = RDExpenditureData(
            impact_entries=impact_entries,
            expenditure_categories=expenditure_categories,
            indirect_cost_rate=float(parsed_data.get('indirect_cost_rate', 0.07)),
            partner_breakdown=partner_breakdown,
            lines_4_5_amount=float(parsed_data.get('lines_4_5_amount', 0.0)),
            lines_4_5_percentage=float(parsed_data.get('lines_4_5_percentage', 0.0)),
            heading_8_amount=float(parsed_data.get('heading_8_amount', 0.0)),
            heading_8_percentage=float(parsed_data.get('heading_8_percentage', 0.0))
        )
        
        return rd_data

class RDExpenditureExcelFiller:
    """Fills Excel form with R&D expenditure data according to exact cell mappings"""
    
    def __init__(self):
        pass
    
    def fill_excel_form(self, template_path: str, output_path: str, rd_data: RDExpenditureData) -> bool:
        """Fill Excel form with R&D expenditure data"""
        
        try:
            logger.info(f"Loading Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            # Use first sheet or find R&D-related sheet
            sheet_name = workbook.sheetnames[0]
            for name in workbook.sheetnames:
                if any(keyword in name.lower() for keyword in ['rd', 'r&d', 'expenditure', 'budget']):
                    sheet_name = name
                    break
            
            sheet = workbook[sheet_name]
            logger.info(f"Using sheet: {sheet_name}")
            
            # Fill Table 1: R&D Activities by Impact
            self._fill_table1_impacts(sheet, rd_data.impact_entries)
            
            # Fill Table 2: R&D Expenditure Categories
            self._fill_table2_categories(sheet, rd_data.expenditure_categories, rd_data.indirect_cost_rate)
            
            # Fill additional calculations for Table 2
            self._fill_table2_calculations(sheet, rd_data)
            
            # Fill Table 3: Partner Breakdown
            self._fill_table3_partners(sheet, rd_data.partner_breakdown)
            
            # Save the filled workbook
            workbook.save(output_path)
            logger.info(f"✅ Successfully saved Excel form: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to fill Excel form: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _fill_table1_impacts(self, sheet, impact_entries: List[RDImpactEntry]):
        """Fill Table 1: R&D Activities by Impact (A5:F15)"""
        
        logger.info("Filling Table 1: R&D Activities by Impact...")
        
        total_direct = 0.0
        total_indirect = 0.0
        total_combined = 0.0
        total_funding = 0.0
        
        for i, entry in enumerate(impact_entries):
            if i < 10:  # Maximum 10 entries (rows 5-14)
                row = 5 + i
                
                # Fill data
                sheet[f"A{row}"] = entry.serial_no
                sheet[f"B{row}"] = entry.impact_name
                sheet[f"C{row}"] = entry.direct_costs
                sheet[f"D{row}"] = entry.indirect_costs
                sheet[f"E{row}"] = entry.total_costs
                sheet[f"F{row}"] = entry.funding_requested
                
                # Add to totals
                total_direct += entry.direct_costs
                total_indirect += entry.indirect_costs
                total_combined += entry.total_costs
                total_funding += entry.funding_requested
                
                logger.info(f"✅ Impact {entry.serial_no}: {entry.impact_name} - €{entry.total_costs:,.2f}")
        
        # Fill totals in row 15
        sheet["C15"] = total_direct
        sheet["D15"] = total_indirect
        sheet["E15"] = total_combined
        sheet["F15"] = total_funding
        
        logger.info(f"✅ Table 1 Totals - Direct: €{total_direct:,.2f}, Total: €{total_combined:,.2f}")
    
    def _fill_table2_categories(self, sheet, categories: List[RDExpenditureCategory], indirect_rate: float):
        """Fill Table 2: R&D Expenditure Categories (C19:D26)"""
        
        logger.info("Filling Table 2: R&D Expenditure Categories...")
        
        total_direct = 0.0
        total_funding = 0.0
        
        for i, category in enumerate(categories):
            if i < 8:  # Exactly 8 categories (rows 19-26)
                row = 19 + i
                
                # Fill data
                sheet[f"C{row}"] = category.direct_costs
                sheet[f"D{row}"] = category.funding_requested
                
                # Add to totals
                total_direct += category.direct_costs
                total_funding += category.funding_requested
                
                logger.info(f"✅ Category {category.category_id}: €{category.direct_costs:,.2f}")
        
        # Fill totals in row 27
        sheet["C27"] = total_direct
        sheet["D27"] = total_funding
        
        # Fill indirect cost rate and calculations
        sheet["B29"] = f"{indirect_rate * 100:.1f}%"  # Indirect cost rate percentage
        
        # Calculate indirect costs
        indirect_amount = total_direct * indirect_rate
        sheet["C31"] = indirect_amount  # Indirect expenditure amount
        sheet["D31"] = indirect_amount * 0.85  # Funding for indirect (typically 85%)
        
        # Calculate project budget (direct + indirect)
        project_budget_direct = total_direct + indirect_amount
        project_budget_funding = total_funding + (indirect_amount * 0.85)
        
        sheet["C33"] = project_budget_direct
        sheet["D33"] = project_budget_funding
        
        logger.info(f"✅ Table 2 Totals - Direct: €{total_direct:,.2f}, Budget: €{project_budget_direct:,.2f}")
    
    def _fill_table2_calculations(self, sheet, rd_data: RDExpenditureData):
        """Fill additional calculations for Table 2"""
        
        # Lines 4 and 5 (R&D services + contractual research)
        sheet["C36"] = rd_data.lines_4_5_amount
        sheet["D36"] = f"{rd_data.lines_4_5_percentage:.1f}%"
        
        # Heading 8 (building/premises rental)
        sheet["C37"] = rd_data.heading_8_amount
        sheet["D37"] = f"{rd_data.heading_8_percentage:.1f}%"
        
        logger.info(f"✅ Additional calculations - Lines 4&5: €{rd_data.lines_4_5_amount:,.2f}")
    
    def _fill_table3_partners(self, sheet, partners: List[PartnerBreakdown]):
        """Fill Table 3: Partner Breakdown (C41:D50)"""
        
        logger.info("Filling Table 3: Partner Breakdown...")
        
        total_eligible = 0.0
        total_funding = 0.0
        total_percentage = 0.0
        
        partner_rows = [41, 43, 45, 47]  # Rows for each partner
        percentage_rows = [42, 44, 46, 48]  # Rows for percentages
        
        for i, partner in enumerate(partners):
            if i < 4:  # Maximum 4 partners
                eligible_row = partner_rows[i]
                percentage_row = percentage_rows[i]
                
                # Fill eligible costs and funding
                sheet[f"C{eligible_row}"] = partner.eligible_costs
                sheet[f"D{eligible_row}"] = partner.funding_requested
                
                # Fill percentage
                sheet[f"C{percentage_row}"] = f"{partner.percentage:.1f}%"
                
                # Add to totals
                total_eligible += partner.eligible_costs
                total_funding += partner.funding_requested
                total_percentage += partner.percentage
                
                logger.info(f"✅ {partner.partner_name}: €{partner.eligible_costs:,.2f} ({partner.percentage:.1f}%)")
        
        # Fill totals
        sheet["C49"] = f"{total_percentage:.1f}%"  # Total percentage
        sheet["C50"] = total_eligible  # Total eligible costs
        sheet["D50"] = total_funding   # Total funding requested
        
        logger.info(f"✅ Table 3 Totals - Eligible: €{total_eligible:,.2f}, Funding: €{total_funding:,.2f}")

class RDExpenditureAgent:
    """Main orchestrator for R&D expenditure processing with streaming"""
    
    def __init__(self, groq_api_key: str):
        self.analyzer = GroqRDAnalyzer(groq_api_key)
        self.excel_filler = RDExpenditureExcelFiller()
    
    def read_data_file(self, file_path: str) -> str:
        """Read content from data file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"✅ Successfully read data file: {file_path} ({len(content)} characters)")
            return content
        except FileNotFoundError:
            logger.error(f"❌ Data file not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"❌ Error reading data file: {e}")
            raise
    
    def process_rd_expenditure(self, data_file_path: str, template_path: str, output_path: str) -> Dict[str, Any]:
        """Complete workflow for processing R&D expenditure"""
        
        results = {
            'success': False,
            'steps_completed': [],
            'rd_data': None,
            'statistics': {},
            'errors': []
        }
        
        try:
            # Step 1: Read data file
            logger.info("🔄 STEP 1: Reading Data File")
            text_content = self.read_data_file(data_file_path)
            results['steps_completed'].append('data_file_read')
            
            # Step 2: Analyze with Groq API (using streaming)
            logger.info("🔄 STEP 2: Analyzing R&D Expenditure with Groq Streaming API")
            rd_data = self.analyzer.analyze_rd_data_streaming(text_content)
            results['rd_data'] = rd_data
            results['steps_completed'].append('data_analyzed')
            
            # Step 3: Generate statistics
            logger.info("🔄 STEP 3: Generating Statistics")
            statistics = self._generate_statistics(rd_data)
            results['statistics'] = statistics
            results['steps_completed'].append('statistics_generated')
            
            # Step 4: Fill Excel form
            logger.info("🔄 STEP 4: Filling Excel Form")
            excel_success = self.excel_filler.fill_excel_form(template_path, output_path, rd_data)
            
            if excel_success:
                results['steps_completed'].append('excel_filled')
                results['success'] = True
                logger.info("✅ PROCESS COMPLETED SUCCESSFULLY")
            else:
                results['errors'].append('Excel filling failed')
                
        except Exception as e:
            error_msg = f"❌ Process failed at step {len(results['steps_completed']) + 1}: {e}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _generate_statistics(self, rd_data: RDExpenditureData) -> Dict[str, Any]:
        """Generate comprehensive statistics"""
        
        # Table 1 statistics
        total_impact_entries = len(rd_data.impact_entries)
        total_direct_impacts = sum(entry.direct_costs for entry in rd_data.impact_entries)
        total_indirect_impacts = sum(entry.indirect_costs for entry in rd_data.impact_entries)
        total_funding_impacts = sum(entry.funding_requested for entry in rd_data.impact_entries)
        
        # Table 2 statistics
        total_direct_categories = sum(cat.direct_costs for cat in rd_data.expenditure_categories)
        total_funding_categories = sum(cat.funding_requested for cat in rd_data.expenditure_categories)
        indirect_amount = total_direct_categories * rd_data.indirect_cost_rate
        project_budget = total_direct_categories + indirect_amount
        
        # Table 3 statistics
        total_eligible_partners = sum(partner.eligible_costs for partner in rd_data.partner_breakdown)
        total_funding_partners = sum(partner.funding_requested for partner in rd_data.partner_breakdown)
        active_partners = len([p for p in rd_data.partner_breakdown if p.eligible_costs > 0])
        
        # Funding efficiency
        funding_rate = (total_funding_categories / total_direct_categories * 100) if total_direct_categories > 0 else 0
        
        statistics = {
            'table1_impacts': {
                'total_entries': total_impact_entries,
                'total_direct_costs': total_direct_impacts,
                'total_indirect_costs': total_indirect_impacts,
                'total_funding_requested': total_funding_impacts
            },
            'table2_categories': {
                'total_direct_costs': total_direct_categories,
                'total_funding_requested': total_funding_categories,
                'indirect_cost_rate': rd_data.indirect_cost_rate * 100,
                'indirect_amount': indirect_amount,
                'project_budget': project_budget,
                'funding_efficiency': funding_rate
            },
            'table3_partners': {
                'active_partners': active_partners,
                'total_eligible_costs': total_eligible_partners,
                'total_funding_requested': total_funding_partners,
                'partner_distribution': [
                    {
                        'name': partner.partner_name,
                        'eligible_costs': partner.eligible_costs,
                        'percentage': partner.percentage
                    } for partner in rd_data.partner_breakdown if partner.eligible_costs > 0
                ]
            },
            'special_calculations': {
                'lines_4_5_amount': rd_data.lines_4_5_amount,
                'lines_4_5_percentage': rd_data.lines_4_5_percentage,
                'heading_8_amount': rd_data.heading_8_amount,
                'heading_8_percentage': rd_data.heading_8_percentage
            },
            'compliance_check': {
                'minimum_funding_met': project_budget >= 40000,
                'indirect_rate_valid': rd_data.indirect_cost_rate in [0.0, 0.07],
                'percentages_sum_100': sum(p.percentage for p in rd_data.partner_breakdown) <= 100.1
            }
        }
        
        return statistics
    
    def print_results(self, results: Dict[str, Any]):
        """Print comprehensive results"""
        
        print("\n" + "="*80)
        print("🚀 R&D EXPENDITURE EXCEL FORM FILLING RESULTS (STREAMING)")
        print("="*80)
        
        # Process Status
        status_icon = "✅ SUCCESS" if results['success'] else "❌ FAILED"
        print(f"\n🔄 PROCESS STATUS: {status_icon}")
        print(f"📋 Steps Completed: {', '.join(results['steps_completed'])}")
        
        if results['errors']:
            print(f"⚠️  Errors: {'; '.join(results['errors'])}")
        
        # Statistics
        if results['rd_data'] and results['statistics']:
            stats = results['statistics']
            
            print(f"\n📊 TABLE 1: R&D ACTIVITIES BY IMPACT")
            table1 = stats['table1_impacts']
            print(f"   Total Entries: {table1['total_entries']}")
            print(f"   Direct Costs: €{table1['total_direct_costs']:,.2f}")
            print(f"   Indirect Costs: €{table1['total_indirect_costs']:,.2f}")
            print(f"   Funding Requested: €{table1['total_funding_requested']:,.2f}")
            
            print(f"\n💰 TABLE 2: R&D EXPENDITURE CATEGORIES")
            table2 = stats['table2_categories']
            print(f"   Direct Costs: €{table2['total_direct_costs']:,.2f}")
            print(f"   Indirect Rate: {table2['indirect_cost_rate']:.1f}%")
            print(f"   Indirect Amount: €{table2['indirect_amount']:,.2f}")
            print(f"   Project Budget: €{table2['project_budget']:,.2f}")
            print(f"   Funding Efficiency: {table2['funding_efficiency']:.1f}%")
            
            print(f"\n🤝 TABLE 3: PARTNER BREAKDOWN")
            table3 = stats['table3_partners']
            print(f"   Active Partners: {table3['active_partners']}")
            print(f"   Total Eligible Costs: €{table3['total_eligible_costs']:,.2f}")
            print(f"   Total Funding Requested: €{table3['total_funding_requested']:,.2f}")
            
            for partner in table3['partner_distribution']:
                print(f"   • {partner['name']}: €{partner['eligible_costs']:,.2f} ({partner['percentage']:.1f}%)")
            
            print(f"\n🔢 SPECIAL CALCULATIONS")
            special = stats['special_calculations']
            print(f"   Lines 4&5 Amount: €{special['lines_4_5_amount']:,.2f} ({special['lines_4_5_percentage']:.1f}%)")
            print(f"   Heading 8 Amount: €{special['heading_8_amount']:,.2f} ({special['heading_8_percentage']:.1f}%)")
            
            print(f"\n✅ COMPLIANCE CHECK")
            compliance = stats['compliance_check']
            min_funding = "✅ Met" if compliance['minimum_funding_met'] else "❌ Not Met"
            print(f"   Minimum Funding (€40K): {min_funding}")
            print(f"   Indirect Rate Valid: {'✅ Valid' if compliance['indirect_rate_valid'] else '❌ Invalid'}")
            print(f"   Partner Percentages: {'✅ Valid' if compliance['percentages_sum_100'] else '❌ >100%'}")
        
        print("\n" + "="*80)

def mainsummary(data_file_path):
    """Main function to run the R&D expenditure agent"""
    
    print("🚀 R&D EXPENDITURE EXCEL FORM FILLER (GROQ STREAMING)")
    print("="*65)
    
    # Configuration
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
    if not GROQ_API_KEY:
        print("❌ Error: GROQ_API_KEY environment variable not set")
        print("Please set your Groq API key in .env file: GROQ_API_KEY='your_api_key_here'")
        return
    
    
    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    template_path = "code/agents/ENGLISH_1A_priedas_InoStartas_en/Summary/inputSummary.xlsx"  # Your Excel template
    output_path = "code/agents/ENGLISH_1A_priedas_InoStartas_en/Summary/Summary.xlsx"
    
    # Validate required files exist
    required_files = [data_file_path, template_path]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print(f"❌ Missing required files: {', '.join(missing_files)}")
        print("Required files:")
        print(f"  • {data_file_path} - Text file with R&D expenditure data")
        print(f"  • {template_path} - Excel template file (R&D Budget Form)")
        return
    
    try:
        # Initialize agent
        print(f"\n🔧 Initializing R&D Expenditure Agent with Groq Streaming...")
        agent = RDExpenditureAgent(GROQ_API_KEY)
        
        # Process R&D expenditure
        print(f"\n🔄 Processing R&D expenditure from {data_file_path}...")
        results = agent.process_rd_expenditure(data_file_path, template_path, output_path)
        
        # Print results
        agent.print_results(results)
        
        if results['success']:
            print(f"\n✅ R&D expenditure Excel form successfully filled: {output_path}")
            print(f"\n🎯 Key Features Implemented:")
            print(f"   • ✅ Groq streaming API with meta-llama/llama-4-scout-17b-16e-instruct")
            print(f"   • ✅ Table 1: R&D activities by impact (A5:F15)")
            print(f"   • ✅ Table 2: 8 expenditure categories (C19:D26)")
            print(f"   • ✅ Table 3: Partner breakdown (C41:D50)")
            print(f"   • ✅ Automatic calculations (totals, percentages, indirect costs)")
            print(f"   • ✅ Compliance checking (€40K minimum, rates validation)")
            print(f"   • ✅ Complete R&D budget workflow")
            
            if results['statistics']:
                budget = results['statistics']['table2_categories']['project_budget']
                funding = results['statistics']['table2_categories']['total_funding_requested']
                print(f"\n📊 Project Overview:")
                print(f"   • Total Project Budget: €{budget:,.2f}")
                print(f"   • EU Funding Requested: €{funding:,.2f}")
                print(f"   • Funding Rate: {(funding/budget*100):,.1f}%")
        else:
            print(f"\n❌ Process failed. Check the logs above for details.")
            
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        print(f"\n❌ Fatal error: {e}")

