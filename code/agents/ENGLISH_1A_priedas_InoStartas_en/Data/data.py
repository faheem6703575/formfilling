import json
import logging
from typing import Dict, List, Optional, Any
from pathlib import Path
import openpyxl
from openpyxl import load_workbook
from groq import Groq
import re
from dotenv import load_dotenv

load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class ExcelFormFillerAgent:
    """
    AI Agent that uses Groq LLM to extract structured data from text files
    and automatically fills Excel forms for EU patent services.
    """
    
    def __init__(self, groq_api_key: Optional[str] = None):
        """Initialize the agent with Groq client."""
        self.client = Groq(api_key=groq_api_key) if groq_api_key else Groq()
        self.model = "meta-llama/llama-4-scout-17b-16e-instruct"  # Using available model
        
        # Define expected contractor types and legal entities
        self.contractor_types = [
            "Applicant", "Partner No 1", "Partner No 2", "Partner No 3",
            "Patent Attorney", "Legal Consultant", "Translation Service",
            "Technical Writer", "IP Consultant", "Research Institution",
            "Law Firm", "Patent Office", "External Consultant"
        ]
        
        self.legal_entities = [
            "Public Institution", "Private Company", "University",
            "Research Institute", "Law Firm", "Consulting Firm",
            "Government Agency", "Non-Profit Organization", "Partnership",
            "Limited Liability Company", "Corporation", "Sole Proprietorship"
        ]

    def read_data_file(self, file_path: str) -> str:
        """Read content from data.txt file."""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read().strip()
                logger.info(f"Successfully read {len(content)} characters from {file_path}")
                return content
        except FileNotFoundError:
            logger.error(f"File {file_path} not found")
            raise
        except Exception as e:
            logger.error(f"Error reading file {file_path}: {str(e)}")
            raise

    def create_analysis_prompt(self, prompt_path: str = "code/prompts/data.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")

    # def extract_data_with_llm(self, text_content: str) -> Dict[str, Any]:
    #     """Use Groq LLM to extract structured data from text content."""
    #     try:
    #         prompt = self.create_analysis_prompt(text_content)
            
    #         completion = self.client.chat.completions.create(
    #             model=self.model,
    #             messages=[
    #                 {
    #                     "role": "system",
    #                     "content": "You are a precise data extraction specialist. Return only valid JSON as requested."
    #                 },
    #                 {
    #                     "role": "user",
    #                     "content": prompt
    #                 }
    #             ],
    #             temperature=0.1,  # Low temperature for consistent extraction
    #             max_tokens=2048,
    #             top_p=0.9,
    #             stream=False,
    #             stop=None,
    #         )
            
    #         response_content = completion.choices[0].message.content.strip()
    #         logger.info(f"LLM Response: {response_content[:200]}...")
            
    #         # Clean the response to ensure it's valid JSON
    #         response_content = self._clean_json_response(response_content)
            
    #         # Parse JSON response
    #         extracted_data = json.loads(response_content)
    #         logger.info(f"Successfully extracted data for {len(extracted_data.get('contractors', []))} contractors")
            
    #         return extracted_data
            
    #     except json.JSONDecodeError as e:
    #         logger.error(f"Failed to parse JSON response: {str(e)}")
    #         logger.error(f"Response content: {response_content}")
    #         raise
    #     except Exception as e:
    #         logger.error(f"Error in LLM extraction: {str(e)}")
    #         raise
    def extract_data_with_llm(self, text_content: str) -> Dict[str, Any]:
        try:
            prompt_template = self.create_analysis_prompt()  # FIXED: No argument passed
            prompt = prompt_template + "\n\n" + text_content  # or use replace("{{INPUT_TEXT}}", text_content)
            
            completion = self.client.chat.completions.create(
                model=self.model,
                messages=[
                    {
                        "role": "system",
                        "content": "You are a precise data extraction specialist. Return only valid JSON as requested."
                    },
                    {
                        "role": "user",
                        "content": prompt
                    }
                ],
                temperature=0.3,
                stream=False  # Fixed: Changed to False for non-streaming
            )

            response_content = completion.choices[0].message.content.strip()
            logger.info(f"LLM Response: {response_content[:200]}...")
            
            response_content = self._clean_json_response(response_content)
            extracted_data = json.loads(response_content)

            logger.info(f"Successfully extracted data for {len(extracted_data.get('contractors', []))} contractors")
            return extracted_data
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON response: {str(e)}")
            logger.error(f"Response content: {response_content}")
            raise
        except Exception as e:
            logger.error(f"Error in LLM extraction: {str(e)}")
            raise
    def _clean_json_response(self, response: str) -> str:
        """Clean and validate JSON response from LLM."""
        # Remove any markdown code blocks
        response = re.sub(r'```json\s*', '', response)
        response = re.sub(r'```\s*$', '', response)
        
        # Find JSON object boundaries
        start_idx = response.find('{')
        end_idx = response.rfind('}') + 1
        
        if start_idx != -1 and end_idx != 0:
            response = response[start_idx:end_idx]
        
        return response.strip()

    def validate_extracted_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """Validate and clean extracted data."""
        if "contractors" not in data:
            logger.warning("No contractors found in extracted data")
            return {"contractors": []}
        
        validated_contractors = []
        
        for contractor in data["contractors"]:
            # Validate required fields
            contractor_type = contractor.get("type_of_contractor", "Unknown")
            legal_entity = contractor.get("legal_entity", "Unknown")
            eligible_costs = float(contractor.get("eligible_costs", 0.0))
            funding_requested = float(contractor.get("funding_requested", 0.0))
            
            # Ensure funding_requested doesn't exceed eligible_costs
            if funding_requested > eligible_costs and eligible_costs > 0:
                funding_requested = eligible_costs * 0.85  # Typical 85% funding rate
            
            validated_contractor = {
                "type_of_contractor": contractor_type,
                "legal_entity": legal_entity,
                "eligible_costs": eligible_costs,
                "funding_requested": funding_requested
            }
            
            validated_contractors.append(validated_contractor)
            logger.info(f"Validated contractor: {contractor_type} - €{eligible_costs}")
        
        return {"contractors": validated_contractors}

    def fill_excel_form(self, excel_path: str, extracted_data: Dict[str, Any], output_path: str = None) -> str:
        """Fill the Excel form with extracted data."""
        try:
            # Load the workbook
            workbook = load_workbook(excel_path)
            
            # Access the DATA sheet
            if "DATA" in workbook.sheetnames:
                sheet = workbook["DATA"]
            else:
                # If DATA sheet doesn't exist, use the first sheet
                sheet = workbook.active
                logger.warning("DATA sheet not found, using active sheet")
            
            contractors = extracted_data.get("contractors", [])
            
            # Fill main contractor data (rows 3-6: Applicant, Partner No 1-3)
            main_contractors = ["Applicant", "Partner No 1", "Partner No 2", "Partner No 3"]
            
            for i, contractor_type in enumerate(main_contractors):
                row = i + 3  # Starting from row 3
                
                # Find matching contractor
                matching_contractor = None
                for contractor in contractors:
                    if contractor["type_of_contractor"] == contractor_type:
                        matching_contractor = contractor
                        break
                
                if matching_contractor:
                    # Fill eligible costs (column C)
                    sheet[f'C{row}'] = matching_contractor["eligible_costs"]
                    # Fill funding requested (column D)
                    sheet[f'D{row}'] = matching_contractor["funding_requested"]
                    # Fill legal entity reference (column H for VLOOKUP)
                    sheet[f'H{row}'] = matching_contractor["type_of_contractor"]
                    
                    logger.info(f"Filled row {row} for {contractor_type}")
                else:
                    # Fill with zeros if no data found
                    sheet[f'C{row}'] = 0.00
                    sheet[f'D{row}'] = 0.00
                    sheet[f'H{row}'] = contractor_type
            
            # Fill lookup table (rows 9-18) with contractor types and legal entities
            lookup_start_row = 9
            for i, contractor in enumerate(contractors[:10]):  # Maximum 10 entries
                row = lookup_start_row + i
                
                # Fill contractor type (column B)
                sheet[f'B{row}'] = contractor["type_of_contractor"]
                # Fill eligible costs (column C)
                sheet[f'C{row}'] = contractor["eligible_costs"]
                # Fill funding requested (column D)
                sheet[f'D{row}'] = contractor["funding_requested"]
                # Fill contractor type for lookup (column H)
                sheet[f'H{row}'] = contractor["type_of_contractor"]
                # Fill legal entity (column I)
                sheet[f'I{row}'] = contractor["legal_entity"]
            
            # Set output path
            if output_path is None:
                output_path = excel_path.replace('.xlsx', '_filled.xlsx')
            
            # Save the workbook
            workbook.save(output_path)
            logger.info(f"Successfully saved filled Excel form to {output_path}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error filling Excel form: {str(e)}")
            raise

    def generate_summary_report(self, extracted_data: Dict[str, Any]) -> str:
        """Generate a summary report of the extraction and filling process."""
        contractors = extracted_data.get("contractors", [])
        
        report = "=" * 60 + "\n"
        report += "EXCEL FORM FILLING SUMMARY REPORT\n"
        report += "=" * 60 + "\n\n"
        
        report += f"Total Contractors Processed: {len(contractors)}\n\n"
        
        total_eligible = 0
        total_funding = 0
        
        for i, contractor in enumerate(contractors, 1):
            report += f"{i}. Contractor: {contractor['type_of_contractor']}\n"
            report += f"   Legal Entity: {contractor['legal_entity']}\n"
            report += f"   Eligible Costs: €{contractor['eligible_costs']:,.2f}\n"
            report += f"   Funding Requested: €{contractor['funding_requested']:,.2f}\n"
            report += f"   Funding Rate: {(contractor['funding_requested']/contractor['eligible_costs']*100) if contractor['eligible_costs'] > 0 else 0:.1f}%\n\n"
            
            total_eligible += contractor['eligible_costs']
            total_funding += contractor['funding_requested']
        
        report += "-" * 40 + "\n"
        report += f"TOTALS:\n"
        report += f"Total Eligible Costs: €{total_eligible:,.2f}\n"
        report += f"Total Funding Requested: €{total_funding:,.2f}\n"
        report += f"Overall Funding Rate: {(total_funding/total_eligible*100) if total_eligible > 0 else 0:.1f}%\n"
        report += "=" * 60 + "\n"
        
        return report

    def process_form(self, data_file_path: str, excel_file_path: str, output_file_path: str = None) -> Dict[str, Any]:
        """Main processing method that orchestrates the entire workflow."""
        try:
            logger.info("Starting Excel form filling process...")
            
            # Step 1: Read input data
            logger.info("Step 1: Reading input data file...")
            text_content = self.read_data_file(data_file_path)
            print(text_content)
            # Step 2: Extract data using LLM
            logger.info("Step 2: Extracting data using Groq LLM...")
            extracted_data = self.extract_data_with_llm(text_content)
            
            # Step 3: Validate extracted data
            logger.info("Step 3: Validating extracted data...")
            validated_data = self.validate_extracted_data(extracted_data)
            
            # Step 4: Fill Excel form
            logger.info("Step 4: Filling Excel form...")
            output_path = self.fill_excel_form(excel_file_path, validated_data, output_file_path)
            
            # Step 5: Generate summary report
            logger.info("Step 5: Generating summary report...")
            summary_report = self.generate_summary_report(validated_data)
            
            logger.info("Process completed successfully!")
            
            return {
                "success": True,
                "output_file": output_path,
                "extracted_data": validated_data,
                "summary_report": summary_report,
                "total_contractors": len(validated_data.get("contractors", [])),
                "total_eligible_costs": sum(c["eligible_costs"] for c in validated_data.get("contractors", [])),
                "total_funding_requested": sum(c["funding_requested"] for c in validated_data.get("contractors", []))
            }
            
        except Exception as e:
            logger.error(f"Process failed: {str(e)}")
            return {
                "success": False,
                "error": str(e),
                "extracted_data": None,
                "summary_report": None
            }


def maindata(data_file_path):
    """Example usage of the ExcelFormFillerAgent."""
    
    # Initialize the agent
    # agent = ExcelFormFillerAgent(groq_api_key="your_groq_api_key_here")  # With API key
    agent = ExcelFormFillerAgent()  # Uses environment variable GROQ_API_KEY
    
    # Define file paths
    
    
    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    excel_file = "code/agents/ENGLISH_1A_priedas_InoStartas_en/Data/inputdata.xlsx"
    output_file = "code/agents/ENGLISH_1A_priedas_InoStartas_en/Data/DATA.xlsx"
    
    # Process the form
    result = agent.process_form(data_file_path, excel_file, output_file)
    
    # Display results
    if result["success"]:
        print("\n" + "="*60)
        print("EXCEL FORM FILLING COMPLETED SUCCESSFULLY!")
        print("="*60)
        print(f"Output file: {result['output_file']}")
        print(f"Total contractors processed: {result['total_contractors']}")
        print(f"Total eligible costs: €{result['total_eligible_costs']:,.2f}")
        print(f"Total funding requested: €{result['total_funding_requested']:,.2f}")
        print("\nSUMMARY REPORT:")
        print(result["summary_report"])
    else:
        print(f"\nProcess failed with error: {result['error']}")
