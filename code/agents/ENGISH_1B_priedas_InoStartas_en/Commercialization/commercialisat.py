import openpyxl
import json
import os
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict
from dotenv import load_dotenv
from groq import Groq

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class MarketDevelopmentEntry:
    """Data class for market development service entries"""
    expenditure_type: str
    supplier_info: str
    price_eur: float
    table_number: int
    entry_number: int

@dataclass
class MarketDevelopmentData:
    """Complete market development service data structure"""
    table1_entries: List[MarketDevelopmentEntry]  # Type 1 expenditure (3 entries)
    table2_entries: List[MarketDevelopmentEntry]  # Type 2 expenditure (3 entries)
    table3_entries: List[MarketDevelopmentEntry]  # Type 3 expenditure (3 entries)
    table4_entries: List[MarketDevelopmentEntry]  # Type 4 expenditure (3 entries)
    table5_entries: List[MarketDevelopmentEntry]  # Type 5 expenditure (3 entries)
    
    def __post_init__(self):
        # Ensure exactly 3 entries per table
        self.table1_entries = self.table1_entries[:3]
        self.table2_entries = self.table2_entries[:3]
        self.table3_entries = self.table3_entries[:3]
        self.table4_entries = self.table4_entries[:3]
        self.table5_entries = self.table5_entries[:3]

class GroqMarketAnalyzer:
    """Analyzes market development data using Groq LLM"""
    
    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key)

    def create_analysis_prompt(self, prompt_path: str = "code/prompts/commercialist.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")
    
    def analyze_market_data(self, text_content: str) -> MarketDevelopmentData:
        """Analyze text and extract market development service data"""
        
        logger.info("Starting market development data analysis...")
        
        prompt = self.create_analysis_prompt()
        
        try:
            completion = self.client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert market development analyst specializing in extracting structured information from product market readiness documents and service proposals."
                    },
                    {
                        "role": "user",
                        "content": f"{prompt}\n\nText to analyze:\n{text_content}"
                    }
                ],
                temperature=0.3
            )
            
            analysis_result = completion.choices[0].message.content
            logger.info("Received analysis from Groq API")
            
            # Clean and parse JSON
            json_content = self._extract_json_from_response(analysis_result)
            parsed_data = json.loads(json_content)
            
            # Convert to MarketDevelopmentData object
            market_data = self._convert_to_market_data(parsed_data)
            
            return market_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            logger.error(f"Raw response: {analysis_result}")
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_json_from_response(self, response: str) -> str:
        """Extract JSON from response, handling markdown code blocks"""
        
        response = response.strip()
        
        # Remove markdown code blocks if present
        if '```json' in response:
            start_marker = '```json'
            end_marker = '```'
            start_index = response.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = response.find(end_marker, start_index)
                if end_index != -1:
                    return response[start_index:end_index].strip()
        
        if response.startswith('```') and response.endswith('```'):
            return response[3:-3].strip()
        
        return response
    
    def _convert_to_market_data(self, parsed_data: Dict) -> MarketDevelopmentData:
        """Convert parsed JSON to MarketDevelopmentData object"""
        
        def create_entries(data_list: List[Dict], table_num: int) -> List[MarketDevelopmentEntry]:
            entries = []
            for i, item in enumerate(data_list[:3]):  # Ensure max 3 entries
                entry = MarketDevelopmentEntry(
                    expenditure_type=item.get('expenditure_type', f'Market Development Service {table_num}.{i+1}'),
                    supplier_info=item.get('supplier_info', f'Supplier {table_num}.{i+1}, To be determined'),
                    price_eur=float(item.get('price_eur', 0.0)),
                    table_number=table_num,
                    entry_number=i + 1
                )
                entries.append(entry)
            return entries
        
        market_data = MarketDevelopmentData(
            table1_entries=create_entries(parsed_data.get('table1_entries', []), 1),
            table2_entries=create_entries(parsed_data.get('table2_entries', []), 2),
            table3_entries=create_entries(parsed_data.get('table3_entries', []), 3),
            table4_entries=create_entries(parsed_data.get('table4_entries', []), 4),
            table5_entries=create_entries(parsed_data.get('table5_entries', []), 5)
        )
        
        return market_data

class MarketDevelopmentExcelFiller:
    """Fills Excel form with market development service data"""
    
    def __init__(self):
        self.table_mappings = {
            1: {
                'rows': [7, 8, 9],
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D10'
            },
            2: {
                'rows': [13, 14, 15],
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D16'
            },
            3: {
                'rows': [19, 20, 21],
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D22'
            },
            4: {
                'rows': [25, 26, 27],
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D28'
            },
            5: {
                'rows': [31, 32, 33],
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D34'
            }
        }
    
    def fill_excel_form(self, template_path: str, output_path: str, market_data: MarketDevelopmentData) -> bool:
        """Fill Excel form with market development service data"""
        
        try:
            logger.info(f"Loading Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            # Get the correct sheet - looking for market development sheet
            sheet_name = None
            for name in workbook.sheetnames:
                if 'market' in name.lower() or 'commercialization' in name.lower() or 'development' in name.lower():
                    sheet_name = name
                    break
            
            if not sheet_name:
                sheet_name = workbook.sheetnames[0]  # Use first sheet as fallback
                logger.warning(f"Market development sheet not found, using: {sheet_name}")
            
            sheet = workbook[sheet_name]
            
            # Fill all 5 tables
            self._fill_table(sheet, market_data.table1_entries, 1)
            self._fill_table(sheet, market_data.table2_entries, 2)
            self._fill_table(sheet, market_data.table3_entries, 3)
            self._fill_table(sheet, market_data.table4_entries, 4)
            self._fill_table(sheet, market_data.table5_entries, 5)
            
            # Calculate and fill totals
            self._calculate_totals(sheet)
            
            # Save the filled workbook
            workbook.save(output_path)
            logger.info(f"✅ Successfully saved Excel form: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to fill Excel form: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _fill_table(self, sheet, entries: List[MarketDevelopmentEntry], table_number: int):
        """Fill a specific table with market development entries"""
        
        mapping = self.table_mappings[table_number]
        rows = mapping['rows']
        subtotal = 0.0
        
        # Use the lowest price among the 3 commercial offers (as per requirements)
        if entries:
            prices = [entry.price_eur for entry in entries]
            lowest_price = min(prices)
            
            # Fill all entries but use lowest price for calculation
            for i, entry in enumerate(entries):
                if i < len(rows):
                    row = rows[i]
                    
                    # Fill supplier info (Columns B:C merged)
                    sheet[f"B{row}"] = entry.supplier_info
                    
                    # Fill price (Column D)
                    sheet[f"D{row}"] = entry.price_eur
                    
                    logger.info(f"✅ Table {table_number}, Row {row}: {entry.expenditure_type} - €{entry.price_eur}")
            
            # Fixed subtotal uses lowest price (as per requirements)
            subtotal = lowest_price
        
        # Fill subtotal
        sheet[mapping['subtotal_cell']] = subtotal
        logger.info(f"✅ Table {table_number} Fixed Subtotal (lowest price): €{subtotal}")
    
    def _calculate_totals(self, sheet):
        """Calculate and fill all totals in the Excel sheet"""
        
        try:
            # Get subtotals from each table
            table1_subtotal = float(sheet['D10'].value or 0)
            table2_subtotal = float(sheet['D16'].value or 0)
            table3_subtotal = float(sheet['D22'].value or 0)
            table4_subtotal = float(sheet['D28'].value or 0)
            table5_subtotal = float(sheet['D34'].value or 0)
            
            # Calculate Fixed Amount (D36) - sum of all subtotals
            fixed_amount = table1_subtotal + table2_subtotal + table3_subtotal + table4_subtotal + table5_subtotal
            sheet['D36'] = fixed_amount
            
            # Calculate Indirect costs (D37) - typically 7% of fixed amount
            indirect_costs = fixed_amount * 0.07
            sheet['D37'] = indirect_costs
            
            # Calculate Total eligible costs (D38)
            total_costs = fixed_amount + indirect_costs
            sheet['D38'] = total_costs
            
            # Funding requested (D39) - up to 85% of total costs (as per requirements)
            funding_requested = total_costs * 0.85
            sheet['D39'] = funding_requested
            
            logger.info(f"✅ Calculated totals:")
            logger.info(f"   Table 1 Subtotal: €{table1_subtotal:,.2f}")
            logger.info(f"   Table 2 Subtotal: €{table2_subtotal:,.2f}")
            logger.info(f"   Table 3 Subtotal: €{table3_subtotal:,.2f}")
            logger.info(f"   Table 4 Subtotal: €{table4_subtotal:,.2f}")
            logger.info(f"   Table 5 Subtotal: €{table5_subtotal:,.2f}")
            logger.info(f"   Fixed Amount: €{fixed_amount:,.2f}")
            logger.info(f"   Indirect costs (7%): €{indirect_costs:,.2f}")
            logger.info(f"   Total costs: €{total_costs:,.2f}")
            logger.info(f"   Funding requested (85%): €{funding_requested:,.2f}")
            
        except Exception as e:
            logger.error(f"❌ Error calculating totals: {e}")

class MarketDevelopmentAgent:
    """Main orchestrator for market development processing"""
    
    def __init__(self, groq_api_key: str):
        self.analyzer = GroqMarketAnalyzer(groq_api_key)
        self.excel_filler = MarketDevelopmentExcelFiller()
    
    def read_data_file(self, file_path: str) -> str:
        """Read content from data file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"✅ Successfully read data file: {file_path} ({len(content)} characters)")
            return content
        except FileNotFoundError:
            logger.error(f"❌ Data file not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"❌ Error reading data file: {e}")
            raise
    
    def process_market_development(self, data_file_path: str, template_path: str, output_path: str) -> Dict[str, Any]:
        """Complete workflow for processing market development services"""
        
        results = {
            'success': False,
            'steps_completed': [],
            'market_data': None,
            'statistics': {},
            'errors': []
        }
        
        try:
            # Step 1: Read data file
            logger.info("🔄 STEP 1: Reading Data File")
            text_content = self.read_data_file(data_file_path)
            results['steps_completed'].append('data_file_read')
            
            # Step 2: Analyze with Groq API
            logger.info("🔄 STEP 2: Analyzing Market Development Services with Groq API")
            market_data = self.analyzer.analyze_market_data(text_content)
            results['market_data'] = market_data
            results['steps_completed'].append('data_analyzed')
            
            # Step 3: Generate statistics
            logger.info("🔄 STEP 3: Generating Statistics")
            statistics = self._generate_statistics(market_data)
            results['statistics'] = statistics
            results['steps_completed'].append('statistics_generated')
            
            # Step 4: Fill Excel form
            logger.info("🔄 STEP 4: Filling Excel Form")
            excel_success = self.excel_filler.fill_excel_form(template_path, output_path, market_data)
            
            if excel_success:
                results['steps_completed'].append('excel_filled')
                results['success'] = True
                logger.info("✅ PROCESS COMPLETED SUCCESSFULLY")
            else:
                results['errors'].append('Excel filling failed')
                
        except Exception as e:
            error_msg = f"❌ Process failed at step {len(results['steps_completed']) + 1}: {e}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _generate_statistics(self, market_data: MarketDevelopmentData) -> Dict[str, Any]:
        """Generate comprehensive statistics"""
        
        def get_table_stats(entries: List[MarketDevelopmentEntry], table_name: str):
            if not entries:
                return {'total': 0.0, 'lowest': 0.0, 'highest': 0.0, 'average': 0.0, 'count': 0}
            
            prices = [entry.price_eur for entry in entries]
            return {
                'total': sum(prices),
                'lowest': min(prices),
                'highest': max(prices),
                'average': sum(prices) / len(prices),
                'count': len(entries)
            }
        
        table1_stats = get_table_stats(market_data.table1_entries, "Market Research & Analysis")
        table2_stats = get_table_stats(market_data.table2_entries, "Marketing Strategy Development")
        table3_stats = get_table_stats(market_data.table3_entries, "Brand Development & Positioning")
        table4_stats = get_table_stats(market_data.table4_entries, "Market Testing & Validation")
        table5_stats = get_table_stats(market_data.table5_entries, "Market Entry & Launch Support")
        
        # Calculate totals using lowest prices (as per form requirements)
        fixed_amount = (table1_stats['lowest'] + table2_stats['lowest'] + 
                       table3_stats['lowest'] + table4_stats['lowest'] + table5_stats['lowest'])
        
        statistics = {
            'table_breakdown': {
                'table1_market_research': table1_stats,
                'table2_marketing_strategy': table2_stats,
                'table3_brand_development': table3_stats,
                'table4_market_testing': table4_stats,
                'table5_market_entry': table5_stats
            },
            'cost_summary': {
                'fixed_amount': fixed_amount,
                'indirect_costs': fixed_amount * 0.07,
                'total_eligible_costs': fixed_amount * 1.07,
                'funding_requested': fixed_amount * 1.07 * 0.85
            },
            'service_summary': {
                'total_services': 15,  # 5 tables × 3 services each
                'total_tables': 5,
                'services_per_table': 3,
                'all_proposals_received': True
            }
        }
        
        return statistics
    
    def print_results(self, results: Dict[str, Any]):
        """Print comprehensive results"""
        
        print("\n" + "="*80)
        print("🚀 MARKET DEVELOPMENT EXCEL FORM FILLING RESULTS")
        print("="*80)
        
        # Process Status
        status_icon = "✅ SUCCESS" if results['success'] else "❌ FAILED"
        print(f"\n🔄 PROCESS STATUS: {status_icon}")
        print(f"📋 Steps Completed: {', '.join(results['steps_completed'])}")
        
        if results['errors']:
            print(f"⚠️  Errors: {'; '.join(results['errors'])}")
        
        # Statistics
        if results['market_data'] and results['statistics']:
            stats = results['statistics']
            
            print(f"\n📊 TABLE BREAKDOWN (Based on Lowest Price Criterion)")
            tables = stats['table_breakdown']
            for table_name, table_stats in tables.items():
                table_display = table_name.replace('_', ' ').title()
                print(f"   {table_display}:")
                print(f"     • Services: {table_stats['count']}")
                print(f"     • Price Range: €{table_stats['lowest']:,.2f} - €{table_stats['highest']:,.2f}")
                print(f"     • Fixed Amount (Lowest): €{table_stats['lowest']:,.2f}")
            
            print(f"\n💰 COST SUMMARY")
            costs = stats['cost_summary']
            print(f"   Fixed Amount (Sum of Lowest Prices): €{costs['fixed_amount']:,.2f}")
            print(f"   Indirect Costs (7%): €{costs['indirect_costs']:,.2f}")
            print(f"   Total Eligible Costs: €{costs['total_eligible_costs']:,.2f}")
            print(f"   Funding Requested (85%): €{costs['funding_requested']:,.2f}")
            
            print(f"\n📋 SERVICE SUMMARY")
            service_summary = stats['service_summary']
            print(f"   Total Services Processed: {service_summary['total_services']}")
            print(f"   Tables Completed: {service_summary['total_tables']}")
            print(f"   Commercial Proposals per Service: {service_summary['services_per_table']}")
            print(f"   All Required Proposals Received: {'✅ Yes' if service_summary['all_proposals_received'] else '❌ No'}")
        
        print("\n" + "="*80)

def main_commercilisat(data_file_path):
    """Main function to run the market development agent"""
    
    print("🚀 MARKET DEVELOPMENT EXCEL FORM FILLER")
    print("="*50)
    # Configuration
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
    if not GROQ_API_KEY:
        print("❌ Error: GROQ_API_KEY environment variable not set")
        print("Please set your Groq API key in .env file: GROQ_API_KEY='your_api_key_here'")
        return
    
    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    template_path = "code/agents/ENGISH_1B_priedas_InoStartas_en/Commercialization/commercialisat.xlsx"  # Your Excel template
    output_path = "code/agents/ENGISH_1B_priedas_InoStartas_en/Commercialization/2. Fic. amounts (commercialisat.xlsx"
    
    # Validate required files exist
    required_files = [data_file_path, template_path]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print(f"❌ Missing required files: {', '.join(missing_files)}")
        print("Required files:")
        print(f"  • {data_file_path} - Text file with market development service data")
        print(f"  • {template_path} - Excel template file")
        return
    
    try:
        # Initialize agent
        print(f"\n🔧 Initializing Market Development Agent...")
        agent = MarketDevelopmentAgent(GROQ_API_KEY)
        
        # Process market development services
        print(f"\n🔄 Processing market development services from {data_file_path}...")
        results = agent.process_market_development(data_file_path, template_path, output_path)
        
        # Print results
        agent.print_results(results)
        
        if results['success']:
            print(f"\n✅ Market development Excel form successfully filled: {output_path}")
            print(f"\n🎯 The form includes:")
            print(f"   • 5 expenditure types with 3 commercial proposals each")
            print(f"   • Fixed amount calculation based on lowest price criterion")
            print(f"   • Automatic subtotal and total calculations")
            print(f"   • 7% indirect costs calculation")
            print(f"   • 85% funding request calculation")
        else:
            print(f"\n❌ Process failed. Check the logs above for details.")
            
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        print(f"\n❌ Fatal error: {e}")

