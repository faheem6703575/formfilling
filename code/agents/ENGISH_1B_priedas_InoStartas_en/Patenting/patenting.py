import openpyxl
import json
import os
import logging
from typing import Dict, List, Any, Optional
from dataclasses import dataclass, asdict
from dotenv import load_dotenv
from groq import Groq

# Load environment variables
load_dotenv()

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

@dataclass
class PatentServiceEntry:
    """Data class for patent service entries"""
    description: str
    supplier_info: str
    price_eur: float
    category: str  # 'phase1_attorney', 'phase1_taxes', 'phase2_attorney', 'phase2_taxes'

@dataclass
class PatentData:
    """Complete patent service data structure"""
    phase1_attorney_services: List[PatentServiceEntry]
    phase1_taxes: List[PatentServiceEntry]
    phase2_attorney_services: List[PatentServiceEntry]
    phase2_taxes: List[PatentServiceEntry]
    
    def __post_init__(self):
        # Ensure we have exactly the right number of entries for each category
        self.phase1_attorney_services = self.phase1_attorney_services[:3]  # Max 3 entries
        self.phase1_taxes = self.phase1_taxes[:3]  # Max 3 entries
        self.phase2_attorney_services = self.phase2_attorney_services[:3]  # Max 3 entries
        self.phase2_taxes = self.phase2_taxes[:4]  # Max 4 entries

class GroqPatentAnalyzer:
    """Analyzes patent service data using Groq LLM"""
    
    def __init__(self, api_key: str):
        self.client = Groq(api_key=api_key)
        
    def create_analysis_prompt(self, prompt_path: str = "code/prompts/patenting.txt") -> str:
        """Read detailed prompt for market development service extraction from a file"""
        try:
            with open(prompt_path, 'r', encoding='utf-8') as file:
                prompt = file.read()
            return prompt
        except FileNotFoundError:
            raise FileNotFoundError(f"Prompt file not found at: {prompt_path}")
        except Exception as e:
            raise RuntimeError(f"Failed to read prompt file: {e}")
    
    def analyze_patent_data(self, text_content: str) -> PatentData:
        """Analyze text and extract patent service data"""
        
        logger.info("Starting patent service data analysis...")
        
        prompt = self.create_analysis_prompt()
        
        try:
            completion = self.client.chat.completions.create(
                model="meta-llama/llama-4-scout-17b-16e-instruct",
                messages=[
                    {
                        "role": "system",
                        "content": "You are an expert patent service data analyst specializing in extracting structured information from patent-related documents and invoices."
                    },
                    {
                        "role": "user",
                        "content": f"{prompt}\n\nText to analyze:\n{text_content}"
                    }
                ],
                temperature=0.3
            )
            
            analysis_result = completion.choices[0].message.content
            logger.info("Received analysis from Groq API")
            
            # Clean and parse JSON
            json_content = self._extract_json_from_response(analysis_result)
            parsed_data = json.loads(json_content)
            
            # Convert to PatentData object
            patent_data = self._convert_to_patent_data(parsed_data)
            
            return patent_data
            
        except json.JSONDecodeError as e:
            logger.error(f"Failed to parse JSON: {e}")
            logger.error(f"Raw response: {analysis_result}")
            raise
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise
    
    def _extract_json_from_response(self, response: str) -> str:
        """Extract JSON from response, handling markdown code blocks"""
        
        response = response.strip()
        
        # Remove markdown code blocks if present
        if '```json' in response:
            start_marker = '```json'
            end_marker = '```'
            start_index = response.find(start_marker)
            if start_index != -1:
                start_index += len(start_marker)
                end_index = response.find(end_marker, start_index)
                if end_index != -1:
                    return response[start_index:end_index].strip()
        
        if response.startswith('```') and response.endswith('```'):
            return response[3:-3].strip()
        
        return response
    
    def _convert_to_patent_data(self, parsed_data: Dict) -> PatentData:
        """Convert parsed JSON to PatentData object"""
        
        def create_entries(data_list: List[Dict], category: str) -> List[PatentServiceEntry]:
            entries = []
            for item in data_list:
                entry = PatentServiceEntry(
                    description=item.get('description', f'Patent service - {category}'),
                    supplier_info=item.get('supplier_info', 'To be determined'),
                    price_eur=float(item.get('price_eur', 0.0)),
                    category=category
                )
                entries.append(entry)
            return entries
        
        patent_data = PatentData(
            phase1_attorney_services=create_entries(
                parsed_data.get('phase1_attorney_services', []), 'phase1_attorney'
            ),
            phase1_taxes=create_entries(
                parsed_data.get('phase1_taxes', []), 'phase1_taxes'
            ),
            phase2_attorney_services=create_entries(
                parsed_data.get('phase2_attorney_services', []), 'phase2_attorney'
            ),
            phase2_taxes=create_entries(
                parsed_data.get('phase2_taxes', []), 'phase2_taxes'
            )
        )
        
        return patent_data

class PatentExcelFiller:
    """Fills Excel form with patent service data"""
    
    def __init__(self):
        self.cell_mappings = {
            # Phase I Attorney Services (rows 7-9)
            'phase1_attorney': {
                'rows': [7, 8, 9],
                'description_cols': 'A',
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D10'
            },
            # Phase I Taxes (rows 13-15)
            'phase1_taxes': {
                'rows': [13, 14, 15],
                'description_cols': 'A',
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D16'
            },
            # Phase II Attorney Services (rows 22-24)
            'phase2_attorney': {
                'rows': [22, 23, 24],
                'description_cols': 'A',
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D25'
            },
            # Phase II Taxes (rows 28-31)
            'phase2_taxes': {
                'rows': [28, 29, 30, 31],
                'description_cols': 'A',
                'supplier_cols': 'B:C',
                'price_col': 'D',
                'subtotal_cell': 'D32'
            }
        }
    
    def fill_excel_form(self, template_path: str, output_path: str, patent_data: PatentData) -> bool:
        """Fill Excel form with patent service data"""
        
        try:
            logger.info(f"Loading Excel template: {template_path}")
            workbook = openpyxl.load_workbook(template_path)
            
            # Get the correct sheet - looking for patenting sheet
            sheet_name = None
            for name in workbook.sheetnames:
                if 'patent' in name.lower() or 'fictitious' in name.lower():
                    sheet_name = name
                    break
            
            if not sheet_name:
                sheet_name = workbook.sheetnames[0]  # Use first sheet as fallback
                logger.warning(f"Patent sheet not found, using: {sheet_name}")
            
            sheet = workbook[sheet_name]
            
            # Fill Phase I Attorney Services
            self._fill_service_section(
                sheet, 
                patent_data.phase1_attorney_services, 
                self.cell_mappings['phase1_attorney']
            )
            
            # Fill Phase I Taxes
            self._fill_service_section(
                sheet, 
                patent_data.phase1_taxes, 
                self.cell_mappings['phase1_taxes']
            )
            
            # Fill Phase II Attorney Services
            self._fill_service_section(
                sheet, 
                patent_data.phase2_attorney_services, 
                self.cell_mappings['phase2_attorney']
            )
            
            # Fill Phase II Taxes
            self._fill_service_section(
                sheet, 
                patent_data.phase2_taxes, 
                self.cell_mappings['phase2_taxes']
            )
            
            # Calculate and fill totals
            self._calculate_totals(sheet)
            
            # Save the filled workbook
            workbook.save(output_path)
            logger.info(f"✅ Successfully saved Excel form: {output_path}")
            
            return True
            
        except Exception as e:
            logger.error(f"❌ Failed to fill Excel form: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _fill_service_section(self, sheet, services: List[PatentServiceEntry], mapping: Dict):
        """Fill a section of services in the Excel sheet"""
        
        rows = mapping['rows']
        subtotal = 0.0
        
        for i, service in enumerate(services):
            if i < len(rows):
                row = rows[i]
                
                # Fill description (Column A)
                # sheet[f"A{row}"] = service.eil_no
                
                # Fill supplier info (Columns B:C merged)
                sheet[f"B{row}"] = service.supplier_info
                
                # Fill price (Column D)
                sheet[f"D{row}"] = service.price_eur
                subtotal += service.price_eur
                
                logger.info(f"✅ Filled row {row}: {service.description} - €{service.price_eur}")
        
        # Fill subtotal
        sheet[mapping['subtotal_cell']] = subtotal
        logger.info(f"✅ Subtotal for section: €{subtotal}")
    
    def _calculate_totals(self, sheet):
        """Calculate and fill all totals in the Excel sheet"""
        
        try:
            # Get subtotals
            phase1_attorney_subtotal = float(sheet['D10'].value or 0)
            phase1_taxes_subtotal = float(sheet['D16'].value or 0)
            phase2_attorney_subtotal = float(sheet['D25'].value or 0)
            phase2_taxes_subtotal = float(sheet['D32'].value or 0)
            
            # Calculate Fixed Amount I (D17)
            fixed_amount_1 = phase1_attorney_subtotal + phase1_taxes_subtotal
            sheet['D17'] = fixed_amount_1
            
            # Calculate Fixed Amount II (D34)
            fixed_amount_2 = phase2_attorney_subtotal + phase2_taxes_subtotal
            sheet['D34'] = fixed_amount_2
            
            # Calculate Direct eligible costs (D36)
            direct_costs = fixed_amount_1 + fixed_amount_2
            sheet['D36'] = direct_costs
            
            # Calculate Indirect costs (D37) - typically 7% of direct costs
            indirect_costs = direct_costs * 0.07
            sheet['D37'] = indirect_costs
            
            # Calculate Total eligible costs (D38)
            total_costs = direct_costs + indirect_costs
            sheet['D38'] = total_costs
            
            # Funding requested (D39) - typically 85% of total costs
            funding_requested = total_costs * 0.85
            sheet['D39'] = funding_requested
            
            logger.info(f"✅ Calculated totals:")
            logger.info(f"   Fixed Amount I: €{fixed_amount_1:,.2f}")
            logger.info(f"   Fixed Amount II: €{fixed_amount_2:,.2f}")
            logger.info(f"   Direct costs: €{direct_costs:,.2f}")
            logger.info(f"   Indirect costs: €{indirect_costs:,.2f}")
            logger.info(f"   Total costs: €{total_costs:,.2f}")
            logger.info(f"   Funding requested: €{funding_requested:,.2f}")
            
        except Exception as e:
            logger.error(f"❌ Error calculating totals: {e}")

class PatentServiceAgent:
    """Main orchestrator for patent service processing"""
    
    def __init__(self, groq_api_key: str):
        self.analyzer = GroqPatentAnalyzer(groq_api_key)
        self.excel_filler = PatentExcelFiller()
    
    def read_data_file(self, file_path: str) -> str:
        """Read content from data file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            logger.info(f"✅ Successfully read data file: {file_path} ({len(content)} characters)")
            return content
        except FileNotFoundError:
            logger.error(f"❌ Data file not found: {file_path}")
            raise
        except Exception as e:
            logger.error(f"❌ Error reading data file: {e}")
            raise
    
    def process_patent_services(self, data_file_path: str, template_path: str, output_path: str) -> Dict[str, Any]:
        """Complete workflow for processing patent services"""
        
        results = {
            'success': False,
            'steps_completed': [],
            'patent_data': None,
            'statistics': {},
            'errors': []
        }
        
        try:
            # Step 1: Read data file
            logger.info("🔄 STEP 1: Reading Data File")
            text_content = self.read_data_file(data_file_path)
            results['steps_completed'].append('data_file_read')
            
            # Step 2: Analyze with Groq API
            logger.info("🔄 STEP 2: Analyzing Patent Services with Groq API")
            patent_data = self.analyzer.analyze_patent_data(text_content)
            results['patent_data'] = patent_data
            results['steps_completed'].append('data_analyzed')
            
            # Step 3: Generate statistics
            logger.info("🔄 STEP 3: Generating Statistics")
            statistics = self._generate_statistics(patent_data)
            results['statistics'] = statistics
            results['steps_completed'].append('statistics_generated')
            
            # Step 4: Fill Excel form
            logger.info("🔄 STEP 4: Filling Excel Form")
            excel_success = self.excel_filler.fill_excel_form(template_path, output_path, patent_data)
            
            if excel_success:
                results['steps_completed'].append('excel_filled')
                results['success'] = True
                logger.info("✅ PROCESS COMPLETED SUCCESSFULLY")
            else:
                results['errors'].append('Excel filling failed')
                
        except Exception as e:
            error_msg = f"❌ Process failed at step {len(results['steps_completed']) + 1}: {e}"
            logger.error(error_msg)
            results['errors'].append(error_msg)
        
        return results
    
    def _generate_statistics(self, patent_data: PatentData) -> Dict[str, Any]:
        """Generate comprehensive statistics"""
        
        phase1_attorney_total = sum(s.price_eur for s in patent_data.phase1_attorney_services)
        phase1_taxes_total = sum(s.price_eur for s in patent_data.phase1_taxes)
        phase2_attorney_total = sum(s.price_eur for s in patent_data.phase2_attorney_services)
        phase2_taxes_total = sum(s.price_eur for s in patent_data.phase2_taxes)
        
        total_attorney_costs = phase1_attorney_total + phase2_attorney_total
        total_tax_costs = phase1_taxes_total + phase2_taxes_total
        grand_total = total_attorney_costs + total_tax_costs
        
        statistics = {
            'phase_breakdown': {
                'phase1_attorney': phase1_attorney_total,
                'phase1_taxes': phase1_taxes_total,
                'phase1_total': phase1_attorney_total + phase1_taxes_total,
                'phase2_attorney': phase2_attorney_total,
                'phase2_taxes': phase2_taxes_total,
                'phase2_total': phase2_attorney_total + phase2_taxes_total
            },
            'cost_summary': {
                'total_attorney_costs': total_attorney_costs,
                'total_tax_costs': total_tax_costs,
                'direct_costs': grand_total,
                'indirect_costs': grand_total * 0.07,
                'total_project_cost': grand_total * 1.07,
                'funding_requested': grand_total * 1.07 * 0.85
            },
            'service_counts': {
                'phase1_attorney_services': len(patent_data.phase1_attorney_services),
                'phase1_tax_services': len(patent_data.phase1_taxes),
                'phase2_attorney_services': len(patent_data.phase2_attorney_services),
                'phase2_tax_services': len(patent_data.phase2_taxes),
                'total_services': (len(patent_data.phase1_attorney_services) + 
                                 len(patent_data.phase1_taxes) + 
                                 len(patent_data.phase2_attorney_services) + 
                                 len(patent_data.phase2_taxes))
            }
        }
        
        return statistics
    
    def print_results(self, results: Dict[str, Any]):
        """Print comprehensive results"""
        
        print("\n" + "="*80)
        print("🚀 PATENT SERVICE EXCEL FORM FILLING RESULTS")
        print("="*80)
        
        # Process Status
        status_icon = "✅ SUCCESS" if results['success'] else "❌ FAILED"
        print(f"\n🔄 PROCESS STATUS: {status_icon}")
        print(f"📋 Steps Completed: {', '.join(results['steps_completed'])}")
        
        if results['errors']:
            print(f"⚠️  Errors: {'; '.join(results['errors'])}")
        
        # Statistics
        if results['patent_data'] and results['statistics']:
            stats = results['statistics']
            
            print(f"\n📊 PHASE BREAKDOWN")
            phase = stats['phase_breakdown']
            print(f"   Phase I Attorney: €{phase['phase1_attorney']:,.2f}")
            print(f"   Phase I Taxes: €{phase['phase1_taxes']:,.2f}")
            print(f"   Phase I Total: €{phase['phase1_total']:,.2f}")
            print(f"   Phase II Attorney: €{phase['phase2_attorney']:,.2f}")
            print(f"   Phase II Taxes: €{phase['phase2_taxes']:,.2f}")
            print(f"   Phase II Total: €{phase['phase2_total']:,.2f}")
            
            print(f"\n💰 COST SUMMARY")
            costs = stats['cost_summary']
            print(f"   Total Attorney Costs: €{costs['total_attorney_costs']:,.2f}")
            print(f"   Total Tax Costs: €{costs['total_tax_costs']:,.2f}")
            print(f"   Direct Costs: €{costs['direct_costs']:,.2f}")
            print(f"   Indirect Costs (7%): €{costs['indirect_costs']:,.2f}")
            print(f"   Total Project Cost: €{costs['total_project_cost']:,.2f}")
            print(f"   Funding Requested (85%): €{costs['funding_requested']:,.2f}")
            
            print(f"\n📋 SERVICE SUMMARY")
            services = stats['service_counts']
            print(f"   Total Services Processed: {services['total_services']}")
            print(f"   Phase I Services: {services['phase1_attorney_services']} attorney + {services['phase1_tax_services']} tax")
            print(f"   Phase II Services: {services['phase2_attorney_services']} attorney + {services['phase2_tax_services']} tax")
        
        print("\n" + "="*80)

def main_patenting(data_file_path):
    """Main function to run the patent service agent"""
    
    print("🚀 PATENT SERVICE EXCEL FORM FILLER")
    print("="*50)
    
    # Configuration
    GROQ_API_KEY = os.environ.get("GROQ_API_KEY")
    if not GROQ_API_KEY:
        print("❌ Error: GROQ_API_KEY environment variable not set")
        print("Please set your Groq API key in .env file: GROQ_API_KEY='your_api_key_here'")
        return

    
    # r"C:\Users\USER\Documents\eu excel form\excel form filling\mainInput\onefile.txt"
    template_path = "code/agents/ENGISH_1B_priedas_InoStartas_en/Patenting/patenting.xlsx"  # Your Excel template
    output_path = "code/agents/ENGISH_1B_priedas_InoStartas_en/Patenting/1. Fic. amounts (patenting) .xlsx"
    
    # Validate required files exist
    required_files = [data_file_path, template_path]
    missing_files = [f for f in required_files if not os.path.exists(f)]
    
    if missing_files:
        print(f"❌ Missing required files: {', '.join(missing_files)}")
        print("Required files:")
        print(f"  • {data_file_path} - Text file with patent service data")
        print(f"  • {template_path} - Excel template file")
        return
    
    try:
        # Initialize agent
        print(f"\n🔧 Initializing Patent Service Agent...")
        agent = PatentServiceAgent(GROQ_API_KEY)
        
        # Process patent services
        print(f"\n🔄 Processing patent services from {data_file_path}...")
        results = agent.process_patent_services(data_file_path, template_path, output_path)
        
        # Print results
        agent.print_results(results)
        
        if results['success']:
            print(f"\n✅ Patent service Excel form successfully filled: {output_path}")
            print(f"\n🎯 The form includes:")
            print(f"   • Phase I & II attorney services with costs")
            print(f"   • Tax and stamp duty calculations")
            print(f"   • Automatic subtotal and total calculations")
            print(f"   • Funding request calculations (85% of total)")
        else:
            print(f"\n❌ Process failed. Check the logs above for details.")
            
    except Exception as e:
        logger.error(f"❌ Fatal error: {e}")
        print(f"\n❌ Fatal error: {e}")
